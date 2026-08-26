"""Geração da planilha Excel do relatório de faturamento — réplica de
`_write_excel()` em `client/app_faturamento.py`, mesma estrutura de 4 abas e
mesma formatação, mas a partir dos dados já calculados pela API.

Sem pandas: o backend não depende de pandas (só os clients desktop
dependem), então as agregações das abas "Por Filial"/"Top Produtos" são
feitas com dicionários simples em vez de `DataFrame.groupby`. Em caso de
empate no total (critério de ordenação), a ordem de desempate pode divergir
da versão pandas — não afeta os valores, só a ordem de linhas empatadas.
"""
import io
from collections import OrderedDict
from datetime import datetime
from decimal import Decimal
from typing import Sequence

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.excel.styles import (
    ALT_FILL, BOLD10, BORDER, BRL, LEFT, PCT, PCT_CELL, QTD_FMT, RIGHT,
    cabecalho,
)
from app.schemas.faturamento import FaturamentoRead

_COLUNAS_ABA1 = [
    ("Filial", "filial", 9),
    ("Dia", "dia", 7),
    ("Código", "codigo", 14),
    ("Descrição", "descricao", 36),
    ("Quantidade", "quantidade", 14),
    ("Faturamento (R$)", "faturamento", 16),
    ("Custo (R$)", "custo", 16),
    ("Preço Médio (R$)", "preco_medio", 16),
    ("Lucro Bruto (R$)", "lucro_bruto", 16),
    ("Margem (%)", "margem", 12),
    ("Markup (%)", "markup", 12),
]


def gerar_excel_faturamento(linhas: Sequence[FaturamentoRead]) -> io.BytesIO:
    wb = openpyxl.Workbook()

    _aba_faturamento(wb.active, linhas)
    _aba_resumo(wb.create_sheet("Resumo"), linhas)
    _aba_por_filial(wb.create_sheet("Por Filial"), linhas)
    _aba_top_produtos(wb.create_sheet("Top Produtos"), linhas)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _aba_faturamento(ws, linhas: Sequence[FaturamentoRead]) -> None:
    ws.title = "Faturamento"
    ws.freeze_panes = "A2"

    for ci, (titulo, _, largura) in enumerate(_COLUNAS_ABA1, 1):
        cabecalho(ws, 1, ci, titulo)
        ws.column_dimensions[get_column_letter(ci)].width = largura
    ws.row_dimensions[1].height = 18
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUNAS_ABA1))}1"

    for ri, linha in enumerate(linhas, 2):
        preencher_linha = ALT_FILL if ri % 2 == 0 else None
        for ci, (_, campo, _largura) in enumerate(_COLUNAS_ABA1, 1):
            valor = getattr(linha, campo)
            # Decimal -> float: openpyxl grava Decimal como texto em vez de
            # número, o que quebraria number_format e qualquer fórmula/
            # ordenação feita depois na planilha pelo usuário.
            if isinstance(valor, Decimal):
                valor = float(valor)
            celula = ws.cell(row=ri, column=ci, value=valor)
            celula.border = BORDER
            if campo in ("faturamento", "custo", "preco_medio", "lucro_bruto"):
                celula.number_format = BRL
                celula.alignment = RIGHT
            elif campo == "quantidade":
                celula.number_format = QTD_FMT
                celula.alignment = RIGHT
            elif campo in ("margem", "markup"):
                celula.number_format = PCT_CELL
                celula.alignment = RIGHT
            else:
                celula.alignment = LEFT
            if preencher_linha:
                celula.fill = preencher_linha


def _aba_resumo(ws, linhas: Sequence[FaturamentoRead]) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    ws.cell(1, 1, "Resumo — Faturamento").font = Font(
        bold=True, color="1A5276", size=14, name="Calibri")
    ws.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(
        italic=True, size=9, color="7F8C8D", name="Calibri")

    faturamento_total = float(sum(l.faturamento for l in linhas))
    custo_total = float(sum(l.custo for l in linhas))
    lucro_total = float(sum(l.lucro_bruto for l in linhas))
    quantidade_total = float(sum(l.quantidade for l in linhas))
    margem_geral = (lucro_total / faturamento_total) if faturamento_total else 0.0
    markup_geral = (lucro_total / custo_total) if custo_total else 0.0

    itens_resumo = [
        ("Total de Registros", len(linhas), None),
        ("Produtos Distintos", len({l.codigo for l in linhas}), None),
        ("Filiais Distintas", len({l.filial for l in linhas}), None),
        ("Quantidade Total", quantidade_total, QTD_FMT),
        ("Faturamento Total (R$)", faturamento_total, BRL),
        ("Custo Total (R$)", custo_total, BRL),
        ("Lucro Bruto Total (R$)", lucro_total, BRL),
        ("Margem Geral (%)", margem_geral, PCT),
        ("Markup Geral (%)", markup_geral, PCT),
    ]
    for i, (rotulo, valor, formato) in enumerate(itens_resumo, 4):
        c1 = ws.cell(i, 1, rotulo)
        c1.font = Font(bold=True, name="Calibri", size=10)
        c2 = ws.cell(i, 2, valor)
        c2.alignment = RIGHT
        if formato:
            c2.number_format = formato


def _aba_por_filial(ws, linhas: Sequence[FaturamentoRead]) -> None:
    agregados: "OrderedDict[str, dict]" = OrderedDict()
    for l in linhas:
        agg = agregados.setdefault(
            l.filial, {"qtd_registros": 0, "faturamento_total": 0.0, "lucro_total": 0.0})
        agg["qtd_registros"] += 1
        agg["faturamento_total"] += float(l.faturamento)
        agg["lucro_total"] += float(l.lucro_bruto)

    ordenados = sorted(agregados.items(), key=lambda kv: kv[1]["faturamento_total"], reverse=True)
    total_geral = sum(v["faturamento_total"] for _, v in ordenados)

    cabecalhos = ["Filial", "Qtd. Registros", "Faturamento (R$)", "Lucro Bruto (R$)", "% do Total"]
    larguras = [12, 14, 18, 18, 12]
    for ci, (titulo, largura) in enumerate(zip(cabecalhos, larguras), 1):
        cabecalho(ws, 1, ci, titulo)
        ws.column_dimensions[get_column_letter(ci)].width = largura

    ri = 2
    for filial, agg in ordenados:
        pct = (agg["faturamento_total"] / total_geral) if total_geral else 0.0
        ws.cell(ri, 1, filial)
        ws.cell(ri, 2, agg["qtd_registros"])
        c3 = ws.cell(ri, 3, agg["faturamento_total"]); c3.number_format = BRL
        c4 = ws.cell(ri, 4, agg["lucro_total"]); c4.number_format = BRL
        c5 = ws.cell(ri, 5, pct); c5.number_format = PCT
        if ri % 2 == 0:
            for ci in range(1, 6):
                ws.cell(ri, ci).fill = ALT_FILL
        ri += 1

    ws.cell(ri, 1, "TOTAL GERAL").font = BOLD10
    ws.cell(ri, 2, sum(v["qtd_registros"] for _, v in ordenados)).font = BOLD10
    c3 = ws.cell(ri, 3, total_geral); c3.number_format = BRL; c3.font = BOLD10
    c4 = ws.cell(ri, 4, sum(v["lucro_total"] for _, v in ordenados)); c4.number_format = BRL; c4.font = BOLD10
    c5 = ws.cell(ri, 5, 1.0); c5.number_format = PCT; c5.font = BOLD10


def _aba_top_produtos(ws, linhas: Sequence[FaturamentoRead], limite: int = 50) -> None:
    agregados: "OrderedDict[tuple, dict]" = OrderedDict()
    for l in linhas:
        chave = (l.codigo, l.descricao)
        agg = agregados.setdefault(
            chave, {"quantidade_total": 0.0, "faturamento_total": 0.0, "lucro_total": 0.0})
        agg["quantidade_total"] += float(l.quantidade)
        agg["faturamento_total"] += float(l.faturamento)
        agg["lucro_total"] += float(l.lucro_bruto)

    top = sorted(agregados.items(), key=lambda kv: kv[1]["faturamento_total"], reverse=True)[:limite]

    cabecalhos = ["Código", "Descrição", "Quantidade Total", "Faturamento (R$)", "Lucro Bruto (R$)"]
    larguras = [14, 36, 18, 18, 18]
    for ci, (titulo, largura) in enumerate(zip(cabecalhos, larguras), 1):
        cabecalho(ws, 1, ci, titulo)
        ws.column_dimensions[get_column_letter(ci)].width = largura

    for ri, ((codigo, descricao), agg) in enumerate(top, 2):
        ws.cell(ri, 1, codigo)
        ws.cell(ri, 2, descricao)
        c3 = ws.cell(ri, 3, agg["quantidade_total"]); c3.number_format = QTD_FMT
        c4 = ws.cell(ri, 4, agg["faturamento_total"]); c4.number_format = BRL
        c5 = ws.cell(ri, 5, agg["lucro_total"]); c5.number_format = BRL
        if ri % 2 == 0:
            for ci in range(1, 6):
                ws.cell(ri, ci).fill = ALT_FILL
