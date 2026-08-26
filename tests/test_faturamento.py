import io
from datetime import date, timedelta
from decimal import Decimal

import openpyxl
import pytest

from app.models.item_faturamento import ItemFaturamento
from app.models.produto import Produto


def _produto(**overrides):
    dados = dict(
        deletado=" ",
        filial="  ",
        codigo="PROD1",
        descricao="Produto Exemplo",
        tipo="PA",
        conversao=Decimal("1.0000"),
    )
    dados.update(overrides)
    return Produto(**dados)


def _item(**overrides):
    dados = dict(
        deletado=" ",
        filial="01",
        emissao="20260805",
        codigo_produto="PROD1",
        operacao="501",
        quantidade=Decimal("10.0000"),
        total=Decimal("1000.00"),
        custo=Decimal("600.00"),
    )
    dados.update(overrides)
    return ItemFaturamento(**dados)


def test_requer_api_key(client):
    assert client.get("/faturamento/").status_code == 401


def test_agrega_vendas_do_mesmo_dia_produto_e_filial(client, auth_headers, db_session):
    db_session.add_all([
        _produto(),
        _item(quantidade=Decimal("10.0000"), total=Decimal("1000.00"), custo=Decimal("600.00")),
        _item(quantidade=Decimal("5.0000"), total=Decimal("500.00"), custo=Decimal("300.00")),
    ])
    db_session.commit()

    resposta = client.get("/faturamento/?data_inicial=20260101", headers=auth_headers)
    assert resposta.status_code == 200
    dados = resposta.json()
    assert len(dados["items"]) == 1
    item = dados["items"][0]
    assert item["filial"] == "01"
    assert item["dia"] == 5
    assert item["codigo"] == "PROD1"
    assert item["descricao"] == "Produto Exemplo"
    assert item["quantidade"] == "15.0000"
    assert item["faturamento"] == "1500.00"
    assert item["custo"] == "900.00"
    assert item["lucro_bruto"] == "600.00"
    assert item["margem"] == "40.00"
    # markup = lucro_bruto/custo*100 = 600/900*100 = 66.666...
    assert item["markup"] == "66.67"


def test_bonificacao_zera_faturamento_e_conta_como_prejuizo_no_lucro(client, auth_headers, db_session):
    # select_faturamento.sql zera FATURAMENTO para bonificação (542/543/544)
    # — dar um produto de bonificação não gera receita — mas QTDE conta
    # normalmente (sem CASE). PRECO_MEDIO e LUCRO_BRUTO são calculados pela
    # consulta externa a partir do FATURAMENTO/CUSTO já somados (não mais
    # linha a linha, e não mais do D2_TOTAL bruto), então a bonificação
    # contribui só com o custo (prejuízo), sem nenhuma "receita" contrapondo,
    # e "puxa para baixo" o preço médio ponderado do grupo.
    db_session.add_all([
        _produto(),
        _item(operacao="501", quantidade=Decimal("10.0000"),
              total=Decimal("1000.00"), custo=Decimal("600.00")),
        _item(operacao="542", quantidade=Decimal("2.0000"),
              total=Decimal("200.00"), custo=Decimal("120.00")),
    ])
    db_session.commit()

    resposta = client.get("/faturamento/?data_inicial=20260101", headers=auth_headers)
    item = resposta.json()["items"][0]
    assert item["quantidade"] == "12.0000"
    assert item["faturamento"] == "1000.00"
    assert item["custo"] == "720.00"
    # lucro_bruto = SUM(faturamento) - SUM(custo) = 1000 - (600+120) = 280
    assert item["lucro_bruto"] == "280.00"
    assert item["margem"] == "28.00"
    # markup = lucro_bruto/custo*100 = 280/720*100 = 38.888...
    assert item["markup"] == "38.89"
    # preco_medio = SUM(faturamento)/SUM(qtde) = 1000/12 — média ponderada
    # pela quantidade (não mais a média simples de uma razão por linha,
    # que daria AVG(1000/10, 0/2) = 50; a versão corrigida reflete que o
    # preço médio real de venda cai quando parte da quantidade movimentada
    # foi dada de bonificação, sem gerar receita).
    assert item["preco_medio"] == "83.3333"


def test_grupo_so_com_bonificacao_nao_quebra_com_faturamento_zero(client, auth_headers, db_session):
    # Achado num teste manual da SPA (piloto de Faturamento, Fase 3): um
    # grupo (filial/dia/produto) só com bonificação, sem nenhuma venda, tem
    # faturamento total zero — margem/markup dividem por zero, e preco_medio
    # também divide por zero se a quantidade der zero. A consulta original
    # (SQL Server) lançava "Divide by zero error" (500); aqui, com
    # func.nullif no denominador, o resultado é None (não é possível
    # calcular a razão), não um erro.
    db_session.add_all([
        _produto(),
        _item(operacao="542", quantidade=Decimal("5.0000"),
              total=Decimal("300.00"), custo=Decimal("150.00")),
    ])
    db_session.commit()

    resposta = client.get("/faturamento/?data_inicial=20260101", headers=auth_headers)
    assert resposta.status_code == 200
    item = resposta.json()["items"][0]
    assert item["quantidade"] == "5.0000"
    assert item["faturamento"] == "0.00"
    assert item["custo"] == "150.00"
    assert item["lucro_bruto"] == "-150.00"
    assert item["margem"] is None
    # markup não depende de faturamento (denominador é custo, não zero aqui)
    assert item["markup"] == "-100.00"
    # preco_medio = faturamento/qtde = 0/5 = 0 — QTDE não é zero, então essa
    # razão é calculável normalmente (diferente de margem, que é None).
    assert item["preco_medio"] == "0.0000"


def test_quantidade_dividida_pela_conversao_do_produto(client, auth_headers, db_session):
    db_session.add_all([
        _produto(conversao=Decimal("2.0000")),
        _item(quantidade=Decimal("10.0000")),
    ])
    db_session.commit()

    resposta = client.get("/faturamento/?data_inicial=20260101", headers=auth_headers)
    assert resposta.json()["items"][0]["quantidade"] == "5.0000"


def test_produto_tipo_diferente_de_pa_e_ignorado(client, auth_headers, db_session):
    db_session.add_all([_produto(tipo="AM"), _item()])
    db_session.commit()

    resposta = client.get("/faturamento/?data_inicial=20260101", headers=auth_headers)
    assert len(resposta.json()["items"]) == 0


def test_operacao_fora_da_lista_e_ignorada(client, auth_headers, db_session):
    db_session.add_all([_produto(), _item(operacao="999")])
    db_session.commit()

    resposta = client.get("/faturamento/?data_inicial=20260101", headers=auth_headers)
    assert len(resposta.json()["items"]) == 0


def test_item_sem_produto_correspondente_e_ignorado(client, auth_headers, db_session):
    db_session.add(_item())
    db_session.commit()

    resposta = client.get("/faturamento/?data_inicial=20260101", headers=auth_headers)
    assert len(resposta.json()["items"]) == 0


def test_registros_deletados_sao_ignorados(client, auth_headers, db_session):
    db_session.add_all([
        _produto(),
        _item(total=Decimal("1000.00")),
        _item(total=Decimal("5000.00"), deletado="*"),
    ])
    db_session.commit()

    resposta = client.get("/faturamento/?data_inicial=20260101", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["faturamento"] == "1000.00"


def test_dias_de_meses_diferentes_com_mesmo_numero_sao_somados(client, auth_headers, db_session):
    # DAY(D2_EMISSAO) só considera o dia do mês — dois meses diferentes com
    # o mesmo número de dia (05/01 e 05/02) caem no mesmo grupo. Réplica
    # fiel de select_faturamento.sql, mesmo sendo um comportamento pouco
    # intuitivo à primeira vista.
    db_session.add_all([
        _produto(),
        _item(emissao="20260105", total=Decimal("1000.00")),
        _item(emissao="20260205", total=Decimal("2000.00")),
    ])
    db_session.commit()

    resposta = client.get(
        "/faturamento/?data_inicial=20260101&data_final=20260228", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["dia"] == 5
    assert dados["items"][0]["faturamento"] == "3000.00"


def test_filtro_data_final_via_query_string(client, auth_headers, db_session):
    db_session.add_all([
        _produto(),
        _item(emissao="20260805", total=Decimal("1000.00")),
        _item(emissao="20260815", total=Decimal("2000.00")),
    ])
    db_session.commit()

    resposta = client.get(
        "/faturamento/?data_inicial=20260801&data_final=20260810", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["dia"] == 5


def test_sem_filtro_assume_data_atual(client, auth_headers, db_session):
    hoje = date.today().strftime("%Y%m%d")
    ontem = (date.today() - timedelta(days=1)).strftime("%Y%m%d")
    db_session.add_all([
        _produto(),
        _item(emissao=ontem, total=Decimal("1000.00")),
        _item(emissao=hoje, total=Decimal("2000.00")),
    ])
    db_session.commit()

    resposta = client.get("/faturamento/", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["faturamento"] == "2000.00"


def test_metodos_de_escrita_nao_existem(client, auth_headers):
    assert client.post("/faturamento/", json={}, headers=auth_headers).status_code == 405
    assert client.put("/faturamento/", json={}, headers=auth_headers).status_code == 405
    assert client.delete("/faturamento/", headers=auth_headers).status_code == 405


def test_export_gera_planilha_com_4_abas_e_valores_corretos(client, auth_headers, db_session):
    # Réplica do cenário de test_bonificacao_zera_faturamento_e_conta_como_
    # prejuizo_no_lucro, mas verificando a saída da planilha em vez do JSON.
    db_session.add_all([
        _produto(),
        _item(operacao="501", quantidade=Decimal("10.0000"),
              total=Decimal("1000.00"), custo=Decimal("600.00")),
        _item(operacao="542", quantidade=Decimal("2.0000"),
              total=Decimal("200.00"), custo=Decimal("120.00")),
    ])
    db_session.commit()

    resposta = client.get("/faturamento/export?data_inicial=20260101", headers=auth_headers)
    assert resposta.status_code == 200
    assert resposta.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert 'filename="faturamento_202601.xlsx"' in resposta.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(resposta.content))
    assert wb.sheetnames == ["Faturamento", "Resumo", "Por Filial", "Top Produtos"]

    # Aba "Faturamento": A=Filial B=Dia C=Código D=Descrição E=Quantidade
    # F=Faturamento G=Custo H=Preço Médio I=Lucro Bruto J=Margem K=Markup
    ws1 = wb["Faturamento"]
    assert ws1["A1"].value == "Filial"
    assert ws1["A2"].value == "01"
    assert ws1["E2"].value == pytest.approx(12.0)
    assert ws1["F2"].value == pytest.approx(1000.0)
    assert ws1["G2"].value == pytest.approx(720.0)
    assert ws1["I2"].value == pytest.approx(280.0)
    assert ws1["J2"].value == pytest.approx(28.0)  # já multiplicado por 100

    # Aba "Resumo": Margem/Markup Geral usam o formato "%" nativo do Excel
    # (multiplica por 100 ao exibir), então a célula guarda a fração crua.
    ws2 = wb["Resumo"]
    resumo = {ws2.cell(i, 1).value: ws2.cell(i, 2).value for i in range(4, 13)}
    assert resumo["Total de Registros"] == 1
    assert resumo["Faturamento Total (R$)"] == pytest.approx(1000.0)
    assert resumo["Lucro Bruto Total (R$)"] == pytest.approx(280.0)
    assert resumo["Margem Geral (%)"] == pytest.approx(0.28)

    # Aba "Por Filial": só uma filial, então TOTAL GERAL == a própria linha.
    ws3 = wb["Por Filial"]
    assert ws3["A2"].value == "01"
    assert ws3["C2"].value == pytest.approx(1000.0)
    assert ws3["A3"].value == "TOTAL GERAL"

    # Aba "Top Produtos": um único produto agregado.
    ws4 = wb["Top Produtos"]
    assert ws4["A2"].value == "PROD1"
    assert ws4["D2"].value == pytest.approx(1000.0)


def test_export_aplica_filtros_locais_de_filial_e_produto(client, auth_headers, db_session):
    # filial/produto na exportação existem só para bater com o que o
    # usuário vê filtrado na tela da SPA — não são filtros de negócio da
    # consulta, por isso são aplicados aqui em Python (ver _filtrar_items).
    db_session.add_all([
        _produto(codigo="PROD1", descricao="Produto Um"),
        _produto(codigo="PROD2", descricao="Produto Dois"),
        _item(filial="01", codigo_produto="PROD1",
              total=Decimal("1000.00"), custo=Decimal("600.00")),
        _item(filial="02", codigo_produto="PROD2",
              total=Decimal("500.00"), custo=Decimal("300.00")),
    ])
    db_session.commit()

    resposta_filial = client.get(
        "/faturamento/export?data_inicial=20260101&filial=01", headers=auth_headers)
    ws_filial = openpyxl.load_workbook(io.BytesIO(resposta_filial.content))["Faturamento"]
    filiais = [ws_filial.cell(r, 1).value for r in range(2, ws_filial.max_row + 1)
               if ws_filial.cell(r, 1).value]
    assert filiais == ["01"]

    resposta_produto = client.get(
        "/faturamento/export?data_inicial=20260101&produto=dois", headers=auth_headers)
    ws_produto = openpyxl.load_workbook(io.BytesIO(resposta_produto.content))["Faturamento"]
    codigos = [ws_produto.cell(r, 3).value for r in range(2, ws_produto.max_row + 1)
               if ws_produto.cell(r, 3).value]
    assert codigos == ["PROD2"]
