"""Geração da planilha Excel do relatório financeiro — réplica de
`_write_excel()` em `client/app_financeiro.py`, mesma estrutura de 8 abas e
mesma formatação, a partir dos dados já calculados pela API + categorização
(ver categorizacao_financeiro.py) e status/mês-ano calculados aqui, do
mesmo jeito que o client desktop calcula ao carregar os dados.

Sem pandas (ver app/excel/faturamento.py — mesmo raciocínio): pivôs
(Evolução Mensal, Total por Dia) montados com dicionários aninhados em vez
de `DataFrame.groupby().unstack()`.
"""
import io
from collections import OrderedDict
from datetime import datetime
from decimal import Decimal
from typing import List, Sequence

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.excel.categorizacao_financeiro import categorizar
from app.excel.styles import ALT_FILL, BOLD10, BORDER, BRL, LEFT, PCT, RIGHT, cabecalho
from app.schemas.financeiro import FinanceiroRead

# Cores extras deste relatório (fundo por status na aba 1) — as demais
# (HDR_FILL, ALT_FILL etc.) vêm de app/excel/styles.py.
VENC_FILL = PatternFill("solid", fgColor="FDEDEC")
BAIX_FILL = PatternFill("solid", fgColor="F2F3F4")

_STATUS_ORDEM = ("Em aberto", "Vencido", "Baixado")


def _num(valor):
    return float(valor) if isinstance(valor, Decimal) else (valor or 0.0)


def _fmt_data(valor) -> str:
    s = str(valor or "").strip()
    if len(s) == 8 and s.isdigit():
        return f"{s[6:8]}/{s[4:6]}/{s[:4]}"
    return s or ""


def _mes_ano(vencimento_real) -> str:
    s = str(vencimento_real or "").strip()
    return s[:6] if len(s) == 8 and s.isdigit() else ""


def _fmt_mes(mes_ano: str) -> str:
    return f"{mes_ano[4:6]}/{mes_ano[:4]}" if len(mes_ano) == 6 else (mes_ano or "")


def _fmt_rec_judicial(v) -> str:
    return "Sim" if str(v or "").strip() == "1" else "Não"


def _status(item: FinanceiroRead, hoje: str) -> str:
    baixa = (item.data_baixa or "").strip()
    if baixa:
        return "Baixado"
    vcto = (item.vencimento_real or "").strip()
    if vcto and vcto < hoje:
        return "Vencido"
    return "Em aberto"


class _LinhaEnriquecida:
    """FinanceiroRead + os campos que a API não expõe (calculados aqui, do
    mesmo jeito que o client desktop calcula ao carregar os dados)."""

    __slots__ = ("item", "status", "categoria", "mes_ano")

    def __init__(self, item: FinanceiroRead, status: str, categoria: str, mes_ano: str):
        self.item = item
        self.status = status
        self.categoria = categoria
        self.mes_ano = mes_ano


def _enriquecer(linhas: Sequence[FinanceiroRead]) -> List[_LinhaEnriquecida]:
    hoje = datetime.now().strftime("%Y%m%d")
    return [
        _LinhaEnriquecida(
            item=item,
            status=_status(item, hoje),
            categoria=categorizar(item.nome_fornecedor, item.historico),
            mes_ano=_mes_ano(item.vencimento_real),
        )
        for item in linhas
    ]


def gerar_excel_financeiro(linhas: Sequence[FinanceiroRead]) -> io.BytesIO:
    enriquecidas = _enriquecer(linhas)

    wb = openpyxl.Workbook()
    _aba_financeiro(wb.active, enriquecidas)
    _aba_resumo(wb.create_sheet("Resumo"), enriquecidas)
    _aba_por_fornecedor(wb.create_sheet("Por Fornecedor"), enriquecidas)
    _aba_por_tipo_operacao(wb.create_sheet("Por Tipo de Operação"), enriquecidas)
    _aba_resumo_por_categoria(wb.create_sheet("Resumo por Categoria"), enriquecidas)
    _aba_evolucao_mensal(wb.create_sheet("Evolução Mensal"), enriquecidas)
    _aba_total_por_dia(wb.create_sheet("Total por Dia"), enriquecidas)
    _aba_resumo_por_filial(wb.create_sheet("Resumo por Filial"), enriquecidas)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── aba 1: Financeiro ───────────────────────────────────────────────────

_COLUNAS_ABA1 = [
    ("Filial", 7), ("Número", 13), ("Parcela", 9), ("Tipo", 7),
    ("Cód. Operação", 12), ("Tipo Operação", 26), ("Fornecedor", 30),
    ("Categoria", 22), ("Emissão", 13), ("Vencimento", 13), ("Mês/Ano", 10),
    ("Valor (R$)", 16), ("Saldo (R$)", 16), ("Histórico", 42),
    ("Rec. Judicial", 14), ("Status", 12),
]
_COLS_VALOR = (12, 13)  # Valor, Saldo — 1-indexadas


def _aba_financeiro(ws, linhas: Sequence[_LinhaEnriquecida]) -> None:
    ws.title = "Financeiro"
    ws.freeze_panes = "A2"
    for ci, (titulo, largura) in enumerate(_COLUNAS_ABA1, 1):
        cabecalho(ws, 1, ci, titulo)
        ws.column_dimensions[get_column_letter(ci)].width = largura
    ws.row_dimensions[1].height = 18
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUNAS_ABA1))}1"

    for ri, linha in enumerate(linhas, 2):
        item = linha.item
        if linha.status == "Vencido":
            preencher = VENC_FILL
        elif linha.status == "Baixado":
            preencher = BAIX_FILL
        else:
            preencher = ALT_FILL if ri % 2 == 0 else None

        valores = [
            item.filial, item.numero, item.parcela, item.tipo, item.codigo_operacao,
            item.descricao_operacao, item.nome_fornecedor, linha.categoria,
            _fmt_data(item.emissao), _fmt_data(item.vencimento_real), _fmt_mes(linha.mes_ano),
            _num(item.valor), _num(item.saldo), item.historico,
            _fmt_rec_judicial(item.recuperacao_judicial), linha.status,
        ]
        for ci, valor in enumerate(valores, 1):
            celula = ws.cell(row=ri, column=ci, value=valor)
            celula.border = BORDER
            if ci in _COLS_VALOR:
                celula.number_format = BRL
                celula.alignment = RIGHT
            else:
                celula.alignment = LEFT
            if preencher:
                celula.fill = preencher


# ── aba 2: Resumo ───────────────────────────────────────────────────────

def _aba_resumo(ws, linhas: Sequence[_LinhaEnriquecida]) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.cell(1, 1, "Resumo — Relatório Financeiro").font = Font(
        bold=True, color="1A5276", size=14, name="Calibri")
    ws.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(
        italic=True, size=9, color="7F8C8D", name="Calibri")

    aberto = [l for l in linhas if l.status == "Em aberto"]
    vencido = [l for l in linhas if l.status == "Vencido"]
    baixado = [l for l in linhas if l.status == "Baixado"]

    # ATENÇÃO: réplica fiel do client desktop — este resumo NÃO exclui
    # títulos de recuperação judicial do valor/saldo, diferente dos KPIs ao
    # vivo da tela da SPA (que excluem — ver features/financeiro/kpis.ts).
    # Inconsistência já existente no client/app_financeiro.py original
    # (_update_kpis exclui RJ, _write_excel não) — mantida de propósito,
    # não corrigida sem confirmação de que a correção é desejada.
    itens_resumo = [
        ("Total de Títulos", len(linhas), None),
        ("Valor Total (R$)", sum(_num(l.item.valor) for l in linhas), BRL),
        ("Saldo Total (R$)", sum(_num(l.item.saldo) for l in linhas), BRL),
        ("", None, None),
        ("Em Aberto — Qtd", len(aberto), None),
        ("Em Aberto — Saldo (R$)", sum(_num(l.item.saldo) for l in aberto), BRL),
        ("", None, None),
        ("Vencidos — Qtd", len(vencido), None),
        ("Vencidos — Saldo (R$)", sum(_num(l.item.saldo) for l in vencido), BRL),
        ("", None, None),
        ("Baixados — Qtd", len(baixado), None),
        ("Baixados — Valor (R$)", sum(_num(l.item.valor) for l in baixado), BRL),
    ]
    for i, (rotulo, valor, formato) in enumerate(itens_resumo, 4):
        c1 = ws.cell(i, 1, rotulo)
        c1.font = Font(bold=bool(rotulo), name="Calibri", size=10)
        c2 = ws.cell(i, 2, valor)
        c2.alignment = RIGHT
        if formato:
            c2.number_format = formato


# ── aba 3: Por Fornecedor / aba 4: Por Tipo de Operação ─────────────────
# Mesma forma de agregação (qtd + valor_total + saldo_total), só troca a
# chave de agrupamento e o texto do "sem valor" — fatorado num helper.

def _agrupar_qtd_valor_saldo(linhas: Sequence[_LinhaEnriquecida], chave_fn, sem_valor: str):
    agregados: "OrderedDict[str, dict]" = OrderedDict()
    for l in linhas:
        chave = chave_fn(l) or sem_valor
        agg = agregados.setdefault(chave, {"qtd": 0, "valor_total": 0.0, "saldo_total": 0.0})
        agg["qtd"] += 1
        agg["valor_total"] += _num(l.item.valor)
        agg["saldo_total"] += _num(l.item.saldo)
    return sorted(agregados.items(), key=lambda kv: kv[1]["saldo_total"], reverse=True)


def _escrever_tabela_qtd_valor_saldo(ws, titulo_chave: str, ordenados) -> None:
    hdrs = [titulo_chave, "Qtd Títulos", "Valor Total (R$)", "Saldo Total (R$)"]
    larguras = [30, 14, 20, 20]
    for ci, (h, w) in enumerate(zip(hdrs, larguras), 1):
        cabecalho(ws, 1, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ri, (chave, agg) in enumerate(ordenados, 2):
        ws.cell(ri, 1, chave)
        ws.cell(ri, 2, agg["qtd"])
        c3 = ws.cell(ri, 3, agg["valor_total"]); c3.number_format = BRL
        c4 = ws.cell(ri, 4, agg["saldo_total"]); c4.number_format = BRL
        if ri % 2 == 0:
            for ci in range(1, 5):
                ws.cell(ri, ci).fill = ALT_FILL


def _aba_por_fornecedor(ws, linhas: Sequence[_LinhaEnriquecida]) -> None:
    ordenados = _agrupar_qtd_valor_saldo(
        linhas, lambda l: l.item.nome_fornecedor, "(sem nome)")
    _escrever_tabela_qtd_valor_saldo(ws, "Fornecedor", ordenados)


def _aba_por_tipo_operacao(ws, linhas: Sequence[_LinhaEnriquecida]) -> None:
    ordenados = _agrupar_qtd_valor_saldo(
        linhas, lambda l: l.item.descricao_operacao, "(não informado)")
    _escrever_tabela_qtd_valor_saldo(ws, "Tipo Operação", ordenados)


# ── aba 5: Resumo por Categoria ──────────────────────────────────────────

def _aba_resumo_por_categoria(ws, linhas: Sequence[_LinhaEnriquecida]) -> None:
    ws.cell(1, 1, "Análise de Contas a Pagar por Categoria").font = Font(
        bold=True, color="1A5276", size=13, name="Calibri")
    ws.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(
        italic=True, size=9, color="7F8C8D", name="Calibri")

    agregados: "OrderedDict[str, dict]" = OrderedDict()
    for l in linhas:
        agg = agregados.setdefault(l.categoria, {"qtd": 0, "valor_total": 0.0})
        agg["qtd"] += 1
        agg["valor_total"] += _num(l.item.valor)
    ordenados = sorted(agregados.items(), key=lambda kv: kv[1]["valor_total"], reverse=True)
    total_geral = sum(v["valor_total"] for _, v in ordenados)

    hdr_row = 4
    hdrs = ["Categoria", "Qtd. Títulos", "Valor Total (R$)", "% do Total", "Valor Médio (R$)"]
    larguras = [26, 14, 20, 12, 20]
    for ci, (h, w) in enumerate(zip(hdrs, larguras), 1):
        cabecalho(ws, hdr_row, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{hdr_row + 1}"

    ri = hdr_row + 1
    for categoria, agg in ordenados:
        pct = (agg["valor_total"] / total_geral) if total_geral else 0.0
        valor_medio = (agg["valor_total"] / agg["qtd"]) if agg["qtd"] else 0.0
        ws.cell(ri, 1, categoria)
        ws.cell(ri, 2, agg["qtd"])
        c3 = ws.cell(ri, 3, agg["valor_total"]); c3.number_format = BRL
        c4 = ws.cell(ri, 4, pct); c4.number_format = PCT
        c5 = ws.cell(ri, 5, valor_medio); c5.number_format = BRL
        if (ri - hdr_row) % 2 == 0:
            for ci in range(1, 6):
                ws.cell(ri, ci).fill = ALT_FILL
        ri += 1

    ws.cell(ri, 1, "TOTAL GERAL").font = BOLD10
    ws.cell(ri, 2, sum(v["qtd"] for _, v in ordenados)).font = BOLD10
    c3 = ws.cell(ri, 3, total_geral); c3.number_format = BRL; c3.font = BOLD10
    c4 = ws.cell(ri, 4, 1.0); c4.number_format = PCT; c4.font = BOLD10


# ── aba 6: Evolução Mensal ───────────────────────────────────────────────

def _aba_evolucao_mensal(ws, linhas: Sequence[_LinhaEnriquecida]) -> None:
    ws.cell(1, 1, "Evolução dos Custos ao Longo do Tempo").font = Font(
        bold=True, color="1A5276", size=13, name="Calibri")
    ws.cell(2, 1, "Valores por vencimento, agrupados por mês e categoria").font = Font(
        italic=True, size=9, color="7F8C8D", name="Calibri")

    meses = sorted({l.mes_ano for l in linhas if l.mes_ano})
    pivo: "OrderedDict[str, dict]" = OrderedDict()
    for l in linhas:
        if not l.mes_ano:
            continue
        linha_pivo = pivo.setdefault(l.categoria, {m: 0.0 for m in meses})
        linha_pivo[l.mes_ano] += _num(l.item.valor)
    # Categorias presentes na pivô (algumas linhas podem ter mes_ano vazio e
    # nunca entrar aqui, mas ainda assim ter categoria) — inclui todas as
    # categorias que apareceram em pelo menos uma linha com mês válido.
    linhas_ordenadas = sorted(
        pivo.items(), key=lambda kv: sum(kv[1].values()), reverse=True)

    hdr_row = 4
    headers = ["Categoria"] + [_fmt_mes(m) for m in meses] + ["Total"]
    for ci, h in enumerate(headers, 1):
        cabecalho(ws, hdr_row, ci, h)
    ws.column_dimensions["A"].width = 26
    for ci in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.freeze_panes = f"B{hdr_row + 1}"

    totais_mes = {m: 0.0 for m in meses}
    ri = hdr_row + 1
    for categoria, valores_mes in linhas_ordenadas:
        ws.cell(ri, 1, categoria)
        total_linha = 0.0
        for ci, m in enumerate(meses, 2):
            v = valores_mes[m]
            total_linha += v
            totais_mes[m] += v
            c = ws.cell(ri, ci, v)
            c.number_format = BRL
        ctotal = ws.cell(ri, len(meses) + 2, total_linha)
        ctotal.number_format = BRL
        ctotal.font = BOLD10
        ri += 1

    ws.cell(ri, 1, "TOTAL MENSAL").font = BOLD10
    for ci, m in enumerate(meses, 2):
        c = ws.cell(ri, ci, totais_mes[m])
        c.number_format = BRL
        c.font = BOLD10
    ctotal_geral = ws.cell(ri, len(meses) + 2, sum(totais_mes.values()))
    ctotal_geral.number_format = BRL
    ctotal_geral.font = BOLD10


# ── aba "Total por Dia" ──────────────────────────────────────────────────

def _aba_total_por_dia(ws, linhas: Sequence[_LinhaEnriquecida]) -> None:
    ws.cell(1, 1, "Total por Dia — A Pagar, Vencido e Baixado").font = Font(
        bold=True, color="1A5276", size=13, name="Calibri")
    ws.cell(2, 1, 'Valores por dia de vencimento ("Em aberto"/"Vencido" somam saldo; '
            '"Baixado" soma valor)').font = Font(
        italic=True, size=9, color="7F8C8D", name="Calibri")

    # "Em aberto"/"Vencido" somam saldo (o que ainda falta pagar); "Baixado"
    # soma valor (já foi pago) — mesma convenção de _aba_resumo/kpis.ts.
    # Sem limite de dias, de propósito (ver comentário equivalente no
    # client desktop): o volume já é controlado pelos filtros de período.
    pivo: "OrderedDict[str, dict]" = OrderedDict()
    for l in linhas:
        linha_pivo = pivo.setdefault(
            l.item.vencimento_real, {s: 0.0 for s in _STATUS_ORDEM})
        v = _num(l.item.valor) if l.status == "Baixado" else _num(l.item.saldo)
        linha_pivo[l.status] += v
    dias_ordenados = sorted(pivo.items(), key=lambda kv: kv[0] or "")

    hdr_row = 4
    hdrs = ["Data", "A Pagar (R$)", "Vencido (R$)", "Baixado (R$)", "Total (R$)"]
    larguras = [13, 16, 16, 16, 16]
    for ci, (h, w) in enumerate(zip(hdrs, larguras), 1):
        cabecalho(ws, hdr_row, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{hdr_row + 1}"

    totais = {s: 0.0 for s in _STATUS_ORDEM}
    ri = hdr_row + 1
    for dia, valores_status in dias_ordenados:
        ws.cell(ri, 1, _fmt_data(dia))
        total_linha = 0.0
        for ci, status in enumerate(_STATUS_ORDEM, 2):
            v = valores_status[status]
            total_linha += v
            totais[status] += v
            c = ws.cell(ri, ci, v)
            c.number_format = BRL
        ctotal = ws.cell(ri, 5, total_linha)
        ctotal.number_format = BRL
        if (ri - hdr_row) % 2 == 0:
            for ci in range(1, 6):
                ws.cell(ri, ci).fill = ALT_FILL
        ri += 1

    ws.cell(ri, 1, "TOTAL GERAL").font = BOLD10
    for ci, status in enumerate(_STATUS_ORDEM, 2):
        c = ws.cell(ri, ci, totais[status])
        c.number_format = BRL
        c.font = BOLD10
    ctotal_geral = ws.cell(ri, 5, sum(totais.values()))
    ctotal_geral.number_format = BRL
    ctotal_geral.font = BOLD10


# ── aba 7: Resumo por Filial ──────────────────────────────────────────────

def _aba_resumo_por_filial(ws, linhas: Sequence[_LinhaEnriquecida]) -> None:
    ws.cell(1, 1, "Contas a Pagar por Filial").font = Font(
        bold=True, color="1A5276", size=13, name="Calibri")

    agregados: "OrderedDict[str, dict]" = OrderedDict()
    for l in linhas:
        agg = agregados.setdefault(l.item.filial, {"qtd": 0, "valor_total": 0.0})
        agg["qtd"] += 1
        agg["valor_total"] += _num(l.item.valor)
    ordenados = sorted(agregados.items(), key=lambda kv: kv[1]["valor_total"], reverse=True)
    total_geral = sum(v["valor_total"] for _, v in ordenados)

    hdr_row = 3
    hdrs = ["Filial", "Qtd. Títulos", "Valor Total (R$)", "% do Total"]
    larguras = [18, 14, 20, 12]
    for ci, (h, w) in enumerate(zip(hdrs, larguras), 1):
        cabecalho(ws, hdr_row, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ri = hdr_row + 1
    for filial, agg in ordenados:
        pct = (agg["valor_total"] / total_geral) if total_geral else 0.0
        ws.cell(ri, 1, filial)
        ws.cell(ri, 2, agg["qtd"])
        c3 = ws.cell(ri, 3, agg["valor_total"]); c3.number_format = BRL
        c4 = ws.cell(ri, 4, pct); c4.number_format = PCT
        if (ri - hdr_row) % 2 == 0:
            for ci in range(1, 5):
                ws.cell(ri, ci).fill = ALT_FILL
        ri += 1

    ws.cell(ri, 1, "TOTAL GERAL").font = BOLD10
    ws.cell(ri, 2, sum(v["qtd"] for _, v in ordenados)).font = BOLD10
    c3 = ws.cell(ri, 3, total_geral); c3.number_format = BRL; c3.font = BOLD10
    c4 = ws.cell(ri, 4, 1.0); c4.number_format = PCT; c4.font = BOLD10
