"""Estilos compartilhados para as planilhas Excel geradas pela API.

Réplica dos estilos hoje duplicados em cada `client/app_*.py` (Tkinter) —
centralizados aqui para que os endpoints `/*/export` e os clients desktop
(se algum dia forem atualizados para consumir isto) produzam exatamente a
mesma formatação.
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

HDR_FILL = PatternFill("solid", fgColor="1A5276")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
ALT_FILL = PatternFill("solid", fgColor="EBF5FB")
THIN = Side(style="thin", color="BDC3C7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# BRL/QTD_FMT: formato numérico comum, sem símbolo (o cabeçalho da coluna já
# diz "(R$)"). PCT: formato percentual nativo do Excel — multiplica o valor
# da célula por 100 para exibir (uso: só para frações "cruas", tipo 0.4 -> 40,0%).
# PCT_CELL: formato customizado que só acrescenta o símbolo "%" ao texto, sem
# multiplicar por 100 — uso: valores que a API já devolve multiplicados por
# 100 (ex. campo `margem`, onde 40.00 já significa "40%").
BRL = "#,##0.00"
QTD_FMT = "#,##0.00"
PCT = "0.0%"
PCT_CELL = '#,##0.00"%"'

BOLD10 = Font(bold=True, name="Calibri", size=10)


def cabecalho(ws: Worksheet, linha: int, coluna: int, texto: str):
    """Escreve e formata uma célula de cabeçalho de tabela (fundo azul,
    fonte branca em negrito, centralizada, com borda)."""
    celula = ws.cell(row=linha, column=coluna, value=texto)
    celula.font = HDR_FONT
    celula.fill = HDR_FILL
    celula.alignment = CENTER
    celula.border = BORDER
    return celula
