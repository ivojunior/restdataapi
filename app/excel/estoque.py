"""Geração da planilha Excel do relatório de saldo de estoque — réplica de
`_write_excel()` em `client/app_estoque.py`, mesma estrutura de 4 abas e
mesma formatação, a partir dos dados já calculados pela API.

Sem pandas (ver app/excel/faturamento.py — mesmo raciocínio): agregações
das abas "Por Filial"/"Top Produtos" feitas com dicionários simples.
"""
import io
from collections import OrderedDict
from datetime import datetime
from decimal import Decimal
from typing import Sequence

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.excel.styles import ALT_FILL, BOLD10, BORDER, BRL, LEFT, PCT, QTD_FMT, RIGHT, cabecalho
from app.schemas.saldo_estoque import SaldoEstoqueRead

_COLUNAS_ABA1 = [
    ("Filial", "filial", 9),
    ("Local", "local", 9),
    ("Código Produto", "codigo_produto", 16),
    ("Descrição", "descricao_produto", 42),
    ("Quantidade", "quantidade", 16),
    ("Valor Atual (R$)", "valor_atual", 18),
]


def _num(valor):
    return float(valor) if isinstance(valor, Decimal) else (valor or 0.0)


def gerar_excel_estoque(linhas: Sequence[SaldoEstoqueRead]) -> io.BytesIO:
    wb = openpyxl.Workbook()

    _aba_estoque(wb.active, linhas)
    _aba_resumo(wb.create_sheet("Resumo"), linhas)
    _aba_por_filial(wb.create_sheet("Por Filial"), linhas)
    _aba_top_produtos(wb.create_sheet("Top Produtos"), linhas)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _aba_estoque(ws, linhas: Sequence[SaldoEstoqueRead]) -> None:
    ws.title = "Estoque"
    ws.freeze_panes = "A2"

    for ci, (titulo, _, largura) in enumerate(_COLUNAS_ABA1, 1):
        cabecalho(ws, 1, ci, titulo)
        ws.column_dimensions[get_column_letter(ci)].width = largura
    ws.row_dimensions[1].height = 18
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUNAS_ABA1))}1"

    for ri, linha in enumerate(linhas, 2):
        preencher_linha = ALT_FILL if ri % 2 == 0 else None
        for ci, (_, campo, _largura) in enumerate(_COLUNAS_ABA1, 1):
            valor = getattr(linha, campo)
            if isinstance(valor, Decimal):
                valor = float(valor)
            celula = ws.cell(row=ri, column=ci, value=valor)
            celula.border = BORDER
            if campo == "valor_atual":
                celula.number_format = BRL
                celula.alignment = RIGHT
            elif campo == "quantidade":
                celula.number_format = QTD_FMT
                celula.alignment = RIGHT
            else:
                celula.alignment = LEFT
            if preencher_linha:
                celula.fill = preencher_linha


def _aba_resumo(ws, linhas: Sequence[SaldoEstoqueRead]) -> None:
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    ws.cell(1, 1, "Resumo — Saldo de Estoque").font = Font(
        bold=True, color="1A5276", size=14, name="Calibri")
    ws.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(
        italic=True, size=9, color="7F8C8D", name="Calibri")

    valor_total = sum(_num(l.valor_atual) for l in linhas)
    quantidade_total = sum(_num(l.quantidade) for l in linhas)
    valor_medio = (valor_total / quantidade_total) if quantidade_total else 0.0

    itens_resumo = [
        ("Total de Itens", len(linhas), None),
        ("Filiais Distintas", len({l.filial for l in linhas}), None),
        ("Quantidade Total", quantidade_total, QTD_FMT),
        ("Valor Total (R$)", valor_total, BRL),
        ("Valor Médio por Item (R$)", valor_medio, BRL),
    ]
    for i, (rotulo, valor, formato) in enumerate(itens_resumo, 4):
        c1 = ws.cell(i, 1, rotulo)
        c1.font = Font(bold=True, name="Calibri", size=10)
        c2 = ws.cell(i, 2, valor)
        c2.alignment = RIGHT
        if formato:
            c2.number_format = formato


def _aba_por_filial(ws, linhas: Sequence[SaldoEstoqueRead]) -> None:
    agregados: "OrderedDict[str, dict]" = OrderedDict()
    for l in linhas:
        agg = agregados.setdefault(
            l.filial, {"qtd_itens": 0, "quantidade_total": 0.0, "valor_total": 0.0})
        agg["qtd_itens"] += 1
        agg["quantidade_total"] += _num(l.quantidade)
        agg["valor_total"] += _num(l.valor_atual)

    ordenados = sorted(agregados.items(), key=lambda kv: kv[1]["valor_total"], reverse=True)
    total_geral = sum(v["valor_total"] for _, v in ordenados)

    cabecalhos = ["Filial", "Qtd. Itens", "Quantidade Total", "Valor Total (R$)", "% do Total"]
    larguras = [12, 12, 18, 18, 12]
    for ci, (titulo, largura) in enumerate(zip(cabecalhos, larguras), 1):
        cabecalho(ws, 1, ci, titulo)
        ws.column_dimensions[get_column_letter(ci)].width = largura

    ri = 2
    for filial, agg in ordenados:
        pct = (agg["valor_total"] / total_geral) if total_geral else 0.0
        ws.cell(ri, 1, filial)
        ws.cell(ri, 2, agg["qtd_itens"])
        c3 = ws.cell(ri, 3, agg["quantidade_total"]); c3.number_format = QTD_FMT
        c4 = ws.cell(ri, 4, agg["valor_total"]); c4.number_format = BRL
        c5 = ws.cell(ri, 5, pct); c5.number_format = PCT
        if ri % 2 == 0:
            for ci in range(1, 6):
                ws.cell(ri, ci).fill = ALT_FILL
        ri += 1

    ws.cell(ri, 1, "TOTAL GERAL").font = BOLD10
    ws.cell(ri, 2, sum(v["qtd_itens"] for _, v in ordenados)).font = BOLD10
    c3 = ws.cell(ri, 3, sum(v["quantidade_total"] for _, v in ordenados)); c3.number_format = QTD_FMT; c3.font = BOLD10
    c4 = ws.cell(ri, 4, total_geral); c4.number_format = BRL; c4.font = BOLD10
    c5 = ws.cell(ri, 5, 1.0); c5.number_format = PCT; c5.font = BOLD10


def _aba_top_produtos(ws, linhas: Sequence[SaldoEstoqueRead], limite: int = 50) -> None:
    agregados: "OrderedDict[tuple, dict]" = OrderedDict()
    for l in linhas:
        chave = (l.codigo_produto, l.descricao_produto)
        agg = agregados.setdefault(chave, {"quantidade_total": 0.0, "valor_total": 0.0})
        agg["quantidade_total"] += _num(l.quantidade)
        agg["valor_total"] += _num(l.valor_atual)

    top = sorted(agregados.items(), key=lambda kv: kv[1]["valor_total"], reverse=True)[:limite]

    cabecalhos = ["Código Produto", "Descrição", "Quantidade Total", "Valor Total (R$)"]
    larguras = [16, 42, 18, 18]
    for ci, (titulo, largura) in enumerate(zip(cabecalhos, larguras), 1):
        cabecalho(ws, 1, ci, titulo)
        ws.column_dimensions[get_column_letter(ci)].width = largura

    for ri, ((codigo, descricao), agg) in enumerate(top, 2):
        ws.cell(ri, 1, codigo)
        ws.cell(ri, 2, descricao)
        c3 = ws.cell(ri, 3, agg["quantidade_total"]); c3.number_format = QTD_FMT
        c4 = ws.cell(ri, 4, agg["valor_total"]); c4.number_format = BRL
        if ri % 2 == 0:
            for ci in range(1, 5):
                ws.cell(ri, ci).fill = ALT_FILL
