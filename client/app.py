"""
Cliente desktop para a RestDataAPI — Contas a Pagar (Protheus SE2070).

Exibe filtros, KPIs, quatro gráficos analíticos e exporta para Excel
com quatro abas (dados, resumo, por fornecedor, por vencimento).

Requisitos:  pip install requests matplotlib pandas openpyxl python-dotenv
"""

from __future__ import annotations

import threading
from datetime import date, datetime
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
from typing import Dict, List, Optional, Tuple

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from matplotlib.patches import Patch
import matplotlib.ticker as mticker
import pandas as pd

from api_client import APIClient
from config import API_BASE_URL, API_KEY, API_KEY_NAME


# ── constantes ───────────────────────────────────────────────────────────────

TODAY = date.today().strftime("%Y%m%d")

STATUS_COLORS: Dict[str, str] = {
    "Em aberto": "#27ae60",
    "Vencido":   "#e74c3c",
    "Baixado":   "#95a5a6",
}

CHART_COLORS = [
    "#2980b9", "#e74c3c", "#27ae60", "#f39c12", "#8e44ad",
    "#16a085", "#d35400", "#2c3e50", "#c0392b", "#1abc9c",
]

TREEVIEW_COLS: List[Tuple[str, str, int, str]] = [
    ("rec_no",     "Rec",           60,  "e"),
    ("filial",     "Filial",        55,  "c"),
    ("prefixo",    "Prefixo",       60,  "c"),
    ("numero",     "Número",        90,  "c"),
    ("parcela",    "Parcela",       60,  "c"),
    ("tipo",       "Tipo",          50,  "c"),
    ("fornecedor", "Fornecedor",    80,  "c"),
    ("loja",       "Loja",          45,  "c"),
    ("emissao",    "Emissão",       85,  "c"),
    ("vencimento", "Vencimento",    85,  "c"),
    ("valor",      "Valor (R$)",   115,  "e"),
    ("saldo",      "Saldo (R$)",   115,  "e"),
    ("moeda",      "Moeda",         50,  "c"),
    ("historico",  "Histórico",    220,  "w"),
    ("status",     "Status",        85,  "c"),
]


# ── helpers ──────────────────────────────────────────────────────────────────

def _brl(v) -> str:
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"


def _fmt_date(s) -> str:
    s = str(s or "").strip()
    if len(s) == 8 and s.isdigit():
        return f"{s[6:8]}/{s[4:6]}/{s[:4]}"
    return s or "—"


def _parse_date_input(s: str) -> str:
    """Aceita DD/MM/AAAA ou AAAAMMDD; retorna AAAAMMDD para comparação."""
    s = s.strip().replace("/", "").replace("-", "").replace(".", "")
    if len(s) != 8 or not s.isdigit():
        return s
    # Tenta AAAAMMDD
    try:
        date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        return s
    except ValueError:
        pass
    # Tenta DDMMAAAA
    try:
        d = date(int(s[4:8]), int(s[2:4]), int(s[:2]))
        return d.strftime("%Y%m%d")
    except ValueError:
        pass
    return s


def _status_from_row(row: Dict) -> str:
    baixa = (row.get("data_baixa") or "").strip()
    if baixa:
        return "Baixado"
    vcto = (row.get("vencimento") or "").strip()
    if vcto and vcto < TODAY:
        return "Vencido"
    return "Em aberto"


# ── aplicação principal ───────────────────────────────────────────────────────

class ContasAPagarApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Contas a Pagar — RestDataAPI")
        self.root.geometry("1300x840")
        self.root.minsize(960, 640)
        self.root.configure(bg="#f0f2f5")

        self._df: Optional[pd.DataFrame] = None
        self._loading = False
        self._sort_ascending: Dict[str, bool] = {}

        self._setup_styles()
        self._build_ui()

    # ── estilos ──────────────────────────────────────────────────────────────

    def _setup_styles(self) -> None:
        s = ttk.Style(self.root)
        s.theme_use("clam")

        BG, WHITE, PRIMARY, ACCENT = "#f0f2f5", "#ffffff", "#1a5276", "#2980b9"

        s.configure("TFrame", background=BG)
        s.configure("White.TFrame", background=WHITE)
        s.configure("TLabel", background=BG, font=("Segoe UI", 9))
        s.configure("TLabelframe", background=BG, font=("Segoe UI", 9, "bold"))
        s.configure("TLabelframe.Label", background=BG, font=("Segoe UI", 9, "bold"),
                    foreground=PRIMARY)

        for name, bg, fg, hover in [
            ("Primary",   PRIMARY,   WHITE, ACCENT),
            ("Secondary", "#7f8c8d", WHITE, "#95a5a6"),
            ("Success",   "#27ae60", WHITE, "#2ecc71"),
        ]:
            s.configure(f"{name}.TButton",
                        background=bg, foreground=fg,
                        font=("Segoe UI", 9, "bold"), padding=(12, 5))
            s.map(f"{name}.TButton", background=[("active", hover), ("pressed", bg)])

        s.configure("TEntry", fieldbackground=WHITE, padding=4)
        s.configure("TNotebook", background=BG)
        s.configure("TNotebook.Tab", padding=(12, 6), font=("Segoe UI", 9))

        s.configure("Treeview", background=WHITE, fieldbackground=WHITE,
                    font=("Segoe UI", 9), rowheight=22)
        s.configure("Treeview.Heading",
                    font=("Segoe UI", 9, "bold"),
                    background=PRIMARY, foreground=WHITE)
        s.map("Treeview",
              background=[("selected", ACCENT)],
              foreground=[("selected", WHITE)])

    # ── construção da UI ─────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        self._build_conn_frame()
        self._build_filter_frame()
        self._build_kpi_frame()
        self._build_toolbar()
        self._build_notebook()
        self._build_statusbar()

    # ── conexão ──────────────────────────────────────────────────────────────

    def _build_conn_frame(self) -> None:
        f = ttk.LabelFrame(self.root, text="Conexão", padding=(10, 6))
        f.pack(fill="x", padx=12, pady=(8, 3))

        ttk.Label(f, text="URL da API:").grid(row=0, column=0, sticky="w", padx=(0, 4))
        self._url_var = tk.StringVar(value=API_BASE_URL)
        ttk.Entry(f, textvariable=self._url_var, width=36).grid(
            row=0, column=1, sticky="ew", padx=(0, 18))

        ttk.Label(f, text="API Key:").grid(row=0, column=2, sticky="w", padx=(0, 4))
        self._key_var = tk.StringVar(value=API_KEY)
        ttk.Entry(f, textvariable=self._key_var, show="•", width=34).grid(
            row=0, column=3, sticky="ew", padx=(0, 12))

        ttk.Button(f, text="Testar conexão", style="Secondary.TButton",
                   command=self._test_connection).grid(row=0, column=4)

        f.columnconfigure(1, weight=1)
        f.columnconfigure(3, weight=1)

    # ── filtros ───────────────────────────────────────────────────────────────

    def _build_filter_frame(self) -> None:
        f = ttk.LabelFrame(self.root, text="Filtros", padding=(10, 6))
        f.pack(fill="x", padx=12, pady=3)

        # linha 0 — filtros da API (server-side)
        api_filters = [
            ("Filial:",       "_f_filial",      5),
            ("Fornecedor:",   "_f_fornecedor",  10),
            ("Prefixo:",      "_f_prefixo",     6),
            ("Número:",       "_f_numero",      12),
        ]
        col = 0
        for label, attr, width in api_filters:
            ttk.Label(f, text=label).grid(
                row=0, column=col, sticky="w", padx=(0 if col == 0 else 14, 3))
            var = tk.StringVar()
            setattr(self, attr, var)
            ttk.Entry(f, textvariable=var, width=width).grid(
                row=0, column=col + 1, sticky="ew")
            col += 2

        # linha 1 — filtros de data (client-side)
        ttk.Label(f, text="Vcto De:").grid(
            row=1, column=0, sticky="w", padx=(0, 3), pady=(6, 0))
        self._f_vcto_de = tk.StringVar()
        ttk.Entry(f, textvariable=self._f_vcto_de, width=12).grid(
            row=1, column=1, sticky="w", pady=(6, 0))

        ttk.Label(f, text="Vcto Até:").grid(
            row=1, column=2, sticky="w", padx=(14, 3), pady=(6, 0))
        self._f_vcto_ate = tk.StringVar()
        ttk.Entry(f, textvariable=self._f_vcto_ate, width=12).grid(
            row=1, column=3, sticky="w", pady=(6, 0))

        ttk.Label(f, text="(DD/MM/AAAA)", foreground="#7f8c8d").grid(
            row=1, column=4, sticky="w", padx=(6, 0), pady=(6, 0))

        # botões
        btn = ttk.Frame(f)
        btn.grid(row=0, column=8, rowspan=2, sticky="e", padx=(20, 0))
        ttk.Button(btn, text="  Consultar  ", style="Primary.TButton",
                   command=self.consultar).pack(side="left", padx=(0, 6))
        ttk.Button(btn, text="Limpar", style="Secondary.TButton",
                   command=self.limpar).pack(side="left")

        f.columnconfigure(8, weight=1)

    # ── KPI cards ─────────────────────────────────────────────────────────────

    def _build_kpi_frame(self) -> None:
        outer = tk.Frame(self.root, bg="#f0f2f5")
        outer.pack(fill="x", padx=12, pady=3)

        self._kpi_vars: Dict[str, tk.StringVar] = {}
        cards_cfg = [
            ("total_titulos", "Total de Títulos",  "#2c3e50"),
            ("valor_total",   "Valor Total",        "#1a5276"),
            ("saldo_total",   "Saldo a Pagar",      "#154360"),
            ("em_aberto",     "Em Aberto",          "#1e8449"),
            ("vencidos",      "Vencidos",           "#922b21"),
            ("baixados",      "Baixados",           "#626567"),
        ]

        for i, (key, label, fg) in enumerate(cards_cfg):
            card = tk.Frame(outer, bg="#ffffff",
                            highlightbackground="#d5d8dc", highlightthickness=1)
            card.grid(row=0, column=i, sticky="nsew", padx=4, pady=2, ipady=10)

            tk.Label(card, text=label, bg="#ffffff",
                     font=("Segoe UI", 8), fg="#7f8c8d").pack(pady=(4, 0))

            var = tk.StringVar(value="—")
            self._kpi_vars[key] = var
            tk.Label(card, textvariable=var, bg="#ffffff",
                     font=("Segoe UI", 12, "bold"), fg=fg).pack()

        for i in range(len(cards_cfg)):
            outer.columnconfigure(i, weight=1)

    # ── toolbar acima do notebook ─────────────────────────────────────────────

    def _build_toolbar(self) -> None:
        bar = ttk.Frame(self.root)
        bar.pack(fill="x", padx=12, pady=(4, 0))
        ttk.Button(bar, text="  Exportar Excel  ", style="Success.TButton",
                   command=self.exportar_excel).pack(side="right")

    # ── notebook ──────────────────────────────────────────────────────────────

    def _build_notebook(self) -> None:
        self._nb = ttk.Notebook(self.root)
        self._nb.pack(fill="both", expand=True, padx=12, pady=(2, 4))

        tab_graficos = ttk.Frame(self._nb)
        self._nb.add(tab_graficos, text="  Gráficos  ")
        self._build_charts_tab(tab_graficos)

        tab_dados = ttk.Frame(self._nb)
        self._nb.add(tab_dados, text="  Dados  ")
        self._build_data_tab(tab_dados)

    def _build_charts_tab(self, parent: ttk.Frame) -> None:
        self._fig = Figure(figsize=(13, 6), dpi=96, facecolor="#f0f2f5")
        self._axes = self._fig.subplots(2, 2)
        self._fig.subplots_adjust(left=0.07, right=0.97, top=0.91,
                                  bottom=0.13, hspace=0.52, wspace=0.38)
        self._canvas = FigureCanvasTkAgg(self._fig, master=parent)
        self._canvas.get_tk_widget().pack(fill="both", expand=True, padx=4, pady=4)
        self._draw_empty_charts()

    def _build_data_tab(self, parent: ttk.Frame) -> None:
        cols = [c[0] for c in TREEVIEW_COLS]
        self._tree = ttk.Treeview(parent, columns=cols,
                                  show="headings", selectmode="browse")

        for col_id, heading, width, anchor in TREEVIEW_COLS:
            self._tree.heading(
                col_id, text=heading,
                command=lambda c=col_id: self._sort_tree(c),
            )
            self._tree.column(col_id, width=width, anchor=anchor,
                              stretch=(col_id == "historico"))

        self._tree.tag_configure("Em aberto", background="#eafaf1")
        self._tree.tag_configure("Vencido",   background="#fdecea")
        self._tree.tag_configure("Baixado",   background="#f2f3f4")

        vsb = ttk.Scrollbar(parent, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        parent.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)

    # ── barra de status ───────────────────────────────────────────────────────

    def _build_statusbar(self) -> None:
        bar = tk.Frame(self.root, bg="#dde1e7", height=24)
        bar.pack(fill="x", side="bottom")

        self._status_var = tk.StringVar(value="Pronto.")
        tk.Label(bar, textvariable=self._status_var,
                 bg="#dde1e7", font=("Segoe UI", 8), anchor="w").pack(
            side="left", padx=8, pady=2)

        self._pct_var = tk.StringVar()
        tk.Label(bar, textvariable=self._pct_var,
                 bg="#dde1e7", font=("Segoe UI", 8)).pack(side="right", padx=4)

        self._progress = ttk.Progressbar(bar, mode="determinate", length=200)
        self._progress.pack(side="right", padx=8, pady=3)

    # ── conexão ───────────────────────────────────────────────────────────────

    def _make_client(self) -> APIClient:
        return APIClient(
            base_url=self._url_var.get().strip(),
            api_key=self._key_var.get().strip(),
            api_key_name=API_KEY_NAME,
        )

    def _test_connection(self) -> None:
        try:
            self._make_client().health()
            messagebox.showinfo("Conexão OK",
                                "A API respondeu com sucesso!")
        except Exception as exc:
            messagebox.showerror("Falha na conexão", str(exc))

    # ── consulta (thread) ─────────────────────────────────────────────────────

    def consultar(self) -> None:
        if self._loading:
            return
        self._loading = True
        self._clear_data()
        self._set_status("Consultando a API…", indeterminate=True)
        threading.Thread(target=self._fetch_thread, daemon=True).start()

    def _fetch_thread(self) -> None:
        try:
            client = self._make_client()

            server_filters: Dict[str, str] = {}
            for attr, key in [
                ("_f_filial",     "filial"),
                ("_f_fornecedor", "fornecedor"),
                ("_f_prefixo",    "prefixo"),
                ("_f_numero",     "numero"),
            ]:
                v = getattr(self, attr).get().strip()
                if v:
                    server_filters[key] = v

            items, total = client.get_all_titulos_pagar(
                progress_callback=self._on_progress, **server_filters
            )
            self.root.after(0, self._on_data_ready, items, total)
        except Exception as exc:
            self.root.after(0, self._on_error, str(exc))

    def _on_progress(self, loaded: int, total: int) -> None:
        pct = int(100 * loaded / total) if total else 0
        self.root.after(0, self._update_progress_bar, loaded, total, pct)

    def _update_progress_bar(self, loaded: int, total: int, pct: int) -> None:
        self._progress.stop()
        self._progress.configure(mode="determinate", value=pct)
        self._pct_var.set(f"{pct}%  ({loaded}/{total})")
        self._status_var.set(f"Carregando… {loaded} de {total} registros")

    def _on_data_ready(self, items: List[Dict], total_api: int) -> None:
        self._loading = False

        for row in items:
            row["status"] = _status_from_row(row)

        df = pd.DataFrame(items) if items else pd.DataFrame()
        if not df.empty:
            df["valor"] = pd.to_numeric(df["valor"], errors="coerce").fillna(0.0)
            df["saldo"] = pd.to_numeric(df["saldo"], errors="coerce").fillna(0.0)
            df = self._apply_date_filter(df)

        self._df = df
        self._update_kpis()
        self._update_charts()
        self._update_table()

        n = len(df)
        self._set_status(
            f"{n} título(s) exibido(s)"
            + (f"  [total na API com esses filtros: {total_api}]"
               if n != total_api else "")
        )
        self._progress["value"] = 100
        self._pct_var.set("100%")

    def _on_error(self, msg: str) -> None:
        self._loading = False
        self._progress.stop()
        self._progress.configure(mode="determinate", value=0)
        self._pct_var.set("")
        self._set_status(f"Erro: {msg}")
        messagebox.showerror("Erro ao consultar a API", msg)

    # ── filtro de datas (client-side) ─────────────────────────────────────────

    def _apply_date_filter(self, df: pd.DataFrame) -> pd.DataFrame:
        vcto_de  = _parse_date_input(self._f_vcto_de.get())
        vcto_ate = _parse_date_input(self._f_vcto_ate.get())
        if vcto_de and len(vcto_de) == 8 and vcto_de.isdigit():
            df = df[df["vencimento"] >= vcto_de]
        if vcto_ate and len(vcto_ate) == 8 and vcto_ate.isdigit():
            df = df[df["vencimento"] <= vcto_ate]
        return df

    # ── KPIs ──────────────────────────────────────────────────────────────────

    def _update_kpis(self) -> None:
        df = self._df
        if df is None or df.empty:
            for v in self._kpi_vars.values():
                v.set("—")
            return

        ab = df[df["status"] == "Em aberto"]
        ve = df[df["status"] == "Vencido"]
        bx = df[df["status"] == "Baixado"]

        self._kpi_vars["total_titulos"].set(f'{len(df):,}'.replace(",", "."))
        self._kpi_vars["valor_total"].set(_brl(df["valor"].sum()))
        self._kpi_vars["saldo_total"].set(_brl(df["saldo"].sum()))
        self._kpi_vars["em_aberto"].set(
            f'{len(ab):,}'.replace(",", ".") + "\n" + _brl(ab["saldo"].sum()))
        self._kpi_vars["vencidos"].set(
            f'{len(ve):,}'.replace(",", ".") + "\n" + _brl(ve["saldo"].sum()))
        self._kpi_vars["baixados"].set(
            f'{len(bx):,}'.replace(",", ".") + "\n" + _brl(bx["valor"].sum()))

    # ── gráficos ──────────────────────────────────────────────────────────────

    def _draw_empty_charts(self) -> None:
        for ax in self._axes.flat:
            ax.clear()
            ax.text(0.5, 0.5, "Sem dados — faça uma consulta",
                    ha="center", va="center", color="#aab7b8", fontsize=9)
            ax.set_facecolor("#f8f9fa")
            ax.axis("off")
        self._canvas.draw()

    def _update_charts(self) -> None:
        for ax in self._axes.flat:
            ax.clear()

        df = self._df
        if df is None or df.empty:
            self._draw_empty_charts()
            return

        ax_status, ax_forn = self._axes[0]
        ax_mes,    ax_fil  = self._axes[1]

        brl_fmt = mticker.FuncFormatter(
            lambda x, _: f"R${x/1000:.0f}k" if abs(x) >= 1000 else f"R${x:.0f}"
        )

        # ── Gráfico 1: pizza — distribuição por status ────────────────────
        status_counts = df["status"].value_counts()
        pie_colors = [STATUS_COLORS.get(s, "#bdc3c7") for s in status_counts.index]
        wedges, texts, autotexts = ax_status.pie(
            status_counts,
            labels=status_counts.index,
            autopct="%1.1f%%",
            colors=pie_colors,
            startangle=90,
            textprops={"fontsize": 8},
            wedgeprops={"edgecolor": "white", "linewidth": 1.5},
        )
        for at in autotexts:
            at.set_fontsize(7.5)
        ax_status.set_title("Distribuição por Status",
                             fontsize=9, fontweight="bold", pad=10)

        # ── Gráfico 2: barras horizontais — Top 10 Fornecedores por Saldo ─
        top10 = (
            df.groupby("fornecedor")["saldo"]
            .sum()
            .nlargest(10)
            .sort_values(ascending=True)
        )
        bar_colors = CHART_COLORS[:len(top10)]
        hbars = ax_forn.barh(
            top10.index, top10.values,
            color=bar_colors[::-1], height=0.65,
            edgecolor="none",
        )
        ax_forn.set_title("Top 10 Fornecedores — Saldo",
                           fontsize=9, fontweight="bold", pad=10)
        ax_forn.set_xlabel("Saldo (R$)", fontsize=7)
        ax_forn.tick_params(axis="y", labelsize=7)
        ax_forn.tick_params(axis="x", labelsize=6)
        ax_forn.xaxis.set_major_formatter(brl_fmt)
        ax_forn.set_facecolor("#fafafa")
        for bar in hbars:
            w = bar.get_width()
            ax_forn.text(
                w * 1.01, bar.get_y() + bar.get_height() / 2,
                _brl(w), va="center", fontsize=6, color="#555",
            )

        # ── Gráfico 3: barras — Saldo por Mês de Vencimento ───────────────
        df_open = df[df["status"] != "Baixado"].copy()
        df_open["mes"] = df_open["vencimento"].str[:6]
        by_month = (
            df_open.groupby("mes")["saldo"]
            .sum()
            .sort_index()
            .tail(14)
        )
        labels_mes = [f"{m[4:6]}/{m[:4]}" for m in by_month.index]
        bar_clrs_mes = [
            "#e74c3c" if m < TODAY[:6] else "#2980b9"
            for m in by_month.index
        ]
        ax_mes.bar(range(len(by_month)), by_month.values,
                   color=bar_clrs_mes, width=0.72, edgecolor="none")
        ax_mes.set_xticks(range(len(by_month)))
        ax_mes.set_xticklabels(labels_mes, rotation=45, ha="right", fontsize=7)
        ax_mes.set_title("Saldo por Mês de Vencimento (em aberto + vencidos)",
                         fontsize=9, fontweight="bold", pad=10)
        ax_mes.set_ylabel("Saldo (R$)", fontsize=7)
        ax_mes.tick_params(axis="y", labelsize=7)
        ax_mes.yaxis.set_major_formatter(brl_fmt)
        ax_mes.set_facecolor("#fafafa")
        ax_mes.legend(handles=[
            Patch(color="#e74c3c", label="Vencido"),
            Patch(color="#2980b9", label="A vencer"),
        ], fontsize=7, loc="upper left", framealpha=0.7)

        # ── Gráfico 4: barras — Saldo por Filial ──────────────────────────
        by_fil = (
            df.groupby("filial")["saldo"]
            .sum()
            .sort_values(ascending=False)
        )
        ax_fil.bar(by_fil.index, by_fil.values,
                   color=CHART_COLORS[:len(by_fil)], width=0.65, edgecolor="none")
        ax_fil.set_title("Saldo Total por Filial",
                         fontsize=9, fontweight="bold", pad=10)
        ax_fil.set_ylabel("Saldo (R$)", fontsize=7)
        ax_fil.tick_params(axis="x", labelsize=8)
        ax_fil.tick_params(axis="y", labelsize=7)
        ax_fil.yaxis.set_major_formatter(brl_fmt)
        ax_fil.set_facecolor("#fafafa")
        for i, (fil, val) in enumerate(by_fil.items()):
            ax_fil.text(i, val * 1.01, _brl(val),
                        ha="center", fontsize=7, color="#555")

        self._canvas.draw()

    # ── tabela ────────────────────────────────────────────────────────────────

    def _update_table(self) -> None:
        for row in self._tree.get_children():
            self._tree.delete(row)

        df = self._df
        if df is None or df.empty:
            return

        for _, row in df.iterrows():
            status = row.get("status", "")
            self._tree.insert("", "end", tags=(status,), values=(
                row.get("rec_no", ""),
                row.get("filial", ""),
                row.get("prefixo", ""),
                row.get("numero", ""),
                row.get("parcela", ""),
                row.get("tipo", ""),
                row.get("fornecedor", ""),
                row.get("loja", ""),
                _fmt_date(row.get("emissao")),
                _fmt_date(row.get("vencimento")),
                _brl(row.get("valor", 0)),
                _brl(row.get("saldo", 0)),
                row.get("moeda", ""),
                row.get("historico", ""),
                status,
            ))

    def _sort_tree(self, col: str) -> None:
        if self._df is None or self._df.empty:
            return
        asc = not self._sort_ascending.get(col, False)
        self._sort_ascending[col] = asc

        if col in ("valor", "saldo", "rec_no"):
            self._df = self._df.sort_values(col, ascending=asc, na_position="last")
        else:
            self._df = self._df.sort_values(
                col, ascending=asc, na_position="last",
                key=lambda s: s.astype(str).str.lower(),
            )
        self._update_table()

    # ── exportar Excel ────────────────────────────────────────────────────────

    def exportar_excel(self) -> None:
        df = self._df
        if df is None or df.empty:
            messagebox.showwarning(
                "Exportar Excel",
                "Nenhum dado carregado. Faça uma consulta primeiro.",
            )
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("Todos os arquivos", "*.*")],
            title="Salvar planilha de Contas a Pagar",
            initialfile=f"contas_pagar_{date.today().strftime('%Y%m%d')}.xlsx",
        )
        if not path:
            return

        try:
            self._set_status("Gerando planilha Excel…")
            self.root.update()
            _write_excel(df, path)
            self._set_status(f"Planilha salva em: {path}")
            messagebox.showinfo("Exportar Excel",
                                f"Arquivo salvo com sucesso:\n{path}")
        except Exception as exc:
            self._set_status("Erro ao exportar.")
            messagebox.showerror("Erro ao exportar", str(exc))

    # ── utilitários ───────────────────────────────────────────────────────────

    def limpar(self) -> None:
        for attr in ("_f_filial", "_f_fornecedor", "_f_prefixo",
                     "_f_numero", "_f_vcto_de", "_f_vcto_ate"):
            getattr(self, attr).set("")
        self._clear_data()
        self._set_status("Filtros limpos.")

    def _clear_data(self) -> None:
        self._df = None
        for v in self._kpi_vars.values():
            v.set("—")
        for row in self._tree.get_children():
            self._tree.delete(row)
        self._draw_empty_charts()
        self._progress["value"] = 0
        self._pct_var.set("")

    def _set_status(self, msg: str, *, indeterminate: bool = False) -> None:
        self._status_var.set(msg)
        if indeterminate:
            self._progress.configure(mode="indeterminate")
            self._progress.start(12)
            self._pct_var.set("")
        else:
            self._progress.stop()
            self._progress.configure(mode="determinate")


# ── exportação Excel (função separada) ───────────────────────────────────────

def _write_excel(df: pd.DataFrame, path: str) -> None:
    import openpyxl
    from openpyxl.styles import (Alignment, Border, Font,
                                 PatternFill, Side)
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── estilos compartilhados ────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1A5276")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    ALT_FILL  = PatternFill("solid", fgColor="EBF5FB")
    VENC_FILL = PatternFill("solid", fgColor="FDEDEC")
    BAIX_FILL = PatternFill("solid", fgColor="F2F3F4")
    THIN      = Side(style="thin", color="BDC3C7")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER    = Alignment(horizontal="center", vertical="center")
    RIGHT     = Alignment(horizontal="right",  vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")
    BRL       = '#,##0.00'
    DATE_COLS = {"emissao", "vencimento_original", "vencimento", "data_baixa"}
    CURR_COLS = {"valor", "saldo"}

    def _hdr(ws, row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = BORDER
        return c

    # ── aba 1: Títulos a Pagar ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Títulos a Pagar"
    ws1.freeze_panes = "A2"

    cols_aba1 = [
        ("Rec No",          "rec_no",              9),
        ("Filial",          "filial",              7),
        ("Prefixo",         "prefixo",             9),
        ("Número",          "numero",             13),
        ("Parcela",         "parcela",             9),
        ("Tipo",            "tipo",                7),
        ("Fornecedor",      "fornecedor",         13),
        ("Loja",            "loja",                7),
        ("Emissão",         "emissao",            13),
        ("Vcto Original",   "vencimento_original", 15),
        ("Vencimento",      "vencimento",         13),
        ("Valor (R$)",      "valor",              16),
        ("Saldo (R$)",      "saldo",              16),
        ("Moeda",           "moeda",               7),
        ("Histórico",       "historico",          42),
        ("Data Baixa",      "data_baixa",         13),
        ("Status",          "status",             12),
    ]

    for ci, (heading, _, w) in enumerate(cols_aba1, 1):
        _hdr(ws1, 1, ci, heading)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[1].height = 18
    ws1.auto_filter.ref = (
        f"A1:{get_column_letter(len(cols_aba1))}1"
    )

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        status = row.get("status", "")
        row_fill = (VENC_FILL if status == "Vencido"
                    else BAIX_FILL if status == "Baixado"
                    else (ALT_FILL if ri % 2 == 0 else None))
        for ci, (_, field, _) in enumerate(cols_aba1, 1):
            val = row.get(field)
            if field in DATE_COLS:
                val = _fmt_date(val)
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if field in CURR_COLS:
                cell.number_format = BRL
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT
            if row_fill:
                cell.fill = row_fill

    # ── aba 2: Resumo ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 22

    ws2.cell(1, 1, "Resumo — Contas a Pagar").font = Font(
        bold=True, color="1A5276", size=14, name="Calibri")
    ws2.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(
        italic=True, size=9, color="7F8C8D", name="Calibri")

    aberto = df[df["status"] == "Em aberto"]
    vencido = df[df["status"] == "Vencido"]
    baixado = df[df["status"] == "Baixado"]

    items_resumo = [
        ("Total de Títulos",         len(df),                       None),
        ("Valor Total (R$)",         float(df["valor"].sum()),      BRL),
        ("Saldo Total (R$)",         float(df["saldo"].sum()),      BRL),
        ("",                          "",                            None),
        ("Em Aberto — Qtd",          len(aberto),                   None),
        ("Em Aberto — Saldo (R$)",   float(aberto["saldo"].sum()),  BRL),
        ("",                          "",                            None),
        ("Vencidos — Qtd",           len(vencido),                  None),
        ("Vencidos — Saldo (R$)",    float(vencido["saldo"].sum()), BRL),
        ("",                          "",                            None),
        ("Baixados — Qtd",           len(baixado),                  None),
        ("Baixados — Valor (R$)",    float(baixado["valor"].sum()), BRL),
    ]
    for i, (lbl, val, fmt) in enumerate(items_resumo, 4):
        c1 = ws2.cell(i, 1, lbl)
        c1.font = Font(bold=bool(lbl), name="Calibri", size=10)
        c2 = ws2.cell(i, 2, val if val != "" else None)
        c2.alignment = RIGHT
        if fmt:
            c2.number_format = fmt

    # ── aba 3: Por Fornecedor ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Por Fornecedor")
    by_forn = (
        df.groupby("fornecedor", as_index=False)
        .agg(qtd=("rec_no", "count"),
             valor_total=("valor", "sum"),
             saldo_total=("saldo", "sum"))
        .sort_values("saldo_total", ascending=False)
    )
    hdrs3 = ["Fornecedor", "Qtd Títulos", "Valor Total (R$)", "Saldo Total (R$)"]
    widths3 = [14, 14, 20, 20]
    for ci, (h, w) in enumerate(zip(hdrs3, widths3), 1):
        _hdr(ws3, 1, ci, h)
        ws3.column_dimensions[get_column_letter(ci)].width = w
    for ri, row in enumerate(by_forn.itertuples(index=False), 2):
        ws3.cell(ri, 1, row.fornecedor)
        ws3.cell(ri, 2, int(row.qtd))
        c3 = ws3.cell(ri, 3, float(row.valor_total))
        c3.number_format = BRL
        c4 = ws3.cell(ri, 4, float(row.saldo_total))
        c4.number_format = BRL
        if ri % 2 == 0:
            for ci in range(1, 5):
                ws3.cell(ri, ci).fill = ALT_FILL

    # ── aba 4: Por Mês de Vencimento ─────────────────────────────────────
    ws4 = wb.create_sheet("Por Vencimento")
    df_open = df[df["status"] != "Baixado"].copy()
    df_open["mes"] = df_open["vencimento"].str[:6]
    by_mes = (
        df_open.groupby("mes", as_index=False)
        .agg(qtd=("rec_no", "count"), saldo=("saldo", "sum"))
        .sort_values("mes")
    )
    hdrs4 = ["Mês/Ano", "Qtd Títulos", "Saldo (R$)", "Situação"]
    widths4 = [12, 14, 20, 12]
    for ci, (h, w) in enumerate(zip(hdrs4, widths4), 1):
        _hdr(ws4, 1, ci, h)
        ws4.column_dimensions[get_column_letter(ci)].width = w
    for ri, row in enumerate(by_mes.itertuples(index=False), 2):
        label = f"{row.mes[4:6]}/{row.mes[:4]}"
        sit = "Vencido" if row.mes < TODAY[:6] else "A vencer"
        ws4.cell(ri, 1, label)
        ws4.cell(ri, 2, int(row.qtd))
        c3 = ws4.cell(ri, 3, float(row.saldo))
        c3.number_format = BRL
        ws4.cell(ri, 4, sit)
        fill = VENC_FILL if sit == "Vencido" else (ALT_FILL if ri % 2 == 0 else None)
        if fill:
            for ci in range(1, 5):
                ws4.cell(ri, ci).fill = fill

    wb.save(path)


# ── entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    ContasAPagarApp(root)
    root.mainloop()
