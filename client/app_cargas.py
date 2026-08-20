"""
Cliente desktop para a RestDataAPI — Relatório de Cargas (endpoint /cargas).

Exibe filtro de data mínima (enviado à API), filtros locais, KPIs, gráficos
analíticos e exporta tudo para Excel.

O endpoint /cargas aplica sempre como regra fixa apenas itens não excluídos
e não cancelados. A data mínima da carga (data_inicial) e o status da carga
(status, campo DAK_ACECAR) são enviados à API como query string — o seletor
de data desta tela já inicia na data atual do sistema (mesmo padrão que a
API assume quando o parâmetro não é enviado) e o usuário pode alterá-lo para
consultar outro período antes de carregar. Os demais filtros (filial,
cliente, caminhão) são aplicados no cliente, após o carregamento completo
dos dados.

Requisitos:  pip install requests matplotlib pandas openpyxl python-dotenv tkcalendar
"""

from __future__ import annotations

import threading
from datetime import date, datetime
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
from typing import Dict, List, Optional, Tuple

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.ticker as mticker
import pandas as pd
from tkcalendar import DateEntry

from api_client import APIClient
from config import API_BASE_URL, API_KEY, API_KEY_NAME


# ── constantes ───────────────────────────────────────────────────────────────

CHART_COLORS = [
    "#2980b9", "#e74c3c", "#27ae60", "#f39c12", "#8e44ad",
    "#16a085", "#d35400", "#2c3e50", "#c0392b", "#1abc9c",
]

# Rótulo exibido -> valor do parâmetro "status" aceito pelo endpoint /cargas
# (campo DAK_ACECAR — customização desta instalação do Protheus, sem lista
# pública de valores; confirmados pelo usuário).
STATUS_CARGA: Dict[str, str] = {
    "Montada":           "1",
    "Disp Conf Gega":    "2",
    "Disp Prest Contas": "3",
    "Disp Prest Títulos":"6",
    "Encerrada":         "7",
    "Juros Pendentes":   "8",
}
STATUS_CARGA_LABELS: Dict[str, str] = {v: k for k, v in STATUS_CARGA.items()}
STATUS_FILTRO_PADRAO = "(Todos)"

# Status considerados "fechados" para o card "Valor em Aberto": tudo que não
# for Encerrada (7) ou Juros Pendentes (8) ainda está em aberto.
STATUS_CARGA_FECHADOS = {"7", "8"}

TREEVIEW_COLS: List[Tuple[str, str, int, str]] = [
    ("filial",        "Filial",      55,  "c"),
    ("codigo",        "Carga",       75,  "c"),
    ("data",          "Data",        85,  "c"),
    ("pedido",        "Pedido",      80,  "c"),
    ("nome_cliente",  "Cliente",    220,  "w"),
    ("nota_fiscal",   "Nota Fiscal", 90,  "c"),
    ("caminhao",      "Caminhão",    90,  "c"),
    ("status_carga",  "Status",     130,  "c"),
    ("peso",          "Peso (kg)",  110,  "e"),
    ("valor",         "Valor (R$)", 120,  "e"),
]


# ── helpers ──────────────────────────────────────────────────────────────────

def _brl(v) -> str:
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"


def _peso(v) -> str:
    try:
        return f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"


def _fmt_date(s) -> str:
    s = str(s or "").strip()
    if len(s) == 8 and s.isdigit():
        return f"{s[6:8]}/{s[4:6]}/{s[:4]}"
    return s or "—"


def _fmt_status(v) -> str:
    s = str(v or "").strip()
    return STATUS_CARGA_LABELS.get(s, s or "—")


# Placa usada para indicar que o próprio cliente retirou a carga (sem
# caminhão terceirizado vinculado) — exibida como "Cliente" em vez da placa.
CAMINHAO_CLIENTE = "KHA0902"


def _fmt_caminhao(v, vazio: str = "—") -> str:
    s = str(v or "").strip()
    if s == CAMINHAO_CLIENTE:
        return "Cliente"
    return s or vazio


def _n_cargas(df: Optional[pd.DataFrame]) -> int:
    """Quantidade de cargas distintas (chave filial + codigo) em df.

    Cada linha retornada por /cargas é um ITEM de uma carga (uma carga pode
    ter vários pedidos e vários itens); não é uma carga por si só.
    """
    if df is None or df.empty:
        return 0
    return int(df["_carga_key"].nunique())


# ── aplicação principal ───────────────────────────────────────────────────────

class CargasApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Relatório de Cargas — Grupo LGH (sem o Broker)")
        self.root.geometry("1280x820")
        self.root.minsize(960, 620)
        self.root.configure(bg="#f0f2f5")

        self._df_raw: Optional[pd.DataFrame] = None
        self._df: Optional[pd.DataFrame] = None
        self._loading = False
        self._sort_ascending: Dict[str, bool] = {}

        self._setup_styles()
        self._build_ui()

    # ── estilos ──────────────────────────────────────────────────────────────

    def _setup_styles(self) -> None:
        s = ttk.Style(self.root)
        s.theme_use("clam")

        BG, WHITE, PRIMARY, ACCENT = "#f0f2f5", "#ffffff", "#1a5276", "#2980b9"

        s.configure("TFrame", background=BG)
        s.configure("White.TFrame", background=WHITE)
        s.configure("TLabel", background=BG, font=("Segoe UI", 10))
        s.configure("TLabelframe", background=BG, font=("Segoe UI", 10, "bold"))
        s.configure("TLabelframe.Label", background=BG, font=("Segoe UI", 10, "bold"),
                    foreground=PRIMARY)

        for name, bg, fg, hover in [
            ("Primary",   PRIMARY,   WHITE, ACCENT),
            ("Secondary", "#7f8c8d", WHITE, "#95a5a6"),
            ("Success",   "#27ae60", WHITE, "#2ecc71"),
        ]:
            s.configure(f"{name}.TButton",
                        background=bg, foreground=fg,
                        font=("Segoe UI", 10, "bold"), padding=(12, 5))
            s.map(f"{name}.TButton", background=[("active", hover), ("pressed", bg)])

        s.configure("TEntry", fieldbackground=WHITE, padding=4)
        s.configure("TNotebook", background=BG)
        s.configure("TNotebook.Tab", padding=(12, 6), font=("Segoe UI", 10))

        s.configure("Treeview", background=WHITE, fieldbackground=WHITE,
                    font=("Segoe UI", 10), rowheight=25)
        s.configure("Treeview.Heading",
                    font=("Segoe UI", 10, "bold"),
                    background=PRIMARY, foreground=WHITE)
        s.map("Treeview",
              background=[("selected", ACCENT)],
              foreground=[("selected", WHITE)])

    # ── construção da UI ─────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        self._build_conn_frame()
        self._build_periodo_frame()
        self._build_filter_frame()
        self._build_kpi_frame()
        self._build_toolbar()
        self._build_notebook()
        self._build_statusbar()

    # ── conexão ──────────────────────────────────────────────────────────────

    def _build_conn_frame(self) -> None:
        f = ttk.LabelFrame(self.root, text="Conexão", padding=(10, 6))
        f.pack(fill="x", padx=12, pady=(8, 3))

        ttk.Label(f, text="URL da API:").grid(row=0, column=0, sticky="w", padx=(0, 4))
        self._url_var = tk.StringVar(value=API_BASE_URL)
        ttk.Entry(f, textvariable=self._url_var, width=36).grid(
            row=0, column=1, sticky="ew", padx=(0, 18))

        ttk.Label(f, text="API Key:").grid(row=0, column=2, sticky="w", padx=(0, 4))
        self._key_var = tk.StringVar(value=API_KEY)
        ttk.Entry(f, textvariable=self._key_var, show="•", width=34).grid(
            row=0, column=3, sticky="ew", padx=(0, 12))

        ttk.Button(f, text="Testar conexão", style="Secondary.TButton",
                   command=self._test_connection).grid(row=0, column=4)

        f.columnconfigure(1, weight=1)
        f.columnconfigure(3, weight=1)

    # ── período (define o data_inicial enviado à API) ──────────────────────────

    def _build_periodo_frame(self) -> None:
        f = ttk.LabelFrame(
            self.root,
            text="Período e status (define data_inicial/data_final/status enviados à API — recarrega ao consultar)",
            padding=(10, 6),
        )
        f.pack(fill="x", padx=12, pady=3)

        ttk.Label(f, text="Data mínima de:").grid(
            row=0, column=0, sticky="w", padx=(0, 6))

        self._data_inicial_var = tk.StringVar()
        self._de_data_inicial = DateEntry(
            f, textvariable=self._data_inicial_var, width=12,
            date_pattern="dd/mm/yyyy",
            background="#1a5276", foreground="white", borderwidth=1,
        )
        self._de_data_inicial.set_date(date.today())
        self._de_data_inicial.grid(row=0, column=1, sticky="w")

        ttk.Label(f, text="Data mínima até:").grid(
            row=0, column=2, sticky="w", padx=(14, 6))

        self._data_final_var = tk.StringVar()
        self._de_data_final = DateEntry(
            f, textvariable=self._data_final_var, width=12,
            date_pattern="dd/mm/yyyy",
            background="#1a5276", foreground="white", borderwidth=1,
        )
        self._de_data_final.set_date(date.today())
        self._de_data_final.grid(row=0, column=3, sticky="w")

        ttk.Label(f, text="Status:").grid(row=0, column=4, sticky="w", padx=(14, 3))
        self._f_status = tk.StringVar(value=STATUS_FILTRO_PADRAO)
        ttk.Combobox(
            f, textvariable=self._f_status, width=16, state="readonly",
            values=[STATUS_FILTRO_PADRAO, *STATUS_CARGA.keys()],
        ).grid(row=0, column=5, sticky="w")

        ttk.Label(
            f,
            text="(por padrão, o período já inicia e termina hoje; amplie as datas para consultar outro período)",
            foreground="#7f8c8d",
        ).grid(row=0, column=6, sticky="w", padx=(10, 0))

    # ── filtros (aplicados no cliente, após o carregamento) ────────────────────

    def _build_filter_frame(self) -> None:
        f = ttk.LabelFrame(
            self.root,
            text="Filtros (aplicados localmente sobre os dados carregados)",
            padding=(10, 6),
        )
        f.pack(fill="x", padx=12, pady=3)

        text_filters = [
            ("Filial:",   "_f_filial",   6),
            ("Cliente:",  "_f_cliente", 26),
            ("Caminhão:", "_f_caminhao", 12),
        ]
        col = 0
        for label, attr, width in text_filters:
            ttk.Label(f, text=label).grid(
                row=0, column=col, sticky="w", padx=(0 if col == 0 else 14, 3))
            var = tk.StringVar()
            setattr(self, attr, var)
            ttk.Entry(f, textvariable=var, width=width).grid(
                row=0, column=col + 1, sticky="ew")
            col += 2

        btn = ttk.Frame(f)
        btn.grid(row=0, column=col, sticky="e", padx=(20, 0))
        ttk.Button(btn, text=" Carregar do Protheus ", style="Primary.TButton",
                   command=self.consultar).pack(side="left", padx=(0, 6))
        ttk.Button(btn, text="Aplicar filtros", style="Secondary.TButton",
                   command=self._on_filtros_alterados).pack(side="left", padx=(0, 6))
        ttk.Button(btn, text="Limpar", style="Secondary.TButton",
                   command=self.limpar).pack(side="left")

        f.columnconfigure(col, weight=1)

    # ── KPI cards ─────────────────────────────────────────────────────────────

    def _build_kpi_frame(self) -> None:
        outer = tk.Frame(self.root, bg="#f0f2f5")
        outer.pack(fill="x", padx=12, pady=3)

        self._kpi_vars: Dict[str, tk.StringVar] = {}
        cards_cfg = [
            ("total_cargas", "Total de Cargas",   "#2c3e50"),
            ("clientes",     "Clientes",          "#7f8c8d"),
            ("peso_total",   "Peso Total (kg)",   "#1a5276"),
            ("valor_total",  "Valor Total",       "#154360"),
            ("valor_aberto", "Valor em Aberto",   "#ca6f1e"),
            ("valor_medio",  "Valor Médio/Carga", "#1e8449"),
        ]

        for i, (key, label, fg) in enumerate(cards_cfg):
            card = tk.Frame(outer, bg="#ffffff",
                            highlightbackground="#d5d8dc", highlightthickness=1)
            card.grid(row=0, column=i, sticky="nsew", padx=4, pady=2, ipady=10)

            tk.Label(card, text=label, bg="#ffffff",
                     font=("Segoe UI", 9), fg="#7f8c8d").pack(pady=(4, 0))

            var = tk.StringVar(value="—")
            self._kpi_vars[key] = var
            tk.Label(card, textvariable=var, bg="#ffffff",
                     font=("Segoe UI", 14, "bold"), fg=fg).pack()

        for i in range(len(cards_cfg)):
            outer.columnconfigure(i, weight=1)

    # ── toolbar acima do notebook ─────────────────────────────────────────────

    def _build_toolbar(self) -> None:
        bar = ttk.Frame(self.root)
        bar.pack(fill="x", padx=12, pady=(4, 0))
        ttk.Button(bar, text="  Exportar Excel  ", style="Success.TButton",
                   command=self.exportar_excel).pack(side="right")

    # ── notebook ──────────────────────────────────────────────────────────────

    def _build_notebook(self) -> None:
        self._nb = ttk.Notebook(self.root)
        self._nb.pack(fill="both", expand=True, padx=12, pady=(2, 4))

        tab_graficos = ttk.Frame(self._nb)
        self._nb.add(tab_graficos, text="  Gráficos  ")
        self._build_charts_tab(tab_graficos)

        tab_dados = ttk.Frame(self._nb)
        self._nb.add(tab_dados, text="  Dados  ")
        self._build_data_tab(tab_dados)

    def _build_charts_tab(self, parent: ttk.Frame) -> None:
        self._fig = Figure(figsize=(13, 6.4), dpi=100, facecolor="#f0f2f5")
        self._axes = self._fig.subplots(2, 2)
        self._fig.subplots_adjust(left=0.12, right=0.97, top=0.9,
                                  bottom=0.18, hspace=0.65, wspace=0.42)
        self._canvas = FigureCanvasTkAgg(self._fig, master=parent)
        self._canvas.get_tk_widget().pack(fill="both", expand=True, padx=4, pady=4)
        self._draw_empty_charts()

    def _build_data_tab(self, parent: ttk.Frame) -> None:
        cols = [c[0] for c in TREEVIEW_COLS]
        self._tree = ttk.Treeview(parent, columns=cols,
                                  show="headings", selectmode="browse")

        for col_id, heading, width, anchor in TREEVIEW_COLS:
            self._tree.heading(
                col_id, text=heading,
                command=lambda c=col_id: self._sort_tree(c),
            )
            self._tree.column(col_id, width=width, anchor=anchor,
                              stretch=(col_id == "nome_cliente"))

        vsb = ttk.Scrollbar(parent, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        parent.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)

    # ── barra de status ───────────────────────────────────────────────────────

    def _build_statusbar(self) -> None:
        bar = tk.Frame(self.root, bg="#dde1e7", height=24)
        bar.pack(fill="x", side="bottom")

        self._status_var = tk.StringVar(value="Pronto.")
        tk.Label(bar, textvariable=self._status_var,
                 bg="#dde1e7", font=("Segoe UI", 9), anchor="w").pack(
            side="left", padx=8, pady=2)

        self._pct_var = tk.StringVar()
        tk.Label(bar, textvariable=self._pct_var,
                 bg="#dde1e7", font=("Segoe UI", 9)).pack(side="right", padx=4)

        self._progress = ttk.Progressbar(bar, mode="determinate", length=200)
        self._progress.pack(side="right", padx=8, pady=3)

    # ── conexão ───────────────────────────────────────────────────────────────

    def _make_client(self) -> APIClient:
        return APIClient(
            base_url=self._url_var.get().strip(),
            api_key=self._key_var.get().strip(),
            api_key_name=API_KEY_NAME,
        )

    def _test_connection(self) -> None:
        try:
            self._make_client().health()
            messagebox.showinfo("Conexão OK",
                                "A API respondeu com sucesso!")
        except Exception as exc:
            messagebox.showerror("Falha na conexão", str(exc))

    # ── consulta (thread) ─────────────────────────────────────────────────────

    def consultar(self) -> None:
        if self._loading:
            return

        # DateEntry sempre contém uma data válida; converte DD/MM/AAAA -> AAAAMMDD.
        data_inicial = self._de_data_inicial.get_date().strftime("%Y%m%d")
        data_final = self._de_data_final.get_date().strftime("%Y%m%d")
        if data_final < data_inicial:
            messagebox.showwarning(
                "Período inválido",
                '"Data mínima até" não pode ser anterior a "Data mínima de".',
            )
            return

        self._loading = True
        self._clear_data()

        status_label = self._f_status.get()
        status = STATUS_CARGA.get(status_label)

        periodo = f"de {self._data_inicial_var.get()} até {self._data_final_var.get()}"
        if status:
            periodo += f", status={status_label}"
        self._set_status(f"Consultando /cargas na API… [{periodo}]", indeterminate=True)
        threading.Thread(
            target=self._fetch_thread, args=(data_inicial, data_final, status), daemon=True
        ).start()

    def _fetch_thread(self, data_inicial: str, data_final: str, status: Optional[str]) -> None:
        try:
            client = self._make_client()
            items, total = client.get_all_cargas(
                data_inicial=data_inicial, data_final=data_final, status=status,
                progress_callback=self._on_progress)
            self.root.after(0, self._on_data_ready, items, total)
        except Exception as exc:
            self.root.after(0, self._on_error, str(exc))

    def _on_progress(self, loaded: int, total: int) -> None:
        pct = int(100 * loaded / total) if total else 0
        self.root.after(0, self._update_progress_bar, loaded, total, pct)

    def _update_progress_bar(self, loaded: int, total: int, pct: int) -> None:
        self._progress.stop()
        self._progress.configure(mode="determinate", value=pct)
        self._pct_var.set(f"{pct}%  ({loaded}/{total})")
        self._status_var.set(f"Carregando… {loaded} de {total} registros")

    def _on_data_ready(self, items: List[Dict], total_api: int) -> None:
        self._loading = False

        df = pd.DataFrame(items) if items else pd.DataFrame()
        if not df.empty:
            df["peso"] = pd.to_numeric(df["peso"], errors="coerce").fillna(0.0)
            df["valor"] = pd.to_numeric(df["valor"], errors="coerce").fillna(0.0)
            # Uma carga é identificada por filial + codigo (DAI_FILIAL + DAI_COD);
            # cada linha da API é um item dessa carga, não uma carga distinta.
            df["_carga_key"] = df["filial"].astype(str) + "-" + df["codigo"].astype(str)

        self._df_raw = df
        self._apply_filters_and_refresh()

        n = _n_cargas(self._df)
        self._set_status(
            f"{n} carga(s) distinta(s) exibida(s)  [{total_api} item(ns) carregado(s) da API]"
        )
        self._progress["value"] = 100
        self._pct_var.set("100%")

    def _on_error(self, msg: str) -> None:
        self._loading = False
        self._progress.stop()
        self._progress.configure(mode="determinate", value=0)
        self._pct_var.set("")
        self._set_status(f"Erro: {msg}")
        messagebox.showerror("Erro ao consultar a API", msg)

    # ── filtros (client-side) ────────────────────────────────────────────────

    def _on_filtros_alterados(self) -> None:
        if self._df_raw is None:
            messagebox.showinfo(
                "Filtros", "Carregue os dados da API antes de aplicar filtros.")
            return
        self._apply_filters_and_refresh()
        n = _n_cargas(self._df)
        self._set_status(f"{n} carga(s) exibida(s) após filtros")

    def _apply_filters_and_refresh(self) -> None:
        df = self._df_raw
        if df is None or df.empty:
            self._df = df
        else:
            self._df = self._apply_filters(df)
        self._update_kpis()
        self._update_charts()
        self._update_table()

    def _apply_filters(self, df: pd.DataFrame) -> pd.DataFrame:
        filial = self._f_filial.get().strip()
        if filial:
            df = df[df["filial"] == filial]

        cliente = self._f_cliente.get().strip().lower()
        if cliente:
            df = df[df["nome_cliente"].fillna("").str.lower().str.contains(cliente)]

        caminhao = self._f_caminhao.get().strip().lower()
        if caminhao:
            labels = df["caminhao"].apply(_fmt_caminhao).str.lower()
            df = df[labels.str.contains(caminhao)]

        return df

    # ── KPIs ──────────────────────────────────────────────────────────────────

    def _update_kpis(self) -> None:
        df = self._df
        if df is None or df.empty:
            for v in self._kpi_vars.values():
                v.set("—")
            return

        valor_total = float(df["valor"].sum())
        n_cargas = df["_carga_key"].nunique()
        valor_medio = valor_total / n_cargas if n_cargas else 0.0

        status_col = df["status_carga"].astype(str).str.strip()
        valor_aberto = float(df.loc[~status_col.isin(STATUS_CARGA_FECHADOS), "valor"].sum())

        self._kpi_vars["total_cargas"].set(f'{n_cargas:,}'.replace(",", "."))
        self._kpi_vars["clientes"].set(f'{df["nome_cliente"].nunique():,}'.replace(",", "."))
        self._kpi_vars["peso_total"].set(_peso(df["peso"].sum()))
        self._kpi_vars["valor_total"].set(_brl(valor_total))
        self._kpi_vars["valor_aberto"].set(_brl(valor_aberto))
        self._kpi_vars["valor_medio"].set(_brl(valor_medio))

    # ── gráficos ──────────────────────────────────────────────────────────────

    def _draw_empty_charts(self) -> None:
        for ax in self._axes.flat:
            ax.clear()
            ax.text(0.5, 0.5, "Sem dados — clique em \"Carregar do Protheus\"",
                    ha="center", va="center", color="#aab7b8", fontsize=10.5)
            ax.set_facecolor("#f8f9fa")
            ax.axis("off")
        self._canvas.draw()

    def _update_charts(self) -> None:
        for ax in self._axes.flat:
            ax.clear()

        df = self._df
        if df is None or df.empty:
            self._draw_empty_charts()
            return

        ax_top_cliente, ax_filial_valor = self._axes[0]
        ax_evolucao,    ax_top_caminhao = self._axes[1]

        brl_fmt = mticker.FuncFormatter(
            lambda x, _: f"R${x/1000:.0f}k" if abs(x) >= 1000 else f"R${x:.0f}"
        )

        # ── Gráfico 1: barras horizontais — Top 10 Clientes por Valor ─────
        top_clientes = (
            df.assign(nome_cliente=df["nome_cliente"].fillna("(sem nome)"))
            .groupby("nome_cliente")["valor"]
            .sum()
            .nlargest(10)
            .sort_values(ascending=True)
        )
        bar_colors = CHART_COLORS[:len(top_clientes)]
        hbars = ax_top_cliente.barh(
            top_clientes.index, top_clientes.values,
            color=bar_colors[::-1], height=0.65, edgecolor="none",
        )
        ax_top_cliente.set_title("Top 10 Clientes — Valor de Frete",
                                 fontsize=13, fontweight="bold", pad=12)
        ax_top_cliente.set_xlabel("Valor (R$)", fontsize=10.5)
        ax_top_cliente.tick_params(axis="y", labelsize=9)
        ax_top_cliente.tick_params(axis="x", labelsize=9)
        ax_top_cliente.xaxis.set_major_formatter(brl_fmt)
        ax_top_cliente.set_facecolor("#fafafa")
        ax_top_cliente.margins(x=0.2)
        for bar in hbars:
            w = bar.get_width()
            ax_top_cliente.text(
                w * 1.01, bar.get_y() + bar.get_height() / 2,
                _brl(w), va="center", fontsize=9, color="#555",
            )

        # ── Gráfico 2: pizza — Valor Total por Filial ─────────────────────
        by_filial_valor = df.groupby("filial")["valor"].sum().sort_values(ascending=False)
        pie_colors = CHART_COLORS[:len(by_filial_valor)]
        wedges, texts, autotexts = ax_filial_valor.pie(
            by_filial_valor,
            labels=by_filial_valor.index,
            autopct="%1.1f%%",
            colors=pie_colors,
            startangle=90,
            textprops={"fontsize": 9.5},
            wedgeprops={"edgecolor": "white", "linewidth": 1.5},
        )
        for at in autotexts:
            at.set_fontsize(9)
        ax_filial_valor.set_title("Valor Total por Filial",
                                  fontsize=13, fontweight="bold", pad=12)

        # ── Gráfico 3: barras — Nº de Cargas por Data ──────────────────────
        # Conta cargas distintas (filial + codigo) por data, não linhas/itens.
        by_data = df.groupby("data")["_carga_key"].nunique().sort_index().tail(12)
        labels_data = [_fmt_date(d) for d in by_data.index]
        ax_evolucao.bar(range(len(by_data)), by_data.values,
                        color="#2980b9", width=0.7, edgecolor="none")
        ax_evolucao.set_xticks(range(len(by_data)))
        ax_evolucao.set_xticklabels(labels_data, rotation=45, ha="right", fontsize=9)
        ax_evolucao.set_title("Nº de Cargas por Data (últimas 12 datas)",
                              fontsize=13, fontweight="bold", pad=12)
        ax_evolucao.set_ylabel("Cargas", fontsize=10.5)
        ax_evolucao.tick_params(axis="y", labelsize=9.5)
        ax_evolucao.set_facecolor("#fafafa")

        # ── Gráfico 4: barras horizontais — Top 10 Caminhões por Peso ─────
        top_caminhoes = (
            df.assign(caminhao=df["caminhao"].apply(lambda v: _fmt_caminhao(v, "(não informado)")))
            .groupby("caminhao")["peso"]
            .sum()
            .nlargest(10)
            .sort_values(ascending=True)
        )
        bar_colors_cam = CHART_COLORS[:len(top_caminhoes)]
        hbars_cam = ax_top_caminhao.barh(
            top_caminhoes.index, top_caminhoes.values,
            color=bar_colors_cam[::-1], height=0.65, edgecolor="none",
        )
        ax_top_caminhao.set_title("Top 10 Caminhões — Peso Transportado",
                                  fontsize=13, fontweight="bold", pad=12)
        ax_top_caminhao.set_xlabel("Peso (kg)", fontsize=10.5)
        ax_top_caminhao.tick_params(axis="y", labelsize=9)
        ax_top_caminhao.tick_params(axis="x", labelsize=9)
        ax_top_caminhao.set_facecolor("#fafafa")
        ax_top_caminhao.margins(x=0.2)
        for bar in hbars_cam:
            w = bar.get_width()
            ax_top_caminhao.text(
                w * 1.01, bar.get_y() + bar.get_height() / 2,
                _peso(w), va="center", fontsize=9, color="#555",
            )

        self._canvas.draw()

    # ── tabela ────────────────────────────────────────────────────────────────

    def _update_table(self) -> None:
        for row in self._tree.get_children():
            self._tree.delete(row)

        df = self._df
        if df is None or df.empty:
            return

        for _, row in df.iterrows():
            self._tree.insert("", "end", values=(
                row.get("filial", ""),
                row.get("codigo", ""),
                _fmt_date(row.get("data")),
                row.get("pedido", ""),
                row.get("nome_cliente", ""),
                row.get("nota_fiscal", ""),
                _fmt_caminhao(row.get("caminhao", "")),
                _fmt_status(row.get("status_carga", "")),
                _peso(row.get("peso", 0)),
                _brl(row.get("valor", 0)),
            ))

    def _sort_tree(self, col: str) -> None:
        if self._df is None or self._df.empty:
            return
        asc = not self._sort_ascending.get(col, False)
        self._sort_ascending[col] = asc

        if col in ("peso", "valor"):
            self._df = self._df.sort_values(col, ascending=asc, na_position="last")
        else:
            self._df = self._df.sort_values(
                col, ascending=asc, na_position="last",
                key=lambda s: s.astype(str).str.lower(),
            )
        self._update_table()

    # ── exportar Excel ────────────────────────────────────────────────────────

    def exportar_excel(self) -> None:
        df = self._df
        if df is None or df.empty:
            messagebox.showwarning(
                "Exportar Excel",
                "Nenhum dado carregado. Faça uma consulta primeiro.",
            )
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("Todos os arquivos", "*.*")],
            title="Salvar planilha de Cargas",
            initialfile=f"cargas_{date.today().strftime('%Y%m%d')}.xlsx",
        )
        if not path:
            return

        try:
            self._set_status("Gerando planilha Excel…")
            self.root.update()
            _write_excel(df, path)
            self._set_status(f"Planilha salva em: {path}")
            messagebox.showinfo("Exportar Excel",
                                f"Arquivo salvo com sucesso:\n{path}")
        except Exception as exc:
            self._set_status("Erro ao exportar.")
            messagebox.showerror("Erro ao exportar", str(exc))

    # ── utilitários ───────────────────────────────────────────────────────────

    def limpar(self) -> None:
        for attr in ("_f_filial", "_f_cliente", "_f_caminhao"):
            getattr(self, attr).set("")
        if self._df_raw is not None:
            self._apply_filters_and_refresh()
        self._set_status("Filtros limpos.")

    def _clear_data(self) -> None:
        self._df_raw = None
        self._df = None
        for v in self._kpi_vars.values():
            v.set("—")
        for row in self._tree.get_children():
            self._tree.delete(row)
        self._draw_empty_charts()
        self._progress["value"] = 0
        self._pct_var.set("")

    def _set_status(self, msg: str, *, indeterminate: bool = False) -> None:
        self._status_var.set(msg)
        if indeterminate:
            self._progress.configure(mode="indeterminate")
            self._progress.start(12)
            self._pct_var.set("")
        else:
            self._progress.stop()
            self._progress.configure(mode="determinate")


# ── exportação Excel (função separada) ───────────────────────────────────────

def _write_excel(df: pd.DataFrame, path: str) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    HDR_FILL  = PatternFill("solid", fgColor="1A5276")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    ALT_FILL  = PatternFill("solid", fgColor="EBF5FB")
    THIN      = Side(style="thin", color="BDC3C7")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER    = Alignment(horizontal="center", vertical="center")
    RIGHT     = Alignment(horizontal="right",  vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")
    BRL       = '#,##0.00'
    PESO_FMT  = '#,##0.00'
    PCT       = '0.0%'
    BOLD10    = Font(bold=True, name="Calibri", size=10)

    def _hdr(ws, row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = BORDER
        return c

    # ── aba 1: Cargas ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Cargas"
    ws1.freeze_panes = "A2"

    cols_aba1 = [
        ("Filial",        "filial",        9),
        ("Carga",         "codigo",       12),
        ("Data",          "data",         13),
        ("Pedido",        "pedido",       13),
        ("Cliente",       "nome_cliente", 34),
        ("Nota Fiscal",   "nota_fiscal",  14),
        ("Caminhão",      "caminhao",     14),
        ("Status",        "status_carga", 18),
        ("Peso (kg)",     "peso",         14),
        ("Valor (R$)",    "valor",        16),
    ]
    for ci, (heading, _, w) in enumerate(cols_aba1, 1):
        _hdr(ws1, 1, ci, heading)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[1].height = 18
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(cols_aba1))}1"

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        row_fill = ALT_FILL if ri % 2 == 0 else None
        for ci, (_, field, _) in enumerate(cols_aba1, 1):
            val = row.get(field)
            if field == "data":
                val = _fmt_date(val)
            elif field == "status_carga":
                val = _fmt_status(val)
            elif field == "caminhao":
                val = _fmt_caminhao(val)
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if field == "valor":
                cell.number_format = BRL
                cell.alignment = RIGHT
            elif field == "peso":
                cell.number_format = PESO_FMT
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT
            if row_fill:
                cell.fill = row_fill

    # ── aba 2: Resumo ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 22

    ws2.cell(1, 1, "Resumo — Cargas").font = Font(
        bold=True, color="1A5276", size=14, name="Calibri")
    ws2.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(
        italic=True, size=9, color="7F8C8D", name="Calibri")

    valor_total = float(df["valor"].sum())
    n_cargas = int(df["_carga_key"].nunique())
    valor_medio = valor_total / n_cargas if n_cargas else 0.0

    status_col = df["status_carga"].astype(str).str.strip()
    valor_aberto = float(df.loc[~status_col.isin(STATUS_CARGA_FECHADOS), "valor"].sum())

    items_resumo = [
        ("Total de Cargas",           n_cargas,                       None),
        ("Clientes Distintos",        int(df["nome_cliente"].nunique()), None),
        ("Peso Total (kg)",           float(df["peso"].sum()),        PESO_FMT),
        ("Valor Total (R$)",          valor_total,                    BRL),
        ("Valor em Aberto (R$)",      valor_aberto,                   BRL),
        ("Valor Médio por Carga (R$)", valor_medio,                   BRL),
    ]
    for i, (lbl, val, fmt) in enumerate(items_resumo, 4):
        c1 = ws2.cell(i, 1, lbl)
        c1.font = Font(bold=True, name="Calibri", size=10)
        c2 = ws2.cell(i, 2, val)
        c2.alignment = RIGHT
        if fmt:
            c2.number_format = fmt

    # ── aba 3: Por Filial ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Por Filial")
    by_filial = (
        df.groupby("filial", as_index=False)
        .agg(qtd_cargas=("_carga_key", "nunique"),
             peso_total=("peso", "sum"),
             valor_total=("valor", "sum"))
        .sort_values("valor_total", ascending=False)
    )
    total_geral_filial = float(by_filial["valor_total"].sum())

    hdrs3 = ["Filial", "Qtd. Cargas", "Peso Total (kg)", "Valor Total (R$)", "% do Total"]
    widths3 = [12, 12, 18, 18, 12]
    for ci, (h, w) in enumerate(zip(hdrs3, widths3), 1):
        _hdr(ws3, 1, ci, h)
        ws3.column_dimensions[get_column_letter(ci)].width = w

    ri = 2
    for row in by_filial.itertuples(index=False):
        pct = (row.valor_total / total_geral_filial) if total_geral_filial else 0.0
        ws3.cell(ri, 1, row.filial)
        ws3.cell(ri, 2, int(row.qtd_cargas))
        c3 = ws3.cell(ri, 3, float(row.peso_total)); c3.number_format = PESO_FMT
        c4 = ws3.cell(ri, 4, float(row.valor_total)); c4.number_format = BRL
        c5 = ws3.cell(ri, 5, pct); c5.number_format = PCT
        if ri % 2 == 0:
            for ci in range(1, 6):
                ws3.cell(ri, ci).fill = ALT_FILL
        ri += 1

    ws3.cell(ri, 1, "TOTAL GERAL").font = BOLD10
    ws3.cell(ri, 2, int(by_filial["qtd_cargas"].sum())).font = BOLD10
    c3 = ws3.cell(ri, 3, float(by_filial["peso_total"].sum())); c3.number_format = PESO_FMT; c3.font = BOLD10
    c4 = ws3.cell(ri, 4, total_geral_filial); c4.number_format = BRL; c4.font = BOLD10
    c5 = ws3.cell(ri, 5, 1.0); c5.number_format = PCT; c5.font = BOLD10

    # ── aba 4: Top Clientes ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Top Clientes")
    by_cliente = (
        df.assign(nome_cliente=df["nome_cliente"].fillna("(sem nome)"))
        .groupby("nome_cliente", as_index=False)
        .agg(qtd_cargas=("_carga_key", "nunique"),
             peso_total=("peso", "sum"),
             valor_total=("valor", "sum"))
        .sort_values("valor_total", ascending=False)
        .head(50)
    )
    hdrs4 = ["Cliente", "Qtd. Cargas", "Peso Total (kg)", "Valor Total (R$)"]
    widths4 = [34, 12, 18, 18]
    for ci, (h, w) in enumerate(zip(hdrs4, widths4), 1):
        _hdr(ws4, 1, ci, h)
        ws4.column_dimensions[get_column_letter(ci)].width = w
    for ri, row in enumerate(by_cliente.itertuples(index=False), 2):
        ws4.cell(ri, 1, row.nome_cliente)
        ws4.cell(ri, 2, int(row.qtd_cargas))
        c3 = ws4.cell(ri, 3, float(row.peso_total)); c3.number_format = PESO_FMT
        c4 = ws4.cell(ri, 4, float(row.valor_total)); c4.number_format = BRL
        if ri % 2 == 0:
            for ci in range(1, 5):
                ws4.cell(ri, ci).fill = ALT_FILL

    wb.save(path)


# ── entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    CargasApp(root)
    root.mainloop()
