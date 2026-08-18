"""
Cliente desktop para a RestDataAPI — Relatório de Saldo de Estoque
(endpoint /saldos-estoque).

Exibe seleção de tipo de estoque, filtros, KPIs, gráficos analíticos e
exporta para Excel com quatro abas (dados, resumo, por filial, top produtos).

O endpoint /saldos-estoque sempre aplica como regra fixa apenas saldo
positivo e registros não excluídos. Tipo de produto e armazém são
parametrizáveis (tipo_produto/local) e, nesta tela, resolvidos pelo seletor
"Tipo de Estoque" (Produtos Acabados: PA/armazém 01, ou Vasilhames:
AM/armazém 20) — a escolha vale para a próxima consulta à API. Os demais
filtros (filial, código do produto e descrição) são aplicados no cliente,
após o carregamento completo dos dados.

Requisitos:  pip install requests matplotlib pandas openpyxl python-dotenv
"""

from __future__ import annotations

import threading
from datetime import date, datetime
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
from typing import Dict, List, Optional, Tuple

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.ticker as mticker
import pandas as pd

from api_client import APIClient
from config import API_BASE_URL, API_KEY, API_KEY_NAME


# ── constantes ───────────────────────────────────────────────────────────────

# Rótulo exibido -> (tipo_produto, local) enviados como query string à API.
TIPO_ESTOQUE_OPCOES: Dict[str, Tuple[str, str]] = {
    "Produtos Acabados  (Tipo PA — Armazém 01)": ("PA", "01"),
    "Vasilhames  (Tipo AM — Armazém 20)":         ("AM", "20"),
}
TIPO_ESTOQUE_PADRAO = next(iter(TIPO_ESTOQUE_OPCOES))

CHART_COLORS = [
    "#2980b9", "#e74c3c", "#27ae60", "#f39c12", "#8e44ad",
    "#16a085", "#d35400", "#2c3e50", "#c0392b", "#1abc9c",
]

TREEVIEW_COLS: List[Tuple[str, str, int, str]] = [
    ("filial",            "Filial",             60,  "c"),
    ("local",              "Local",              60,  "c"),
    ("codigo_produto",    "Código",             110,  "c"),
    ("descricao_produto", "Descrição",          340,  "w"),
    ("quantidade",        "Quantidade",         120,  "e"),
    ("valor_atual",       "Valor Atual (R$)",   140,  "e"),
]


# ── helpers ──────────────────────────────────────────────────────────────────

def _brl(v) -> str:
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"


def _qtd(v) -> str:
    try:
        s = f"{float(v):,.4f}"
        if "." in s:
            s = s.rstrip("0").rstrip(".")
        return s.replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"


def _qtd2(v) -> str:
    try:
        return f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"


# ── aplicação principal ───────────────────────────────────────────────────────

class EstoqueApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Relatório de Saldo de Estoque — Grupo LGH (sem o Broker)")
        self.root.geometry("1280x820")
        self.root.minsize(960, 620)
        self.root.configure(bg="#f0f2f5")

        self._df_raw: Optional[pd.DataFrame] = None
        self._df: Optional[pd.DataFrame] = None
        self._loading = False
        self._sort_ascending: Dict[str, bool] = {}

        self._setup_styles()
        self._build_ui()

    # ── estilos ──────────────────────────────────────────────────────────────

    def _setup_styles(self) -> None:
        s = ttk.Style(self.root)
        s.theme_use("clam")

        BG, WHITE, PRIMARY, ACCENT = "#f0f2f5", "#ffffff", "#1a5276", "#2980b9"

        s.configure("TFrame", background=BG)
        s.configure("White.TFrame", background=WHITE)
        s.configure("TLabel", background=BG, font=("Segoe UI", 9))
        s.configure("TLabelframe", background=BG, font=("Segoe UI", 9, "bold"))
        s.configure("TLabelframe.Label", background=BG, font=("Segoe UI", 9, "bold"),
                    foreground=PRIMARY)

        for name, bg, fg, hover in [
            ("Primary",   PRIMARY,   WHITE, ACCENT),
            ("Secondary", "#7f8c8d", WHITE, "#95a5a6"),
            ("Success",   "#27ae60", WHITE, "#2ecc71"),
        ]:
            s.configure(f"{name}.TButton",
                        background=bg, foreground=fg,
                        font=("Segoe UI", 9, "bold"), padding=(12, 5))
            s.map(f"{name}.TButton", background=[("active", hover), ("pressed", bg)])

        s.configure("TEntry", fieldbackground=WHITE, padding=4)
        s.configure("TNotebook", background=BG)
        s.configure("TNotebook.Tab", padding=(12, 6), font=("Segoe UI", 9))

        s.configure("Treeview", background=WHITE, fieldbackground=WHITE,
                    font=("Segoe UI", 9), rowheight=22)
        s.configure("Treeview.Heading",
                    font=("Segoe UI", 9, "bold"),
                    background=PRIMARY, foreground=WHITE)
        s.map("Treeview",
              background=[("selected", ACCENT)],
              foreground=[("selected", WHITE)])

    # ── construção da UI ─────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        self._build_conn_frame()
        self._build_tipo_estoque_frame()
        self._build_filter_frame()
        self._build_kpi_frame()
        self._build_toolbar()
        self._build_notebook()
        self._build_statusbar()

    # ── conexão ──────────────────────────────────────────────────────────────

    def _build_conn_frame(self) -> None:
        f = ttk.LabelFrame(self.root, text="Conexão", padding=(10, 6))
        f.pack(fill="x", padx=12, pady=(8, 3))

        ttk.Label(f, text="URL da API:").grid(row=0, column=0, sticky="w", padx=(0, 4))
        self._url_var = tk.StringVar(value=API_BASE_URL)
        ttk.Entry(f, textvariable=self._url_var, width=36).grid(
            row=0, column=1, sticky="ew", padx=(0, 18))

        ttk.Label(f, text="API Key:").grid(row=0, column=2, sticky="w", padx=(0, 4))
        self._key_var = tk.StringVar(value=API_KEY)
        ttk.Entry(f, textvariable=self._key_var, show="•", width=34).grid(
            row=0, column=3, sticky="ew", padx=(0, 12))

        ttk.Button(f, text="Testar conexão", style="Secondary.TButton",
                   command=self._test_connection).grid(row=0, column=4)

        f.columnconfigure(1, weight=1)
        f.columnconfigure(3, weight=1)

    # ── tipo de estoque (define o que é consultado na API) ─────────────────────

    def _build_tipo_estoque_frame(self) -> None:
        f = ttk.LabelFrame(
            self.root,
            text="Tipo de Estoque a analisar (recarrega da API ao mudar)",
            padding=(10, 6),
        )
        f.pack(fill="x", padx=12, pady=3)

        ttk.Label(f, text="Analisar:").grid(row=0, column=0, sticky="w", padx=(0, 6))
        self._tipo_estoque_var = tk.StringVar(value=TIPO_ESTOQUE_PADRAO)
        combo = ttk.Combobox(
            f, textvariable=self._tipo_estoque_var, width=42, state="readonly",
            values=list(TIPO_ESTOQUE_OPCOES.keys()),
        )
        combo.grid(row=0, column=1, sticky="w")
        combo.bind("<<ComboboxSelected>>", lambda e: self.consultar())

        ttk.Label(
            f, text="(a seleção define tipo_produto/local enviados à API)",
            foreground="#7f8c8d",
        ).grid(row=0, column=2, sticky="w", padx=(10, 0))

    # ── filtros (aplicados no cliente, após o carregamento) ────────────────────

    def _build_filter_frame(self) -> None:
        f = ttk.LabelFrame(
            self.root,
            text="Filtros (aplicados localmente sobre os dados carregados)",
            padding=(10, 6),
        )
        f.pack(fill="x", padx=12, pady=3)

        text_filters = [
            ("Filial:",     "_f_filial", 6),
            ("Código:",     "_f_codigo", 14),
            ("Descrição:",  "_f_desc",   26),
        ]
        col = 0
        for label, attr, width in text_filters:
            ttk.Label(f, text=label).grid(
                row=0, column=col, sticky="w", padx=(0 if col == 0 else 14, 3))
            var = tk.StringVar()
            setattr(self, attr, var)
            ttk.Entry(f, textvariable=var, width=width).grid(
                row=0, column=col + 1, sticky="ew")
            col += 2

        # botões
        btn = ttk.Frame(f)
        btn.grid(row=0, column=col, sticky="e", padx=(20, 0))
        ttk.Button(btn, text=" Carregar da API ", style="Primary.TButton",
                   command=self.consultar).pack(side="left", padx=(0, 6))
        ttk.Button(btn, text="Aplicar filtros", style="Secondary.TButton",
                   command=self._on_filtros_alterados).pack(side="left", padx=(0, 6))
        ttk.Button(btn, text="Limpar", style="Secondary.TButton",
                   command=self.limpar).pack(side="left")

        f.columnconfigure(col, weight=1)

    # ── KPI cards ─────────────────────────────────────────────────────────────

    def _build_kpi_frame(self) -> None:
        outer = tk.Frame(self.root, bg="#f0f2f5")
        outer.pack(fill="x", padx=12, pady=3)

        self._kpi_vars: Dict[str, tk.StringVar] = {}
        cards_cfg = [
            ("total_itens", "Itens em Estoque",   "#2c3e50"),
            ("filiais",     "Filiais",            "#7f8c8d"),
            ("qtd_total",   "Quantidade Total",   "#1a5276"),
            ("valor_total", "Valor Total",        "#154360"),
            ("valor_medio", "Valor Médio / Item", "#1e8449"),
        ]

        for i, (key, label, fg) in enumerate(cards_cfg):
            card = tk.Frame(outer, bg="#ffffff",
                            highlightbackground="#d5d8dc", highlightthickness=1)
            card.grid(row=0, column=i, sticky="nsew", padx=4, pady=2, ipady=10)

            tk.Label(card, text=label, bg="#ffffff",
                     font=("Segoe UI", 8), fg="#7f8c8d").pack(pady=(4, 0))

            var = tk.StringVar(value="—")
            self._kpi_vars[key] = var
            tk.Label(card, textvariable=var, bg="#ffffff",
                     font=("Segoe UI", 12, "bold"), fg=fg).pack()

        for i in range(len(cards_cfg)):
            outer.columnconfigure(i, weight=1)

    # ── toolbar acima do notebook ─────────────────────────────────────────────

    def _build_toolbar(self) -> None:
        bar = ttk.Frame(self.root)
        bar.pack(fill="x", padx=12, pady=(4, 0))
        ttk.Button(bar, text="  Exportar Excel  ", style="Success.TButton",
                   command=self.exportar_excel).pack(side="right")

    # ── notebook ──────────────────────────────────────────────────────────────

    def _build_notebook(self) -> None:
        self._nb = ttk.Notebook(self.root)
        self._nb.pack(fill="both", expand=True, padx=12, pady=(2, 4))

        tab_graficos = ttk.Frame(self._nb)
        self._nb.add(tab_graficos, text="  Gráficos  ")
        self._build_charts_tab(tab_graficos)

        tab_dados = ttk.Frame(self._nb)
        self._nb.add(tab_dados, text="  Dados  ")
        self._build_data_tab(tab_dados)

    def _build_charts_tab(self, parent: ttk.Frame) -> None:
        self._fig = Figure(figsize=(13, 6), dpi=96, facecolor="#f0f2f5")
        self._axes = self._fig.subplots(2, 2)
        self._fig.subplots_adjust(left=0.09, right=0.97, top=0.91,
                                  bottom=0.13, hspace=0.52, wspace=0.4)
        self._canvas = FigureCanvasTkAgg(self._fig, master=parent)
        self._canvas.get_tk_widget().pack(fill="both", expand=True, padx=4, pady=4)
        self._draw_empty_charts()

    def _build_data_tab(self, parent: ttk.Frame) -> None:
        cols = [c[0] for c in TREEVIEW_COLS]
        self._tree = ttk.Treeview(parent, columns=cols,
                                  show="headings", selectmode="browse")

        for col_id, heading, width, anchor in TREEVIEW_COLS:
            self._tree.heading(
                col_id, text=heading,
                command=lambda c=col_id: self._sort_tree(c),
            )
            self._tree.column(col_id, width=width, anchor=anchor,
                              stretch=(col_id == "descricao_produto"))

        vsb = ttk.Scrollbar(parent, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        parent.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)

    # ── barra de status ───────────────────────────────────────────────────────

    def _build_statusbar(self) -> None:
        bar = tk.Frame(self.root, bg="#dde1e7", height=24)
        bar.pack(fill="x", side="bottom")

        self._status_var = tk.StringVar(value="Pronto.")
        tk.Label(bar, textvariable=self._status_var,
                 bg="#dde1e7", font=("Segoe UI", 8), anchor="w").pack(
            side="left", padx=8, pady=2)

        self._pct_var = tk.StringVar()
        tk.Label(bar, textvariable=self._pct_var,
                 bg="#dde1e7", font=("Segoe UI", 8)).pack(side="right", padx=4)

        self._progress = ttk.Progressbar(bar, mode="determinate", length=200)
        self._progress.pack(side="right", padx=8, pady=3)

    # ── conexão ───────────────────────────────────────────────────────────────

    def _make_client(self) -> APIClient:
        return APIClient(
            base_url=self._url_var.get().strip(),
            api_key=self._key_var.get().strip(),
            api_key_name=API_KEY_NAME,
        )

    def _test_connection(self) -> None:
        try:
            self._make_client().health()
            messagebox.showinfo("Conexão OK",
                                "A API respondeu com sucesso!")
        except Exception as exc:
            messagebox.showerror("Falha na conexão", str(exc))

    # ── consulta (thread) ─────────────────────────────────────────────────────

    def consultar(self) -> None:
        if self._loading:
            return
        self._loading = True
        self._clear_data()
        self._set_status(
            f"Consultando /saldos-estoque na API… [{self._tipo_estoque_var.get()}]",
            indeterminate=True,
        )
        threading.Thread(target=self._fetch_thread, daemon=True).start()

    def _fetch_thread(self) -> None:
        try:
            client = self._make_client()
            tipo_produto, local = TIPO_ESTOQUE_OPCOES[self._tipo_estoque_var.get()]
            items, total = client.get_all_saldos_estoque(
                tipo_produto=tipo_produto, local=local,
                progress_callback=self._on_progress)
            self.root.after(0, self._on_data_ready, items, total)
        except Exception as exc:
            self.root.after(0, self._on_error, str(exc))

    def _on_progress(self, loaded: int, total: int) -> None:
        pct = int(100 * loaded / total) if total else 0
        self.root.after(0, self._update_progress_bar, loaded, total, pct)

    def _update_progress_bar(self, loaded: int, total: int, pct: int) -> None:
        self._progress.stop()
        self._progress.configure(mode="determinate", value=pct)
        self._pct_var.set(f"{pct}%  ({loaded}/{total})")
        self._status_var.set(f"Carregando… {loaded} de {total} registros")

    def _on_data_ready(self, items: List[Dict], total_api: int) -> None:
        self._loading = False

        df = pd.DataFrame(items) if items else pd.DataFrame()
        if not df.empty:
            df["quantidade"] = pd.to_numeric(df["quantidade"], errors="coerce").fillna(0.0)
            df["valor_atual"] = pd.to_numeric(df["valor_atual"], errors="coerce").fillna(0.0)

        self._df_raw = df
        self._apply_filters_and_refresh()

        n = len(self._df) if self._df is not None else 0
        self._set_status(
            f"{n} item(ns) exibido(s) — {self._tipo_estoque_var.get()}"
            + (f"  [total carregado da API: {total_api}]" if n != total_api else "")
        )
        self._progress["value"] = 100
        self._pct_var.set("100%")

    def _on_error(self, msg: str) -> None:
        self._loading = False
        self._progress.stop()
        self._progress.configure(mode="determinate", value=0)
        self._pct_var.set("")
        self._set_status(f"Erro: {msg}")
        messagebox.showerror("Erro ao consultar a API", msg)

    # ── filtros (client-side) ────────────────────────────────────────────────

    def _on_filtros_alterados(self) -> None:
        if self._df_raw is None:
            messagebox.showinfo(
                "Filtros", "Carregue os dados da API antes de aplicar filtros.")
            return
        self._apply_filters_and_refresh()
        n = len(self._df) if self._df is not None else 0
        self._set_status(f"{n} item(ns) exibido(s) após filtros")

    def _apply_filters_and_refresh(self) -> None:
        df = self._df_raw
        if df is None or df.empty:
            self._df = df
        else:
            self._df = self._apply_filters(df)
        self._update_kpis()
        self._update_charts()
        self._update_table()

    def _apply_filters(self, df: pd.DataFrame) -> pd.DataFrame:
        filial = self._f_filial.get().strip()
        if filial:
            df = df[df["filial"] == filial]

        codigo = self._f_codigo.get().strip().lower()
        if codigo:
            df = df[df["codigo_produto"].fillna("").str.lower().str.contains(codigo)]

        descricao = self._f_desc.get().strip().lower()
        if descricao:
            df = df[df["descricao_produto"].fillna("").str.lower().str.contains(descricao)]

        return df

    # ── KPIs ──────────────────────────────────────────────────────────────────

    def _update_kpis(self) -> None:
        df = self._df
        if df is None or df.empty:
            for v in self._kpi_vars.values():
                v.set("—")
            return

        valor_total = float(df["valor_atual"].sum())
        valor_medio = valor_total / len(df) if len(df) else 0.0

        self._kpi_vars["total_itens"].set(f'{len(df):,}'.replace(",", "."))
        self._kpi_vars["filiais"].set(f'{df["filial"].nunique():,}'.replace(",", "."))
        self._kpi_vars["qtd_total"].set(_qtd2(df["quantidade"].sum()))
        self._kpi_vars["valor_total"].set(_brl(valor_total))
        self._kpi_vars["valor_medio"].set(_brl(valor_medio))

    # ── gráficos ──────────────────────────────────────────────────────────────

    def _draw_empty_charts(self) -> None:
        for ax in self._axes.flat:
            ax.clear()
            ax.text(0.5, 0.5, "Sem dados — clique em \"Carregar da API\"",
                    ha="center", va="center", color="#aab7b8", fontsize=9)
            ax.set_facecolor("#f8f9fa")
            ax.axis("off")
        self._canvas.draw()

    def _update_charts(self) -> None:
        for ax in self._axes.flat:
            ax.clear()

        df = self._df
        if df is None or df.empty:
            self._draw_empty_charts()
            return

        ax_top_valor, ax_filial_valor = self._axes[0]
        ax_top_qtd,   ax_filial_qtd = self._axes[1]

        brl_fmt = mticker.FuncFormatter(
            lambda x, _: f"R${x/1000:.0f}k" if abs(x) >= 1000 else f"R${x:.0f}"
        )

        # ── Gráfico 1: barras horizontais — Top 15 Produtos por Valor ─────
        by_produto_valor = (
            df.assign(rotulo=df["codigo_produto"] + " — " + df["descricao_produto"].fillna(""))
            .groupby("rotulo")["valor_atual"]
            .sum()
            .nlargest(15)
            .sort_values(ascending=True)
        )
        bar_colors = CHART_COLORS[:len(by_produto_valor)]
        hbars = ax_top_valor.barh(
            by_produto_valor.index, by_produto_valor.values,
            color=(bar_colors * (len(by_produto_valor) // len(bar_colors) + 1))[::-1][:len(by_produto_valor)],
            height=0.65, edgecolor="none",
        )
        ax_top_valor.set_title("Top 15 Produtos — Valor em Estoque",
                               fontsize=9, fontweight="bold", pad=10)
        ax_top_valor.set_xlabel("Valor (R$)", fontsize=7)
        ax_top_valor.tick_params(axis="y", labelsize=6.5)
        ax_top_valor.tick_params(axis="x", labelsize=6)
        ax_top_valor.xaxis.set_major_formatter(brl_fmt)
        ax_top_valor.set_facecolor("#fafafa")
        for bar in hbars:
            w = bar.get_width()
            ax_top_valor.text(
                w * 1.01, bar.get_y() + bar.get_height() / 2,
                _brl(w), va="center", fontsize=6, color="#555",
            )

        # ── Gráfico 2: pizza — Valor Total por Filial ─────────────────────
        by_filial_valor = df.groupby("filial")["valor_atual"].sum().sort_values(ascending=False)
        pie_colors = CHART_COLORS[:len(by_filial_valor)]
        wedges, texts, autotexts = ax_filial_valor.pie(
            by_filial_valor,
            labels=by_filial_valor.index,
            autopct="%1.1f%%",
            colors=pie_colors,
            startangle=90,
            textprops={"fontsize": 8},
            wedgeprops={"edgecolor": "white", "linewidth": 1.5},
        )
        for at in autotexts:
            at.set_fontsize(7.5)
        ax_filial_valor.set_title("Valor Total por Filial",
                                  fontsize=9, fontweight="bold", pad=10)

        # ── Gráfico 3: barras horizontais — Top 15 Produtos por Quantidade ─
        by_produto_qtd = (
            df.assign(rotulo=df["codigo_produto"] + " — " + df["descricao_produto"].fillna(""))
            .groupby("rotulo")["quantidade"]
            .sum()
            .nlargest(15)
            .sort_values(ascending=True)
        )
        bar_colors_qtd = CHART_COLORS[:len(by_produto_qtd)]
        hbars_qtd = ax_top_qtd.barh(
            by_produto_qtd.index, by_produto_qtd.values,
            color=(bar_colors_qtd * (len(by_produto_qtd) // len(bar_colors_qtd) + 1))[::-1][:len(by_produto_qtd)],
            height=0.65, edgecolor="none",
        )
        ax_top_qtd.set_title("Top 15 Produtos — Quantidade em Estoque",
                             fontsize=9, fontweight="bold", pad=10)
        ax_top_qtd.set_xlabel("Quantidade", fontsize=7)
        ax_top_qtd.tick_params(axis="y", labelsize=6.5)
        ax_top_qtd.tick_params(axis="x", labelsize=6)
        ax_top_qtd.set_facecolor("#fafafa")
        for bar in hbars_qtd:
            w = bar.get_width()
            ax_top_qtd.text(
                w * 1.01, bar.get_y() + bar.get_height() / 2,
                _qtd(w), va="center", fontsize=6, color="#555",
            )

        # ── Gráfico 4: barras — Quantidade Total por Filial ────────────────
        by_filial_qtd = df.groupby("filial")["quantidade"].sum().sort_values(ascending=False)
        bar_clrs_filial = CHART_COLORS[:len(by_filial_qtd)]
        ax_filial_qtd.bar(by_filial_qtd.index, by_filial_qtd.values,
                          color=bar_clrs_filial, width=0.6, edgecolor="none")
        ax_filial_qtd.set_title("Quantidade Total por Filial",
                                fontsize=9, fontweight="bold", pad=10)
        ax_filial_qtd.set_ylabel("Quantidade", fontsize=7)
        ax_filial_qtd.tick_params(axis="both", labelsize=7)
        ax_filial_qtd.set_facecolor("#fafafa")

        self._canvas.draw()

    # ── tabela ────────────────────────────────────────────────────────────────

    def _update_table(self) -> None:
        for row in self._tree.get_children():
            self._tree.delete(row)

        df = self._df
        if df is None or df.empty:
            return

        for _, row in df.iterrows():
            self._tree.insert("", "end", values=(
                row.get("filial", ""),
                row.get("local", ""),
                row.get("codigo_produto", ""),
                row.get("descricao_produto", ""),
                _qtd(row.get("quantidade", 0)),
                _brl(row.get("valor_atual", 0)),
            ))

    def _sort_tree(self, col: str) -> None:
        if self._df is None or self._df.empty:
            return
        asc = not self._sort_ascending.get(col, False)
        self._sort_ascending[col] = asc

        if col in ("quantidade", "valor_atual"):
            self._df = self._df.sort_values(col, ascending=asc, na_position="last")
        else:
            self._df = self._df.sort_values(
                col, ascending=asc, na_position="last",
                key=lambda s: s.astype(str).str.lower(),
            )
        self._update_table()

    # ── exportar Excel ────────────────────────────────────────────────────────

    def exportar_excel(self) -> None:
        df = self._df
        if df is None or df.empty:
            messagebox.showwarning(
                "Exportar Excel",
                "Nenhum dado carregado. Faça uma consulta primeiro.",
            )
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("Todos os arquivos", "*.*")],
            title="Salvar planilha do Saldo de Estoque",
            initialfile=f"saldo_estoque_{date.today().strftime('%Y%m%d')}.xlsx",
        )
        if not path:
            return

        try:
            self._set_status("Gerando planilha Excel…")
            self.root.update()
            _write_excel(df, path)
            self._set_status(f"Planilha salva em: {path}")
            messagebox.showinfo("Exportar Excel",
                                f"Arquivo salvo com sucesso:\n{path}")
        except Exception as exc:
            self._set_status("Erro ao exportar.")
            messagebox.showerror("Erro ao exportar", str(exc))

    # ── utilitários ───────────────────────────────────────────────────────────

    def limpar(self) -> None:
        for attr in ("_f_filial", "_f_codigo", "_f_desc"):
            getattr(self, attr).set("")
        if self._df_raw is not None:
            self._apply_filters_and_refresh()
        self._set_status("Filtros limpos.")

    def _clear_data(self) -> None:
        self._df_raw = None
        self._df = None
        for v in self._kpi_vars.values():
            v.set("—")
        for row in self._tree.get_children():
            self._tree.delete(row)
        self._draw_empty_charts()
        self._progress["value"] = 0
        self._pct_var.set("")

    def _set_status(self, msg: str, *, indeterminate: bool = False) -> None:
        self._status_var.set(msg)
        if indeterminate:
            self._progress.configure(mode="indeterminate")
            self._progress.start(12)
            self._pct_var.set("")
        else:
            self._progress.stop()
            self._progress.configure(mode="determinate")


# ── exportação Excel (função separada) ───────────────────────────────────────

def _write_excel(df: pd.DataFrame, path: str) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    HDR_FILL  = PatternFill("solid", fgColor="1A5276")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    ALT_FILL  = PatternFill("solid", fgColor="EBF5FB")
    THIN      = Side(style="thin", color="BDC3C7")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER    = Alignment(horizontal="center", vertical="center")
    RIGHT     = Alignment(horizontal="right",  vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")
    BRL       = '#,##0.00'
    QTD_FMT   = '#,##0.0000'
    PCT       = '0.0%'
    BOLD10    = Font(bold=True, name="Calibri", size=10)

    def _hdr(ws, row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = BORDER
        return c

    # ── aba 1: Estoque ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Estoque"
    ws1.freeze_panes = "A2"

    cols_aba1 = [
        ("Filial",            "filial",              9),
        ("Local",              "local",               9),
        ("Código Produto",    "codigo_produto",      16),
        ("Descrição",         "descricao_produto",   42),
        ("Quantidade",        "quantidade",          16),
        ("Valor Atual (R$)",  "valor_atual",         18),
    ]
    for ci, (heading, _, w) in enumerate(cols_aba1, 1):
        _hdr(ws1, 1, ci, heading)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[1].height = 18
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(cols_aba1))}1"

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        row_fill = ALT_FILL if ri % 2 == 0 else None
        for ci, (_, field, _) in enumerate(cols_aba1, 1):
            val = row.get(field)
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if field == "valor_atual":
                cell.number_format = BRL
                cell.alignment = RIGHT
            elif field == "quantidade":
                cell.number_format = QTD_FMT
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT
            if row_fill:
                cell.fill = row_fill

    # ── aba 2: Resumo ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 22

    ws2.cell(1, 1, "Resumo — Saldo de Estoque").font = Font(
        bold=True, color="1A5276", size=14, name="Calibri")
    ws2.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(
        italic=True, size=9, color="7F8C8D", name="Calibri")

    valor_total = float(df["valor_atual"].sum())
    valor_medio = valor_total / len(df) if len(df) else 0.0

    items_resumo = [
        ("Total de Itens",           len(df),                        None),
        ("Filiais Distintas",        int(df["filial"].nunique()),    None),
        ("Quantidade Total",         float(df["quantidade"].sum()),  QTD_FMT),
        ("Valor Total (R$)",         valor_total,                    BRL),
        ("Valor Médio por Item (R$)", valor_medio,                   BRL),
    ]
    for i, (lbl, val, fmt) in enumerate(items_resumo, 4):
        c1 = ws2.cell(i, 1, lbl)
        c1.font = Font(bold=True, name="Calibri", size=10)
        c2 = ws2.cell(i, 2, val)
        c2.alignment = RIGHT
        if fmt:
            c2.number_format = fmt

    # ── aba 3: Por Filial ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Por Filial")
    by_filial = (
        df.groupby("filial", as_index=False)
        .agg(qtd_itens=("codigo_produto", "count"),
             quantidade_total=("quantidade", "sum"),
             valor_total=("valor_atual", "sum"))
        .sort_values("valor_total", ascending=False)
    )
    total_geral_filial = float(by_filial["valor_total"].sum())

    hdrs3 = ["Filial", "Qtd. Itens", "Quantidade Total", "Valor Total (R$)", "% do Total"]
    widths3 = [12, 12, 18, 18, 12]
    for ci, (h, w) in enumerate(zip(hdrs3, widths3), 1):
        _hdr(ws3, 1, ci, h)
        ws3.column_dimensions[get_column_letter(ci)].width = w

    ri = 2
    for row in by_filial.itertuples(index=False):
        pct = (row.valor_total / total_geral_filial) if total_geral_filial else 0.0
        ws3.cell(ri, 1, row.filial)
        ws3.cell(ri, 2, int(row.qtd_itens))
        c3 = ws3.cell(ri, 3, float(row.quantidade_total)); c3.number_format = QTD_FMT
        c4 = ws3.cell(ri, 4, float(row.valor_total)); c4.number_format = BRL
        c5 = ws3.cell(ri, 5, pct); c5.number_format = PCT
        if ri % 2 == 0:
            for ci in range(1, 6):
                ws3.cell(ri, ci).fill = ALT_FILL
        ri += 1

    ws3.cell(ri, 1, "TOTAL GERAL").font = BOLD10
    ws3.cell(ri, 2, int(by_filial["qtd_itens"].sum())).font = BOLD10
    c3 = ws3.cell(ri, 3, float(by_filial["quantidade_total"].sum())); c3.number_format = QTD_FMT; c3.font = BOLD10
    c4 = ws3.cell(ri, 4, total_geral_filial); c4.number_format = BRL; c4.font = BOLD10
    c5 = ws3.cell(ri, 5, 1.0); c5.number_format = PCT; c5.font = BOLD10

    # ── aba 4: Top Produtos ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Top Produtos")
    by_produto = (
        df.groupby(["codigo_produto", "descricao_produto"], as_index=False, dropna=False)
        .agg(quantidade_total=("quantidade", "sum"),
             valor_total=("valor_atual", "sum"))
        .sort_values("valor_total", ascending=False)
        .head(50)
    )
    hdrs4 = ["Código Produto", "Descrição", "Quantidade Total", "Valor Total (R$)"]
    widths4 = [16, 42, 18, 18]
    for ci, (h, w) in enumerate(zip(hdrs4, widths4), 1):
        _hdr(ws4, 1, ci, h)
        ws4.column_dimensions[get_column_letter(ci)].width = w
    for ri, row in enumerate(by_produto.itertuples(index=False), 2):
        ws4.cell(ri, 1, row.codigo_produto)
        ws4.cell(ri, 2, row.descricao_produto)
        c3 = ws4.cell(ri, 3, float(row.quantidade_total)); c3.number_format = QTD_FMT
        c4 = ws4.cell(ri, 4, float(row.valor_total)); c4.number_format = BRL
        if ri % 2 == 0:
            for ci in range(1, 5):
                ws4.cell(ri, ci).fill = ALT_FILL

    wb.save(path)


# ── entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    EstoqueApp(root)
    root.mainloop()
