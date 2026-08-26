import io
from decimal import Decimal

import openpyxl
import pytest

from app.models.produto import Produto
from app.models.saldo_estoque import SaldoEstoque


def _produto(**overrides):
    dados = dict(
        deletado=" ",
        filial="  ",
        codigo="PROD001",
        descricao="Produto Exemplo",
        tipo="PA",
        conversao=Decimal("1.0000"),
    )
    dados.update(overrides)
    return Produto(**dados)


def _saldo(**overrides):
    dados = dict(
        deletado=" ",
        filial="01",
        codigo_produto="PROD001",
        local="01",
        saldo_atual=Decimal("120.0000"),
        valor_atual=Decimal("3500.00"),
    )
    dados.update(overrides)
    return SaldoEstoque(**dados)


def test_requer_api_key(client):
    assert client.get("/saldos-estoque/").status_code == 401


def test_lista_com_join_de_produto_e_conversao(client, auth_headers, db_session):
    db_session.add_all([_produto(), _saldo()])
    db_session.commit()

    resposta = client.get("/saldos-estoque/", headers=auth_headers)
    assert resposta.status_code == 200
    dados = resposta.json()
    assert len(dados["items"]) == 1
    item = dados["items"][0]
    assert item["codigo_produto"] == "PROD001"
    assert item["descricao_produto"] == "Produto Exemplo"
    assert item["quantidade"] == "120"
    assert item["valor_atual"] == "3500.00"


def test_quantidade_e_dividida_pela_conversao_do_produto(client, auth_headers, db_session):
    db_session.add_all([
        _produto(conversao=Decimal("2.0000")),
        _saldo(saldo_atual=Decimal("100.0000")),
    ])
    db_session.commit()

    resposta = client.get("/saldos-estoque/", headers=auth_headers)
    assert resposta.json()["items"][0]["quantidade"] == "50"


def test_saldo_sem_produto_correspondente_e_ignorado(client, auth_headers, db_session):
    db_session.add(_saldo())
    db_session.commit()

    resposta = client.get("/saldos-estoque/", headers=auth_headers)
    assert len(resposta.json()["items"]) == 0


def test_sem_filtro_de_tipo_ou_local_traz_qualquer_combinacao(client, auth_headers, db_session):
    db_session.add_all([
        _produto(codigo="PROD001", tipo="PA"),
        _saldo(codigo_produto="PROD001", local="01"),
        _produto(codigo="PROD002", tipo="AM"),
        _saldo(codigo_produto="PROD002", local="20"),
    ])
    db_session.commit()

    resposta = client.get("/saldos-estoque/", headers=auth_headers)
    assert len(resposta.json()["items"]) == 2


def test_filtro_tipo_produto_via_query_string(client, auth_headers, db_session):
    db_session.add_all([
        _produto(codigo="PROD001", tipo="PA"),
        _saldo(codigo_produto="PROD001", local="01"),
        _produto(codigo="PROD002", tipo="AM"),
        _saldo(codigo_produto="PROD002", local="20"),
    ])
    db_session.commit()

    resposta = client.get("/saldos-estoque/?tipo_produto=AM", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["codigo_produto"] == "PROD002"


def test_filtro_local_via_query_string(client, auth_headers, db_session):
    db_session.add_all([
        _produto(codigo="PROD001"),
        _saldo(codigo_produto="PROD001", local="01"),
        _produto(codigo="PROD002"),
        _saldo(codigo_produto="PROD002", local="20"),
    ])
    db_session.commit()

    resposta = client.get("/saldos-estoque/?local=20", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["codigo_produto"] == "PROD002"


def test_filtro_tipo_e_local_combinados(client, auth_headers, db_session):
    db_session.add_all([
        _produto(codigo="PROD001", tipo="PA"),
        _saldo(codigo_produto="PROD001", local="01"),
        _produto(codigo="PROD002", tipo="AM"),
        _saldo(codigo_produto="PROD002", local="20"),
        _produto(codigo="PROD003", tipo="PA"),
        _saldo(codigo_produto="PROD003", local="20"),
    ])
    db_session.commit()

    resposta = client.get(
        "/saldos-estoque/?tipo_produto=AM&local=20", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["codigo_produto"] == "PROD002"


def test_saldo_zerado_ou_negativo_e_ignorado(client, auth_headers, db_session):
    db_session.add_all([
        _produto(),
        _saldo(codigo_produto="PROD001", saldo_atual=Decimal("0.0000")),
    ])
    db_session.commit()

    resposta = client.get("/saldos-estoque/", headers=auth_headers)
    assert len(resposta.json()["items"]) == 0


def test_registros_deletados_sao_ignorados(client, auth_headers, db_session):
    db_session.add_all([
        _produto(),
        _saldo(codigo_produto="PROD001"),
        _saldo(codigo_produto="PROD001", deletado="*"),
    ])
    db_session.commit()

    resposta = client.get("/saldos-estoque/", headers=auth_headers)
    assert len(resposta.json()["items"]) == 1


def test_metodos_de_escrita_nao_existem(client, auth_headers):
    assert client.post("/saldos-estoque/", json={}, headers=auth_headers).status_code == 405
    assert client.put("/saldos-estoque/", json={}, headers=auth_headers).status_code == 405
    assert client.delete("/saldos-estoque/", headers=auth_headers).status_code == 405


def test_export_gera_planilha_com_4_abas_e_valores_corretos(client, auth_headers, db_session):
    db_session.add_all([
        _produto(codigo="PROD001", descricao="Produto Um"),
        _saldo(codigo_produto="PROD001", filial="01",
               saldo_atual=Decimal("120.0000"), valor_atual=Decimal("3500.00")),
    ])
    db_session.commit()

    resposta = client.get("/saldos-estoque/export", headers=auth_headers)
    assert resposta.status_code == 200
    assert resposta.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "filename=" in resposta.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(resposta.content))
    assert wb.sheetnames == ["Estoque", "Resumo", "Por Filial", "Top Produtos"]

    # Aba "Estoque": A=Filial B=Local C=Código Produto D=Descrição
    # E=Quantidade F=Valor Atual
    ws1 = wb["Estoque"]
    assert ws1["A1"].value == "Filial"
    assert ws1["A2"].value == "01"
    assert ws1["C2"].value == "PROD001"
    assert ws1["E2"].value == pytest.approx(120.0)
    assert ws1["F2"].value == pytest.approx(3500.0)

    ws2 = wb["Resumo"]
    resumo = {ws2.cell(i, 1).value: ws2.cell(i, 2).value for i in range(4, 9)}
    assert resumo["Total de Itens"] == 1
    assert resumo["Valor Total (R$)"] == pytest.approx(3500.0)
    assert resumo["Valor Médio por Item (R$)"] == pytest.approx(3500.0 / 120.0)

    ws3 = wb["Por Filial"]
    assert ws3["A2"].value == "01"
    assert ws3["A3"].value == "TOTAL GERAL"

    ws4 = wb["Top Produtos"]
    assert ws4["A2"].value == "PROD001"


def test_export_aplica_filtros_locais_de_filial_codigo_e_descricao(client, auth_headers, db_session):
    db_session.add_all([
        _produto(codigo="PROD001", descricao="Parafuso Sextavado"),
        _saldo(codigo_produto="PROD001", filial="01"),
        _produto(codigo="PROD002", descricao="Porca Sextavada"),
        _saldo(codigo_produto="PROD002", filial="02"),
    ])
    db_session.commit()

    resposta_filial = client.get(
        "/saldos-estoque/export?filial=01", headers=auth_headers)
    ws_filial = openpyxl.load_workbook(io.BytesIO(resposta_filial.content))["Estoque"]
    filiais = [ws_filial.cell(r, 1).value for r in range(2, ws_filial.max_row + 1)
               if ws_filial.cell(r, 1).value]
    assert filiais == ["01"]

    resposta_descricao = client.get(
        "/saldos-estoque/export?descricao=porca", headers=auth_headers)
    ws_desc = openpyxl.load_workbook(io.BytesIO(resposta_descricao.content))["Estoque"]
    codigos = [ws_desc.cell(r, 3).value for r in range(2, ws_desc.max_row + 1)
               if ws_desc.cell(r, 3).value]
    assert codigos == ["PROD002"]
