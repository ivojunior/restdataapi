"""Geração da planilha Excel do relatório de cargas — réplica de
`_write_excel()` em `client/app_cargas.py`, mesma estrutura de 4 abas e
mesma formatação, a partir dos dados já calculados pela API.

Sem pandas (ver app/excel/faturamento.py — mesmo raciocínio). Uma "carga" é
identificada por (filial, codigo) — cada linha da API é um ITEM de uma
carga, não uma carga distinta; as agregações de "quantidade de cargas"
contam pares (filial, codigo) únicos, não linhas.
"""
import io
from collections import OrderedDict
from datetime import datetime
from decimal import Decimal
from typing import Sequence

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.excel.styles import ALT_FILL, BOLD10, BORDER, BRL, LEFT, PCT, QTD_FMT, RIGHT, cabecalho
from app.schemas.carga import CargaRead

_STATUS_CARGA_FECHADO = "Fechada"

_COLUNAS_ABA1 = [
    ("Filial", "filial", 9),
    ("Carga", "codigo", 12),
    ("Data", "data", 13),
    ("Pedido", "pedido", 13),
    ("Motorista", "motorista", 26),
    ("Cliente", "nome_cliente", 34),
    ("Bairro", "bairro_cliente", 24),
    ("Município", "municipio_cliente", 24),
    ("Nota Fiscal", "nota_fiscal", 14),
    ("Caminhão", "caminhao", 14),
    ("Status", "status_carga", 18),
    ("Peso (kg)", "peso", 14),
    ("Valor (R$)", "valor", 16),
]


def _num(valor):
    return float(valor) if isinstance(valor, Decimal) else (valor or 0.0)


def _fmt_data(valor: str) -> str:
    s = str(valor or "").strip()
    if len(s) == 8 and s.isdigit():
        return f"{s[6:8]}/{s[4:6]}/{s[:4]}"
    return s or ""


def gerar_excel_cargas(linhas: Sequence[CargaRead]) -> io.BytesIO:
    wb = openpyxl.Workbook()

    _aba_cargas(wb.active, linhas)
    _aba_resumo(wb.create_sheet("Resumo"), linhas)
    _aba_por_filial(wb.create_sheet("Por Filial"), linhas)
    _aba_top_clientes(wb.create_sheet("Top Clientes"), linhas)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _aba_cargas(ws, linhas: Sequence[CargaRead]) -> None:
    ws.title = "Cargas"
    ws.freeze_panes = "A2"

    for ci, (titulo, _, largura) in enumerate(_COLUNAS_ABA1, 1):
        cabecalho(ws, 1, ci, titulo)
        ws.column_dimensions[get_column_letter(ci)].width = largura
    ws.row_dimensions[1].height = 18
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUNAS_ABA1))}1"

    for ri, linha in enumerate(linhas, 2):
        preencher_linha = ALT_FILL if ri % 2 == 0 else None
        for ci, (_, campo, _largura) in enumerate(_COLUNAS_ABA1, 1):
            valor = getattr(linha, campo)
            if campo == "data":
                valor = _fmt_data(valor)
            elif isinstance(valor, Decimal):
                valor = float(valor)
            celula = ws.cell(row=ri, column=ci, value=valor)
            celula.border = BORDER
            if campo == "valor":
                celula.number_format = BRL
                celula.alignment = RIGHT
            elif campo == "peso":
                celula.number_format = QTD_FMT
                celula.alignment = RIGHT
            else:
                celula.alignment = LEFT
            if preencher_linha:
                celula.fill = preencher_linha


def _aba_resumo(ws, linhas: Sequence[CargaRead]) -> None:
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    ws.cell(1, 1, "Resumo — Cargas").font = Font(
        bold=True, color="1A5276", size=14, name="Calibri")
    ws.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(
        italic=True, size=9, color="7F8C8D", name="Calibri")

    n_cargas = len({(l.filial, l.codigo) for l in linhas})
    pedidos_distintos = len({l.pedido for l in linhas})
    peso_total = sum(_num(l.peso) for l in linhas)
    valor_total = sum(_num(l.valor) for l in linhas)
    valor_aberto = sum(_num(l.valor) for l in linhas if l.status_carga != _STATUS_CARGA_FECHADO)
    valor_acertado = valor_total - valor_aberto
    valor_medio = (valor_total / n_cargas) if n_cargas else 0.0

    itens_resumo = [
        ("Total de Cargas", n_cargas, None),
        ("Pedidos Distintos", pedidos_distintos, None),
        ("Peso Total (kg)", peso_total, QTD_FMT),
        ("Valor Total (R$)", valor_total, BRL),
        ("Valor em Aberto (R$)", valor_aberto, BRL),
        ("Valor Acertado (R$)", valor_acertado, BRL),
        ("Valor Médio por Carga (R$)", valor_medio, BRL),
    ]
    for i, (rotulo, valor, formato) in enumerate(itens_resumo, 4):
        c1 = ws.cell(i, 1, rotulo)
        c1.font = Font(bold=True, name="Calibri", size=10)
        c2 = ws.cell(i, 2, valor)
        c2.alignment = RIGHT
        if formato:
            c2.number_format = formato


def _aba_por_filial(ws, linhas: Sequence[CargaRead]) -> None:
    agregados: "OrderedDict[str, dict]" = OrderedDict()
    for l in linhas:
        agg = agregados.setdefault(
            l.filial, {"cargas": set(), "peso_total": 0.0, "valor_total": 0.0})
        agg["cargas"].add((l.filial, l.codigo))
        agg["peso_total"] += _num(l.peso)
        agg["valor_total"] += _num(l.valor)

    ordenados = sorted(agregados.items(), key=lambda kv: kv[1]["valor_total"], reverse=True)
    total_geral = sum(v["valor_total"] for _, v in ordenados)

    cabecalhos = ["Filial", "Qtd. Cargas", "Peso Total (kg)", "Valor Total (R$)", "% do Total"]
    larguras = [12, 12, 18, 18, 12]
    for ci, (titulo, largura) in enumerate(zip(cabecalhos, larguras), 1):
        cabecalho(ws, 1, ci, titulo)
        ws.column_dimensions[get_column_letter(ci)].width = largura

    ri = 2
    total_cargas = 0
    for filial, agg in ordenados:
        qtd_cargas = len(agg["cargas"])
        total_cargas += qtd_cargas
        pct = (agg["valor_total"] / total_geral) if total_geral else 0.0
        ws.cell(ri, 1, filial)
        ws.cell(ri, 2, qtd_cargas)
        c3 = ws.cell(ri, 3, agg["peso_total"]); c3.number_format = QTD_FMT
        c4 = ws.cell(ri, 4, agg["valor_total"]); c4.number_format = BRL
        c5 = ws.cell(ri, 5, pct); c5.number_format = PCT
        if ri % 2 == 0:
            for ci in range(1, 6):
                ws.cell(ri, ci).fill = ALT_FILL
        ri += 1

    ws.cell(ri, 1, "TOTAL GERAL").font = BOLD10
    ws.cell(ri, 2, total_cargas).font = BOLD10
    c3 = ws.cell(ri, 3, sum(v["peso_total"] for _, v in ordenados)); c3.number_format = QTD_FMT; c3.font = BOLD10
    c4 = ws.cell(ri, 4, total_geral); c4.number_format = BRL; c4.font = BOLD10
    c5 = ws.cell(ri, 5, 1.0); c5.number_format = PCT; c5.font = BOLD10


def _aba_top_clientes(ws, linhas: Sequence[CargaRead], limite: int = 50) -> None:
    agregados: "OrderedDict[str, dict]" = OrderedDict()
    for l in linhas:
        nome = l.nome_cliente or "(sem nome)"
        agg = agregados.setdefault(
            nome, {"cargas": set(), "peso_total": 0.0, "valor_total": 0.0})
        agg["cargas"].add((l.filial, l.codigo))
        agg["peso_total"] += _num(l.peso)
        agg["valor_total"] += _num(l.valor)

    top = sorted(agregados.items(), key=lambda kv: kv[1]["valor_total"], reverse=True)[:limite]

    cabecalhos = ["Cliente", "Qtd. Cargas", "Peso Total (kg)", "Valor Total (R$)"]
    larguras = [34, 12, 18, 18]
    for ci, (titulo, largura) in enumerate(zip(cabecalhos, larguras), 1):
        cabecalho(ws, 1, ci, titulo)
        ws.column_dimensions[get_column_letter(ci)].width = largura

    for ri, (nome, agg) in enumerate(top, 2):
        ws.cell(ri, 1, nome)
        ws.cell(ri, 2, len(agg["cargas"]))
        c3 = ws.cell(ri, 3, agg["peso_total"]); c3.number_format = QTD_FMT
        c4 = ws.cell(ri, 4, agg["valor_total"]); c4.number_format = BRL
        if ri % 2 == 0:
            for ci in range(1, 5):
                ws.cell(ri, ci).fill = ALT_FILL
