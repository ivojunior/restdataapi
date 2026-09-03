import io
from datetime import date, timedelta

import openpyxl
import pytest

from app.models.fornecedor import Fornecedor
from app.models.tipo_operacao import TipoOperacao
from app.models.titulo_pagar import TituloPagar


def _fornecedor(**overrides):
    dados = dict(
        deletado=" ",
        filial="  ",
        codigo="000123",
        loja="01",
        nome="Fornecedor Exemplo Ltda",
    )
    dados.update(overrides)
    return Fornecedor(**dados)


def _tipo_operacao(**overrides):
    dados = dict(
        deletado=" ",
        filial="  ",
        codigo="001",
        descricao="Compra de Mercadoria",
    )
    dados.update(overrides)
    return TipoOperacao(**dados)


def _titulo(**overrides):
    dados = dict(
        deletado=" ",
        filial="01",
        prefixo="A",
        numero="000001",
        parcela="01",
        tipo="NF",
        fornecedor="000123",
        loja="01",
        emissao="20260301",
        vencimento=date.today().strftime("%Y%m%d"),
        vencimento_real=date.today().strftime("%Y%m%d"),
        valor="1000.00",
        saldo="1000.00",
        moeda="01",
        historico="Compra de mercadorias",
        codigo_operacao="001",
        nome_fornecedor="Fornecedor Exemplo Ltda",
    )
    dados.update(overrides)
    return TituloPagar(**dados)


def test_requer_api_key(client):
    assert client.get("/financeiro/").status_code == 401


def test_lista_com_join_de_fornecedor_e_tipo_operacao(client, auth_headers, db_session):
    db_session.add_all([_fornecedor(), _tipo_operacao(), _titulo()])
    db_session.commit()

    resposta = client.get("/financeiro/", headers=auth_headers)
    assert resposta.status_code == 200
    dados = resposta.json()
    assert len(dados["items"]) == 1
    item = dados["items"][0]
    assert item["numero"] == "000001"
    assert item["nome_fornecedor"] == "Fornecedor Exemplo Ltda"
    assert item["descricao_operacao"] == "Compra de Mercadoria"


def test_expoe_recuperacao_judicial(client, auth_headers, db_session):
    db_session.add_all(
        [
            _fornecedor(),
            _titulo(numero="000001", recuperacao_judicial="1"),
            _titulo(numero="000002", recuperacao_judicial="2"),
            _titulo(numero="000003"),
        ]
    )
    db_session.commit()

    resposta = client.get("/financeiro/", headers=auth_headers)
    valores = {item["numero"]: item["recuperacao_judicial"] for item in resposta.json()["items"]}
    assert valores["000001"] == "1"
    assert valores["000002"] == "2"
    assert valores["000003"] is None


def test_valor_soma_retencoes_ao_valor_bruto(client, auth_headers, db_session):
    # Réplica de select_financeiro.sql: E2_VALOR exposto = E2_IRRF + E2_CSLL +
    # E2_PIS + E2_COFINS + E2_VALOR (coluna bruta), não a coluna bruta sozinha.
    db_session.add_all([
        _fornecedor(),
        _titulo(valor="1000.00", irrf="10.00", csll="20.00", pis="5.00", cofins="15.00"),
    ])
    db_session.commit()

    resposta = client.get("/financeiro/", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert float(dados["items"][0]["valor"]) == pytest.approx(1050.0)


def test_valor_trata_retencoes_nulas_como_zero(client, auth_headers, db_session):
    db_session.add_all([_fornecedor(), _titulo(valor="1000.00")])
    db_session.commit()

    resposta = client.get("/financeiro/", headers=auth_headers)
    dados = resposta.json()
    assert float(dados["items"][0]["valor"]) == pytest.approx(1000.0)


def test_titulo_sem_tipo_operacao_correspondente_ainda_aparece(client, auth_headers, db_session):
    db_session.add_all([_fornecedor(), _titulo(codigo_operacao="999")])
    db_session.commit()

    resposta = client.get("/financeiro/", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["descricao_operacao"] is None


def test_titulo_sem_fornecedor_correspondente_e_ignorado(client, auth_headers, db_session):
    db_session.add(_titulo())
    db_session.commit()

    resposta = client.get("/financeiro/", headers=auth_headers)
    assert len(resposta.json()["items"]) == 0


def test_tipos_excluidos_sao_ignorados(client, auth_headers, db_session):
    db_session.add_all(
        [
            _fornecedor(),
            _titulo(numero="000001", tipo="NF"),
            _titulo(numero="000002", tipo="PA"),
            _titulo(numero="000003", tipo="PR"),
            _titulo(numero="000004", tipo="NDF"),
        ]
    )
    db_session.commit()

    resposta = client.get("/financeiro/", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["numero"] == "000001"


def test_sem_filtro_ignora_vencimento_anterior_a_hoje(client, auth_headers, db_session):
    ontem = (date.today() - timedelta(days=1)).strftime("%Y%m%d")
    db_session.add_all(
        [
            _fornecedor(),
            _titulo(numero="000001", vencimento_real=ontem),
        ]
    )
    db_session.commit()

    resposta = client.get("/financeiro/", headers=auth_headers)
    assert len(resposta.json()["items"]) == 0


def test_filtro_vencimento_de_via_query_string(client, auth_headers, db_session):
    ontem = (date.today() - timedelta(days=1)).strftime("%Y%m%d")
    hoje = date.today().strftime("%Y%m%d")
    db_session.add_all(
        [
            _fornecedor(),
            _titulo(numero="000001", vencimento_real=ontem),
            _titulo(numero="000002", vencimento_real=hoje),
        ]
    )
    db_session.commit()

    resposta = client.get(f"/financeiro/?vencimento_de={ontem}", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 2

    resposta = client.get(f"/financeiro/?vencimento_de={hoje}", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["numero"] == "000002"


def test_filtro_vencimento_ate_via_query_string(client, auth_headers, db_session):
    hoje = date.today().strftime("%Y%m%d")
    amanha = (date.today() + timedelta(days=1)).strftime("%Y%m%d")
    db_session.add_all(
        [
            _fornecedor(),
            _titulo(numero="000001", vencimento_real=hoje),
            _titulo(numero="000002", vencimento_real=amanha),
        ]
    )
    db_session.commit()

    resposta = client.get(
        f"/financeiro/?vencimento_de={hoje}&vencimento_ate={hoje}", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["numero"] == "000001"


def test_filtro_status_baixado_via_query_string(client, auth_headers, db_session):
    hoje = date.today().strftime("%Y%m%d")
    db_session.add_all(
        [
            _fornecedor(),
            _titulo(numero="000001", vencimento_real=hoje, data_baixa=hoje),
            _titulo(numero="000002", vencimento_real=hoje, data_baixa=""),
        ]
    )
    db_session.commit()

    resposta = client.get("/financeiro/?status=baixado", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["numero"] == "000001"


def test_filtro_status_baixado_ignora_data_baixa_so_com_espacos(client, auth_headers, db_session):
    hoje = date.today().strftime("%Y%m%d")
    db_session.add_all(
        [
            _fornecedor(),
            _titulo(numero="000001", vencimento_real=hoje, data_baixa="        "),
        ]
    )
    db_session.commit()

    resposta = client.get("/financeiro/?status=baixado", headers=auth_headers)
    assert len(resposta.json()["items"]) == 0

    resposta = client.get("/financeiro/?status=em_aberto", headers=auth_headers)
    assert len(resposta.json()["items"]) == 1


def test_filtro_status_em_aberto_via_query_string(client, auth_headers, db_session):
    hoje = date.today().strftime("%Y%m%d")
    db_session.add_all(
        [
            _fornecedor(),
            _titulo(numero="000001", vencimento_real=hoje, data_baixa=""),
            _titulo(numero="000002", vencimento_real=hoje, data_baixa=hoje),
        ]
    )
    db_session.commit()

    resposta = client.get("/financeiro/?status=em_aberto", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["numero"] == "000001"


def test_filtro_status_vencido_via_query_string(client, auth_headers, db_session):
    # vencimento_de precisa ser recuado, senão o filtro padrão de período
    # (vencimento_real >= hoje) já exclui o título vencido antes do status.
    ontem = (date.today() - timedelta(days=1)).strftime("%Y%m%d")
    hoje = date.today().strftime("%Y%m%d")
    db_session.add_all(
        [
            _fornecedor(),
            _titulo(numero="000001", vencimento_real=ontem, data_baixa=""),
            _titulo(numero="000002", vencimento_real=hoje, data_baixa=""),
        ]
    )
    db_session.commit()

    resposta = client.get(
        f"/financeiro/?vencimento_de={ontem}&status=vencido", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["numero"] == "000001"


def test_filtro_status_invalido_retorna_422(client, auth_headers):
    resposta = client.get("/financeiro/?status=quitado", headers=auth_headers)
    assert resposta.status_code == 422


def test_metodos_de_escrita_nao_existem(client, auth_headers):
    assert client.post("/financeiro/", json={}, headers=auth_headers).status_code == 405
    assert client.put("/financeiro/", json={}, headers=auth_headers).status_code == 405
    assert client.delete("/financeiro/", headers=auth_headers).status_code == 405


def test_export_gera_planilha_com_8_abas_e_valores_corretos(client, auth_headers, db_session):
    db_session.add_all([_fornecedor(), _tipo_operacao(), _titulo()])
    db_session.commit()

    resposta = client.get("/financeiro/export", headers=auth_headers)
    assert resposta.status_code == 200
    assert resposta.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "filename=" in resposta.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(resposta.content))
    assert wb.sheetnames == [
        "Financeiro", "Resumo", "Por Fornecedor", "Por Tipo de Operação",
        "Resumo por Categoria", "Evolução Mensal", "Total por Dia", "Resumo por Filial",
    ]

    # Aba "Financeiro": A=Filial B=Número ... G=Fornecedor H=Categoria ...
    # L=Valor M=Saldo ... P=Status
    ws1 = wb["Financeiro"]
    assert ws1["A1"].value == "Filial"
    assert ws1["A2"].value == "01"
    assert ws1["G2"].value == "Fornecedor Exemplo Ltda"
    # sem regra correspondente no fornecedor de teste -> "Não Classificado"
    assert ws1["H2"].value == "Não Classificado"
    assert ws1["L2"].value == pytest.approx(1000.0)
    # vencimento_real = hoje, sem data_baixa -> "Em aberto"
    assert ws1["P2"].value == "Em aberto"

    ws2 = wb["Resumo"]
    assert ws2.cell(4, 1).value == "Total de Títulos"
    assert ws2.cell(4, 2).value == 1

    ws3 = wb["Por Fornecedor"]
    assert ws3["A2"].value == "Fornecedor Exemplo Ltda"

    ws4 = wb["Por Tipo de Operação"]
    assert ws4["A2"].value == "Compra de Mercadoria"

    ws5 = wb["Resumo por Categoria"]
    assert ws5.cell(5, 1).value == "Não Classificado"

    ws7 = wb["Total por Dia"]
    assert ws7.cell(5, 1).value is not None  # data de vencimento formatada

    # "Resumo por Filial" usa hdr_row=3 (diferente das outras abas, que
    # usam 4) — réplica fiel de hdr_row7 em client/app_financeiro.py.
    ws8 = wb["Resumo por Filial"]
    assert ws8.cell(3, 1).value == "Filial"
    assert ws8.cell(4, 1).value == "01"


def test_export_categoriza_fornecedor_conhecido(client, auth_headers, db_session):
    # "TOTVS" está nas regras reais extraídas de client/categorias.xlsx
    # (ver app/excel/data/categorias_financeiro.json) — confirma que a
    # categorização de verdade (não só o fallback "Não Classificado") funciona.
    db_session.add_all([
        _fornecedor(codigo="000124", nome="TOTVS"),
        _titulo(fornecedor="000124", nome_fornecedor="TOTVS"),
    ])
    db_session.commit()

    resposta = client.get("/financeiro/export", headers=auth_headers)
    ws1 = openpyxl.load_workbook(io.BytesIO(resposta.content))["Financeiro"]
    assert ws1["H2"].value == "T.I."


def test_export_aplica_filtros_locais(client, auth_headers, db_session):
    db_session.add_all([
        _fornecedor(codigo="000123", nome="Fornecedor Um"),
        _tipo_operacao(),
        _titulo(numero="000001", filial="01", fornecedor="000123",
                nome_fornecedor="Fornecedor Um", tipo="NF"),
        _fornecedor(codigo="000124", nome="Fornecedor Dois"),
        _titulo(numero="000002", filial="02", fornecedor="000124",
                nome_fornecedor="Fornecedor Dois", tipo="NF"),
    ])
    db_session.commit()

    resposta_filial = client.get("/financeiro/export?filial=01", headers=auth_headers)
    ws_filial = openpyxl.load_workbook(io.BytesIO(resposta_filial.content))["Financeiro"]
    filiais = [ws_filial.cell(r, 1).value for r in range(2, ws_filial.max_row + 1)
               if ws_filial.cell(r, 1).value]
    assert filiais == ["01"]

    resposta_forn = client.get("/financeiro/export?fornecedor=dois", headers=auth_headers)
    ws_forn = openpyxl.load_workbook(io.BytesIO(resposta_forn.content))["Financeiro"]
    fornecedores = [ws_forn.cell(r, 7).value for r in range(2, ws_forn.max_row + 1)
                    if ws_forn.cell(r, 7).value]
    assert fornecedores == ["Fornecedor Dois"]
