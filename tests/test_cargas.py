import io
from datetime import date, timedelta

import openpyxl
import pytest

from app.models.cliente import Cliente
from app.models.item_carga import ItemCarga
from app.models.motorista import Motorista
from app.models.nota_fiscal_saida import NotaFiscalSaida
from app.models.veiculo_carga import VeiculoCarga


def _cliente(**overrides):
    dados = dict(
        deletado=" ",
        filial="01",
        codigo="000456",
        loja="01",
        nome="Cliente Exemplo Ltda",
        bairro="Centro",
        municipio="Sao Paulo",
    )
    dados.update(overrides)
    return Cliente(**dados)


def _veiculo(**overrides):
    dados = dict(
        deletado=" ",
        filial="01",
        codigo="000001",
        sequencia_carga="001",
        # data_inicial/data_final filtram por VeiculoCarga.data (DAK_DATA),
        # não ItemCarga.data — ver app/routers/cargas.py.
        data=date.today().strftime("%Y%m%d"),
        caminhao="ABC1234",
        status="1",
        motorista="M0001",
    )
    dados.update(overrides)
    return VeiculoCarga(**dados)


def _item(**overrides):
    dados = dict(
        deletado=" ",
        filial="01",
        codigo="000001",
        sequencia_carga="001",
        sequencia="000001",
        data=date.today().strftime("%Y%m%d"),
        pedido="000123",
        cliente="000456",
        loja="01",
        peso="1000.0000",
        nota_fiscal="000789",
        serie="1",
    )
    dados.update(overrides)
    return ItemCarga(**dados)


def _motorista(**overrides):
    dados = dict(
        deletado=" ",
        filial="01",
        codigo="M0001",
        nome="Motorista Exemplo",
    )
    dados.update(overrides)
    return Motorista(**dados)


def _nota_fiscal(**overrides):
    dados = dict(
        deletado=" ",
        filial="01",
        prefixo="1",
        numero="000789",
        cliente="000456",
        loja="01",
        valor="1500.00",
        parcela="  ",
        tipo="NF ",
        carga="000001",
        sequencia_carga="001",
    )
    dados.update(overrides)
    return NotaFiscalSaida(**dados)


def test_requer_api_key(client):
    assert client.get("/cargas/").status_code == 401


def test_lista_com_join_de_veiculo_e_cliente(client, auth_headers, db_session):
    db_session.add_all([_cliente(), _veiculo(), _item(), _nota_fiscal(), _motorista()])
    db_session.commit()

    resposta = client.get("/cargas/", headers=auth_headers)
    assert resposta.status_code == 200
    dados = resposta.json()
    assert len(dados["items"]) == 1
    item = dados["items"][0]
    assert item["codigo"] == "000001"
    assert item["nome_cliente"] == "Cliente Exemplo Ltda"
    assert item["bairro_cliente"] == "Centro"
    assert item["municipio_cliente"] == "Sao Paulo"
    assert item["caminhao"] == "ABC1234"
    # status "1" não está em ('7','8') -> case() com ELSE classifica como "Aberta".
    assert item["status_carga"] == "Aberta"
    assert item["valor"] == "1500.00"
    assert item["motorista"] == "Motorista Exemplo"


def test_caminhao_placa_cliente_e_status_fechada_sao_formatados(client, auth_headers, db_session):
    # select_cargas.sql calcula CAMINHAO ("Cliente" no lugar da placa
    # KHA0902) e STATUS ("Fechada"/"Aberta") com CASE — replicado em
    # app/routers/cargas.py com case() do SQLAlchemy.
    db_session.add_all(
        [
            _cliente(),
            _veiculo(caminhao="KHA0902", status="8"),
            _item(),
        ]
    )
    db_session.commit()

    resposta = client.get("/cargas/", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["caminhao"] == "Cliente"
    assert dados["items"][0]["status_carga"] == "Fechada"


def test_item_sem_motorista_correspondente_ainda_aparece(client, auth_headers, db_session):
    # DA4070 (motorista) é LEFT JOIN: sem motorista cadastrado no veículo, o
    # item aparece normalmente com motorista nulo.
    db_session.add_all([_cliente(), _veiculo(), _item()])
    db_session.commit()

    resposta = client.get("/cargas/", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["motorista"] is None


def test_valor_vem_da_nota_fiscal_nao_do_veiculo(client, auth_headers, db_session):
    # select_cargas.sql busca o valor em SE1070 (nota fiscal), casando
    # filial/série/número/cliente/loja — não é um campo de DAK070.
    db_session.add_all(
        [
            _cliente(),
            _veiculo(),
            _item(serie="2", nota_fiscal="000999"),
            _nota_fiscal(prefixo="2", numero="000999", valor="777.77"),
            _nota_fiscal(prefixo="1", numero="000789", valor="1500.00"),  # não deve casar
        ]
    )
    db_session.commit()

    resposta = client.get("/cargas/", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["valor"] == "777.77"


def test_valor_e_zero_sem_nota_fiscal_correspondente(client, auth_headers, db_session):
    db_session.add_all([_cliente(), _veiculo(), _item()])
    db_session.commit()

    resposta = client.get("/cargas/", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["valor"] == "0.00"


def test_filtro_status_encerrada_agrupa_codigos_7_e_8(client, auth_headers, db_session):
    db_session.add_all(
        [
            _cliente(),
            _veiculo(codigo="000001", sequencia_carga="001", status="1"),
            _item(codigo="000001", sequencia_carga="001"),
            _veiculo(codigo="000002", sequencia_carga="001", status="7"),
            _item(codigo="000002", sequencia_carga="001"),
            _veiculo(codigo="000003", sequencia_carga="001", status="8"),
            _item(codigo="000003", sequencia_carga="001"),
        ]
    )
    db_session.commit()

    resposta = client.get("/cargas/?status=encerrada", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 2
    assert {item["codigo"] for item in dados["items"]} == {"000002", "000003"}


def test_filtro_status_aberta_agrupa_demais_codigos(client, auth_headers, db_session):
    db_session.add_all(
        [
            _cliente(),
            _veiculo(codigo="000001", sequencia_carga="001", status="1"),
            _item(codigo="000001", sequencia_carga="001"),
            _veiculo(codigo="000002", sequencia_carga="001", status="6"),
            _item(codigo="000002", sequencia_carga="001"),
            _veiculo(codigo="000003", sequencia_carga="001", status="7"),
            _item(codigo="000003", sequencia_carga="001"),
        ]
    )
    db_session.commit()

    resposta = client.get("/cargas/?status=aberta", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 2
    assert {item["codigo"] for item in dados["items"]} == {"000001", "000002"}


def test_filtro_aberta_inclui_status_nulo_e_codigo_desconhecido(client, auth_headers, db_session):
    # VeiculoCarga.status.notin_(("7","8")) sozinho nunca é verdadeiro quando
    # a coluna é NULL (lógica de três valores do SQL) — sem o or_(is_(None)),
    # essa carga sumiria tanto de "aberta" quanto de "encerrada", mesmo não
    # estando em ('7','8'). status_carga (case() com ELSE 'Aberta') já
    # classifica NULL — e qualquer código fora de ('7','8'), como "9" — como
    # "Aberta"; o filtro precisa ser consistente com isso.
    db_session.add_all(
        [
            _cliente(),
            _veiculo(codigo="000001", sequencia_carga="001", status=None),
            _item(codigo="000001", sequencia_carga="001"),
            _veiculo(codigo="000002", sequencia_carga="001", status="9"),
            _item(codigo="000002", sequencia_carga="001"),
        ]
    )
    db_session.commit()

    sem_filtro = client.get("/cargas/", headers=auth_headers).json()
    assert len(sem_filtro["items"]) == 2
    assert {item["status_carga"] for item in sem_filtro["items"]} == {"Aberta"}

    resposta_aberta = client.get("/cargas/?status=aberta", headers=auth_headers).json()
    assert len(resposta_aberta["items"]) == 2

    resposta_encerrada = client.get("/cargas/?status=encerrada", headers=auth_headers).json()
    assert len(resposta_encerrada["items"]) == 0


def test_filtro_status_invalido_retorna_422(client, auth_headers):
    resposta = client.get("/cargas/?status=fechada", headers=auth_headers)
    assert resposta.status_code == 422


def test_filtro_data_final_via_query_string(client, auth_headers, db_session):
    db_session.add_all(
        [
            _cliente(),
            _veiculo(codigo="000001", sequencia_carga="001", data="20260805"),
            _item(codigo="000001", sequencia_carga="001"),
            _veiculo(codigo="000002", sequencia_carga="001", data="20260815"),
            _item(codigo="000002", sequencia_carga="001"),
        ]
    )
    db_session.commit()

    resposta = client.get(
        "/cargas/?data_inicial=20260801&data_final=20260810", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["codigo"] == "000001"


def test_filtro_data_usa_veiculo_nao_item(client, auth_headers, db_session):
    # data_inicial/data_final filtram por VeiculoCarga.data (DAK_DATA), não
    # ItemCarga.data (DAI_DATA) — o campo `data` retornado no item continua
    # vindo de DAI070; só o predicado do filtro migrou de tabela (ver
    # app/routers/cargas.py). DAI_DATA do item aqui está bem fora do período
    # pedido; se o filtro estivesse (por engano) em ItemCarga.data, esta
    # carga não apareceria.
    db_session.add_all(
        [
            _cliente(),
            _veiculo(data="20260805"),
            _item(data="20250101"),
        ]
    )
    db_session.commit()

    resposta = client.get(
        "/cargas/?data_inicial=20260801&data_final=20260810", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["data"] == "20250101"


def test_item_sem_veiculo_correspondente_e_ignorado(client, auth_headers, db_session):
    db_session.add_all([_cliente(), _item()])
    db_session.commit()

    resposta = client.get("/cargas/", headers=auth_headers)
    assert len(resposta.json()["items"]) == 0


def test_item_sem_cliente_correspondente_e_ignorado(client, auth_headers, db_session):
    db_session.add_all([_veiculo(), _item()])
    db_session.commit()

    resposta = client.get("/cargas/", headers=auth_headers)
    assert len(resposta.json()["items"]) == 0


def test_sequencia_cancelada_e_ignorada(client, auth_headers, db_session):
    db_session.add_all([_cliente(), _veiculo(), _item(sequencia="999999")])
    db_session.commit()

    resposta = client.get("/cargas/", headers=auth_headers)
    assert len(resposta.json()["items"]) == 0


def test_registros_deletados_sao_ignorados(client, auth_headers, db_session):
    db_session.add_all([_cliente(), _veiculo(), _item(deletado="*")])
    db_session.commit()

    resposta = client.get("/cargas/", headers=auth_headers)
    assert len(resposta.json()["items"]) == 0


def test_sem_filtro_assume_data_atual(client, auth_headers, db_session):
    hoje = date.today().strftime("%Y%m%d")
    ontem = (date.today() - timedelta(days=1)).strftime("%Y%m%d")
    db_session.add_all(
        [
            _cliente(),
            _veiculo(codigo="000001", sequencia_carga="001", data=ontem),
            _item(codigo="000001", sequencia_carga="001"),
            _veiculo(codigo="000002", sequencia_carga="001", data=hoje),
            _item(codigo="000002", sequencia_carga="001"),
        ]
    )
    db_session.commit()

    resposta = client.get("/cargas/", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["codigo"] == "000002"


def test_filtro_data_inicial_via_query_string(client, auth_headers, db_session):
    db_session.add_all(
        [
            _cliente(),
            _veiculo(codigo="000001", sequencia_carga="001", data="20260101"),
            _item(codigo="000001", sequencia_carga="001"),
            _veiculo(codigo="000002", sequencia_carga="001", data="20260801"),
            _item(codigo="000002", sequencia_carga="001"),
        ]
    )
    db_session.commit()

    resposta = client.get("/cargas/?data_inicial=20260801", headers=auth_headers)
    dados = resposta.json()
    assert len(dados["items"]) == 1
    assert dados["items"][0]["codigo"] == "000002"


def test_metodos_de_escrita_nao_existem(client, auth_headers):
    assert client.post("/cargas/", json={}, headers=auth_headers).status_code == 405
    assert client.put("/cargas/", json={}, headers=auth_headers).status_code == 405
    assert client.delete("/cargas/", headers=auth_headers).status_code == 405


def test_export_gera_planilha_com_4_abas_e_valores_corretos(client, auth_headers, db_session):
    db_session.add_all([_cliente(), _veiculo(), _item(), _nota_fiscal(), _motorista()])
    db_session.commit()

    resposta = client.get("/cargas/export", headers=auth_headers)
    assert resposta.status_code == 200
    assert resposta.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "filename=" in resposta.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(resposta.content))
    assert wb.sheetnames == ["Cargas", "Resumo", "Por Filial", "Top Clientes"]

    # Aba "Cargas": A=Filial B=Carga C=Data D=Pedido E=Motorista F=Cliente
    # ... L=Peso M=Valor
    ws1 = wb["Cargas"]
    assert ws1["A1"].value == "Filial"
    assert ws1["A2"].value == "01"
    assert ws1["B2"].value == "000001"
    assert ws1["F2"].value == "Cliente Exemplo Ltda"
    assert ws1["L2"].value == pytest.approx(1000.0)
    assert ws1["M2"].value == pytest.approx(1500.0)

    ws2 = wb["Resumo"]
    resumo = {ws2.cell(i, 1).value: ws2.cell(i, 2).value for i in range(4, 11)}
    assert resumo["Total de Cargas"] == 1
    assert resumo["Valor Total (R$)"] == pytest.approx(1500.0)
    # status "1" -> "Aberta" (case() com ELSE) -> tudo em aberto, nada acertado
    assert resumo["Valor em Aberto (R$)"] == pytest.approx(1500.0)
    assert resumo["Valor Acertado (R$)"] == pytest.approx(0.0)

    ws3 = wb["Por Filial"]
    assert ws3["A2"].value == "01"
    assert ws3["B2"].value == 1
    assert ws3["A3"].value == "TOTAL GERAL"

    ws4 = wb["Top Clientes"]
    assert ws4["A2"].value == "Cliente Exemplo Ltda"


def test_export_conta_cargas_distintas_nao_linhas(client, auth_headers, db_session):
    # Uma carga (mesma filial+codigo) com 2 itens deve contar como 1 carga
    # no resumo/por filial, não 2 — cada linha da API é um item, não uma
    # carga (ver comentário em app/excel/cargas.py).
    db_session.add_all([
        _cliente(),
        _veiculo(codigo="000001", sequencia_carga="001"),
        _item(codigo="000001", sequencia_carga="001", sequencia="000001",
              pedido="000001", peso="1000.0000"),
        _item(codigo="000001", sequencia_carga="001", sequencia="000002",
              pedido="000002", peso="500.0000"),
        _nota_fiscal(),
        _motorista(),
    ])
    db_session.commit()

    resposta = client.get("/cargas/export", headers=auth_headers)
    wb = openpyxl.load_workbook(io.BytesIO(resposta.content))
    ws1 = wb["Cargas"]
    assert ws1.max_row == 3  # cabeçalho + 2 itens

    ws2 = wb["Resumo"]
    resumo = {ws2.cell(i, 1).value: ws2.cell(i, 2).value for i in range(4, 11)}
    assert resumo["Total de Cargas"] == 1
    assert resumo["Pedidos Distintos"] == 2
    assert resumo["Peso Total (kg)"] == pytest.approx(1500.0)


def test_export_aplica_filtros_locais_de_filial_cliente_e_caminhao(client, auth_headers, db_session):
    db_session.add_all([
        _cliente(codigo="000456", nome="Cliente Um"),
        _veiculo(codigo="000001", sequencia_carga="001", caminhao="AAA1111"),
        _item(codigo="000001", sequencia_carga="001", filial="01"),
        _cliente(codigo="000457", nome="Cliente Dois", filial="02"),
        _veiculo(codigo="000002", sequencia_carga="001", filial="02", caminhao="BBB2222"),
        _item(codigo="000002", sequencia_carga="001", filial="02", cliente="000457",
              nota_fiscal="000790"),
        _nota_fiscal(),
        _nota_fiscal(numero="000790", cliente="000457", carga="000002"),
        _motorista(),
    ])
    db_session.commit()

    resposta_filial = client.get("/cargas/export?filial=01", headers=auth_headers)
    ws_filial = openpyxl.load_workbook(io.BytesIO(resposta_filial.content))["Cargas"]
    filiais = [ws_filial.cell(r, 1).value for r in range(2, ws_filial.max_row + 1)
               if ws_filial.cell(r, 1).value]
    assert filiais == ["01"]

    resposta_cliente = client.get("/cargas/export?cliente=dois", headers=auth_headers)
    ws_cliente = openpyxl.load_workbook(io.BytesIO(resposta_cliente.content))["Cargas"]
    clientes = [ws_cliente.cell(r, 6).value for r in range(2, ws_cliente.max_row + 1)
                if ws_cliente.cell(r, 6).value]
    assert clientes == ["Cliente Dois"]

    resposta_caminhao = client.get("/cargas/export?caminhao=bbb", headers=auth_headers)
    ws_caminhao = openpyxl.load_workbook(io.BytesIO(resposta_caminhao.content))["Cargas"]
    caminhoes = [ws_caminhao.cell(r, 10).value for r in range(2, ws_caminhao.max_row + 1)
                 if ws_caminhao.cell(r, 10).value]
    assert caminhoes == ["BBB2222"]
