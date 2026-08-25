"""
Cliente desktop para a RestDataAPI — Relatório de Faturamento (endpoint
/faturamento).

Exibe seletor de Mês/Ano, filtros locais, KPIs, gráficos analíticos e
exporta tudo para Excel.

O endpoint /faturamento agrega notas fiscais de saída de produtos acabados
(tipo 'PA') por filial, dia do mês e produto — aplica sempre como regra
fixa apenas os tipos de operação 501 (venda) e 542/543/544 (bonificação).
`dia` é o dia do mês da emissão (1 a 31), não uma data completa: por isso
esta tela sempre consulta a API por um único mês por vez (Mês/Ano abaixo
define data_inicial/data_final como o primeiro e o último dia do mês
escolhido) — consultar mais de um mês de uma vez faria dias iguais de
meses diferentes aparecerem somados na mesma linha (comportamento da API,
não deste cliente). Os demais filtros (filial, produto) são aplicados no
cliente, após o carregamento completo dos dados.

Requisitos:  pip install requests matplotlib pandas openpyxl python-dotenv
"""

from __future__ import annotations

import calendar
import threading
from datetime import date, datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
from typing import Dict, List, Optional, Tuple

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.ticker as mticker
import pandas as pd
from PIL import Image, ImageTk

from api_client import APIClient
from config import API_BASE_URL, API_KEY, API_KEY_NAME

LOGO_PATH = Path(__file__).resolve().parent / "logo.jpg"
LOGO_ALTURA_PX = 46


# ── constantes ───────────────────────────────────────────────────────────────

CHART_COLORS = [
    "#2980b9", "#e74c3c", "#27ae60", "#f39c12", "#8e44ad",
    "#16a085", "#d35400", "#2c3e50", "#c0392b", "#1abc9c",
]

MESES_NOMES = [
    "01 - Janeiro", "02 - Fevereiro", "03 - Março", "04 - Abril",
    "05 - Maio", "06 - Junho", "07 - Julho", "08 - Agosto",
    "09 - Setembro", "10 - Outubro", "11 - Novembro", "12 - Dezembro",
]

TREEVIEW_COLS: List[Tuple[str, str, int, str]] = [
    ("filial",       "Filial",           60,  "c"),
    ("dia",          "Dia",              50,  "c"),
    ("codigo",       "Código",          110,  "c"),
    ("descricao",    "Descrição",       300,  "w"),
    ("quantidade",   "Quantidade",      110,  "e"),
    ("faturamento",  "Faturamento (R$)", 130,  "e"),
    ("preco_medio",  "Preço Médio (R$)", 120,  "e"),
    ("lucro_bruto",  "Lucro Bruto (R$)", 130,  "e"),
    ("margem",       "Margem (%)",       90,  "e"),
]


# ── helpers ──────────────────────────────────────────────────────────────────

def _brl(v) -> str:
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"


def _qtd(v) -> str:
    try:
        return f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"


def _pct(v) -> str:
    try:
        return f"{float(v):,.1f}%".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"


def _periodo_do_mes(ano: int, mes: int) -> Tuple[str, str]:
    """(data_inicial, data_final) — primeiro e último dia do mês, formato
    AAAAMMDD. `dia` na API é só o dia do mês (não uma data completa), então
    este client sempre consulta um único mês por vez, para não misturar
    dias iguais de meses diferentes num mesmo grupo (ver módulo)."""
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    return f"{ano:04d}{mes:02d}01", f"{ano:04d}{mes:02d}{ultimo_dia:02d}"


# ── aplicação principal ───────────────────────────────────────────────────────

class FaturamentoApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Relatório de Faturamento — Grupo LGH (sem o Broker)")
        self.root.geometry("1300x1100")
        self.root.minsize(980, 640)
        self.root.configure(bg="#f0f2f5")

        self._df_raw: Optional[pd.DataFrame] = None
        self._df: Optional[pd.DataFrame] = None
        self._loading = False
        self._sort_ascending: Dict[str, bool] = {}

        self._setup_styles()
        self._build_ui()

    # ── estilos ──────────────────────────────────────────────────────────────

    def _setup_styles(self) -> None:
        s = ttk.Style(self.root)
        s.theme_use("clam")

        BG, WHITE, PRIMARY, ACCENT = "#f0f2f5", "#ffffff", "#1a5276", "#2980b9"

        s.configure("TFrame", background=BG)
        s.configure("White.TFrame", background=WHITE)
        s.configure("TLabel", background=BG, font=("Segoe UI", 10))
        s.configure("TLabelframe", background=BG, font=("Segoe UI", 10, "bold"))
        s.configure("TLabelframe.Label", background=BG, font=("Segoe UI", 10, "bold"),
                    foreground=PRIMARY)

        for name, bg, fg, hover in [
            ("Primary",   PRIMARY,   WHITE, ACCENT),
            ("Secondary", "#7f8c8d", WHITE, "#95a5a6"),
            ("Success",   "#27ae60", WHITE, "#2ecc71"),
        ]:
            s.configure(f"{name}.TButton",
                        background=bg, foreground=fg,
                        font=("Segoe UI", 10, "bold"), padding=(12, 5))
            s.map(f"{name}.TButton", background=[("active", hover), ("pressed", bg)])

        s.configure("TEntry", fieldbackground=WHITE, padding=4)
        s.configure("TNotebook", background=BG)
        s.configure("TNotebook.Tab", padding=(12, 6), font=("Segoe UI", 10))

        s.configure("Treeview", background=WHITE, fieldbackground=WHITE,
                    font=("Segoe UI", 10), rowheight=25)
        s.configure("Treeview.Heading",
                    font=("Segoe UI", 10, "bold"),
                    background=PRIMARY, foreground=WHITE)
        s.map("Treeview",
              background=[("selected", ACCENT)],
              foreground=[("selected", WHITE)])

    # ── construção da UI ─────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        self._build_header_frame()
        self._build_periodo_frame()
        self._build_filter_frame()
        self._build_kpi_frame()
        self._build_toolbar()
        self._build_statusbar()
        self._build_notebook()

    # ── cabeçalho (logotipo + conexão, na mesma linha) ──────────────────────

    def _build_header_frame(self) -> None:
        row = tk.Frame(self.root, bg="#f0f2f5")
        row.pack(fill="x")

        logo_holder = tk.Frame(row, bg="#ffffff",
                               highlightbackground="#d5d8dc", highlightthickness=1)
        logo_holder.pack(side="left", fill="y", padx=(12, 8), pady=(8, 3))
        try:
            img = Image.open(LOGO_PATH)
            largura = round(img.width * (LOGO_ALTURA_PX / img.height))
            img = img.resize((largura, LOGO_ALTURA_PX), Image.LANCZOS)
            self._logo_img = ImageTk.PhotoImage(img)
            tk.Label(logo_holder, image=self._logo_img, bg="#ffffff").pack(
                padx=10, pady=6)
        except Exception:
            pass

        self._build_conn_frame(row)

    # ── conexão ──────────────────────────────────────────────────────────────

    def _build_conn_frame(self, parent: tk.Frame) -> None:
        f = ttk.LabelFrame(parent, text="Conexão", padding=(10, 6))
        f.pack(side="left", fill="both", expand=True, padx=(0, 12), pady=(8, 3))

        ttk.Label(f, text="URL da API:").grid(row=0, column=0, sticky="w", padx=(0, 4))
        self._url_var = tk.StringVar(value=API_BASE_URL)
        ttk.Entry(f, textvariable=self._url_var, width=36).grid(
            row=0, column=1, sticky="ew", padx=(0, 18))

        ttk.Label(f, text="API Key:").grid(row=0, column=2, sticky="w", padx=(0, 4))
        self._key_var = tk.StringVar(value=API_KEY)
        ttk.Entry(f, textvariable=self._key_var, show="•", width=34).grid(
            row=0, column=3, sticky="ew", padx=(0, 12))

        ttk.Button(f, text="Testar conexão", style="Secondary.TButton",
                   command=self._test_connection).grid(row=0, column=4)

        f.columnconfigure(1, weight=1)
        f.columnconfigure(3, weight=1)

    # ── período (define data_inicial/data_final enviados à API) ────────────────

    def _build_periodo_frame(self) -> None:
        f = ttk.LabelFrame(
            self.root,
            text="Mês/Ano a analisar (define data_inicial/data_final enviados ao Protheus — "
                 "recarrega ao consultar)",
            padding=(10, 6),
        )
        f.pack(fill="x", padx=12, pady=3)

        hoje = date.today()

        ttk.Label(f, text="Mês:").grid(row=0, column=0, sticky="w", padx=(0, 6))
        self._mes_var = tk.StringVar(value=MESES_NOMES[hoje.month - 1])
        ttk.Combobox(
            f, textvariable=self._mes_var, width=14, state="readonly",
            values=MESES_NOMES,
        ).grid(row=0, column=1, sticky="w")

        ttk.Label(f, text="Ano:").grid(row=0, column=2, sticky="w", padx=(14, 6))
        anos = [str(a) for a in range(hoje.year - 5, hoje.year + 2)]
        self._ano_var = tk.StringVar(value=str(hoje.year))
        ttk.Combobox(
            f, textvariable=self._ano_var, width=8, state="readonly",
            values=anos,
        ).grid(row=0, column=3, sticky="w")

        ttk.Label(
            f,
            text='(o relatório agrega por "dia do mês" — a API sempre traz um único mês por vez)',
            foreground="#7f8c8d",
        ).grid(row=0, column=4, sticky="w", padx=(10, 0))

    # ── filtros (aplicados no cliente, após o carregamento) ────────────────────

    def _build_filter_frame(self) -> None:
        f = ttk.LabelFrame(
            self.root,
            text="Filtros (aplicados localmente sobre os dados carregados)",
            padding=(10, 6),
        )
        f.pack(fill="x", padx=12, pady=3)

        text_filters = [
            ("Filial:",   "_f_filial",  6),
            ("Produto:",  "_f_produto", 26),
        ]
        col = 0
        for label, attr, width in text_filters:
            ttk.Label(f, text=label).grid(
                row=0, column=col, sticky="w", padx=(0 if col == 0 else 14, 3))
            var = tk.StringVar()
            setattr(self, attr, var)
            ttk.Entry(f, textvariable=var, width=width).grid(
                row=0, column=col + 1, sticky="ew")
            col += 2

        ttk.Label(
            f, text="(Produto busca por código ou descrição)",
            foreground="#7f8c8d",
        ).grid(row=0, column=col, sticky="w", padx=(6, 0))
        col += 1

        # botões
        btn = ttk.Frame(f)
        btn.grid(row=0, column=col, sticky="e", padx=(20, 0))
        ttk.Button(btn, text=" Carregar do Protheus ", style="Primary.TButton",
                   command=self.consultar).pack(side="left", padx=(0, 6))
        ttk.Button(btn, text="Aplicar filtros", style="Secondary.TButton",
                   command=self._on_filtros_alterados).pack(side="left", padx=(0, 6))
        ttk.Button(btn, text="Limpar", style="Secondary.TButton",
                   command=self.limpar).pack(side="left")

        f.columnconfigure(col, weight=1)

    # ── KPI cards ─────────────────────────────────────────────────────────────

    def _build_kpi_frame(self) -> None:
        outer = tk.Frame(self.root, bg="#f0f2f5")
        outer.pack(fill="x", padx=12, pady=3)

        self._kpi_vars: Dict[str, tk.StringVar] = {}
        cards_cfg = [
            ("registros",      "Registros",           "#2c3e50"),
            ("produtos",       "Produtos",             "#7f8c8d"),
            ("qtd_total",      "Quantidade Total",     "#1a5276"),
            ("faturamento",    "Faturamento Total",    "#154360"),
            ("lucro_bruto",    "Lucro Bruto Total",    "#1e8449"),
            ("margem_geral",   "Margem Geral",         "#ca6f1e"),
        ]

        for i, (key, label, fg) in enumerate(cards_cfg):
            card = tk.Frame(outer, bg="#ffffff",
                            highlightbackground="#d5d8dc", highlightthickness=1)
            card.grid(row=0, column=i, sticky="nsew", padx=4, pady=2, ipady=10)

            tk.Label(card, text=label, bg="#ffffff",
                     font=("Segoe UI", 9), fg="#7f8c8d").pack(pady=(4, 0))

            var = tk.StringVar(value="—")
            self._kpi_vars[key] = var
            tk.Label(card, textvariable=var, bg="#ffffff",
                     font=("Segoe UI", 14, "bold"), fg=fg).pack()

        for i in range(len(cards_cfg)):
            outer.columnconfigure(i, weight=1)

    # ── toolbar acima do notebook ─────────────────────────────────────────────

    def _build_toolbar(self) -> None:
        bar = ttk.Frame(self.root)
        bar.pack(fill="x", padx=12, pady=(4, 0))
        ttk.Button(bar, text="  Exportar Excel  ", style="Success.TButton",
                   command=self.exportar_excel).pack(side="right")

    # ── notebook ──────────────────────────────────────────────────────────────

    def _build_notebook(self) -> None:
        self._nb = ttk.Notebook(self.root)
        self._nb.pack(fill="both", expand=True, padx=12, pady=(2, 4))

        tab_graficos = ttk.Frame(self._nb)
        self._nb.add(tab_graficos, text="  Gráficos  ")
        self._build_charts_tab(tab_graficos)

        tab_dados = ttk.Frame(self._nb)
        self._nb.add(tab_dados, text="  Dados  ")
        self._build_data_tab(tab_dados)

    def _build_charts_tab(self, parent: ttk.Frame) -> None:
        self._fig = Figure(figsize=(13, 6.4), dpi=100, facecolor="#f0f2f5")
        self._axes = self._fig.subplots(2, 2)
        self._fig.subplots_adjust(left=0.17, right=0.97, top=0.9,
                                  bottom=0.16, hspace=0.62, wspace=0.42)
        self._canvas = FigureCanvasTkAgg(self._fig, master=parent)
        self._canvas.get_tk_widget().pack(fill="both", expand=True, padx=4, pady=4)
        self._draw_empty_charts()

    def _build_data_tab(self, parent: ttk.Frame) -> None:
        cols = [c[0] for c in TREEVIEW_COLS]
        self._tree = ttk.Treeview(parent, columns=cols,
                                  show="headings", selectmode="browse")

        for col_id, heading, width, anchor in TREEVIEW_COLS:
            self._tree.heading(
                col_id, text=heading,
                command=lambda c=col_id: self._sort_tree(c),
            )
            self._tree.column(col_id, width=width, anchor=anchor,
                              stretch=(col_id == "descricao"))

        vsb = ttk.Scrollbar(parent, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        parent.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)

    # ── barra de status ───────────────────────────────────────────────────────

    def _build_statusbar(self) -> None:
        bar = tk.Frame(self.root, bg="#dde1e7", height=24)
        bar.pack(fill="x", side="bottom")

        self._status_var = tk.StringVar(value="Pronto.")
        tk.Label(bar, textvariable=self._status_var,
                 bg="#dde1e7", font=("Segoe UI", 9), anchor="w").pack(
            side="left", padx=8, pady=2)

        self._pct_var = tk.StringVar()
        tk.Label(bar, textvariable=self._pct_var,
                 bg="#dde1e7", font=("Segoe UI", 9)).pack(side="right", padx=4)

        self._progress = ttk.Progressbar(bar, mode="determinate", length=200)
        self._progress.pack(side="right", padx=8, pady=3)

    # ── conexão ───────────────────────────────────────────────────────────────

    def _make_client(self) -> APIClient:
        return APIClient(
            base_url=self._url_var.get().strip(),
            api_key=self._key_var.get().strip(),
            api_key_name=API_KEY_NAME,
        )

    def _test_connection(self) -> None:
        try:
            self._make_client().health()
            messagebox.showinfo("Conexão OK",
                                "A API respondeu com sucesso!")
        except Exception as exc:
            messagebox.showerror("Falha na conexão", str(exc))

    # ── consulta (thread) ─────────────────────────────────────────────────────

    def consultar(self) -> None:
        if self._loading:
            return

        mes = int(self._mes_var.get()[:2])
        ano = int(self._ano_var.get())
        data_inicial, data_final = _periodo_do_mes(ano, mes)

        self._loading = True
        self._clear_data()
        self._set_status(
            f"Consultando /faturamento na API… [{self._mes_var.get()} de {ano}]",
            indeterminate=True,
        )
        threading.Thread(
            target=self._fetch_thread, args=(data_inicial, data_final), daemon=True
        ).start()

    def _fetch_thread(self, data_inicial: str, data_final: str) -> None:
        try:
            client = self._make_client()
            items = client.get_all_faturamento(
                data_inicial=data_inicial, data_final=data_final,
                progress_callback=self._on_progress)
            self.root.after(0, self._on_data_ready, items)
        except Exception as exc:
            self.root.after(0, self._on_error, str(exc))

    def _on_progress(self, loaded: int) -> None:
        self.root.after(0, self._update_progress_bar, loaded)

    def _update_progress_bar(self, loaded: int) -> None:
        # A API não expõe o total de registros por página (um count() exato
        # sobre os mesmos joins/agregações da consulta é caro demais para
        # calcular a cada página — ver README): a barra fica indeterminada
        # (giratória, iniciada em _set_status) durante o carregamento; só o
        # contador de registros já lidos é atualizado.
        self._pct_var.set(f"{loaded} registro(s)")
        self._status_var.set(f"Carregando… {loaded} registro(s)")

    def _on_data_ready(self, items: List[Dict]) -> None:
        self._loading = False

        df = pd.DataFrame(items) if items else pd.DataFrame()
        if not df.empty:
            for col in ("quantidade", "faturamento", "preco_medio", "lucro_bruto", "margem"):
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
            df["dia"] = pd.to_numeric(df["dia"], errors="coerce").fillna(0).astype(int)

        self._df_raw = df
        self._apply_filters_and_refresh()

        n = len(self._df) if self._df is not None else 0
        total_api = len(items)
        self._set_status(
            f"{n} registro(s) exibido(s)"
            + (f"  [total carregado da API: {total_api}]" if n != total_api else "")
        )
        self._progress.stop()
        self._progress.configure(mode="determinate", value=100)
        self._pct_var.set("100%")

    def _on_error(self, msg: str) -> None:
        self._loading = False
        self._progress.stop()
        self._progress.configure(mode="determinate", value=0)
        self._pct_var.set("")
        self._set_status(f"Erro: {msg}")
        messagebox.showerror("Erro ao consultar a API", msg)

    # ── filtros (client-side) ────────────────────────────────────────────────

    def _on_filtros_alterados(self) -> None:
        if self._df_raw is None:
            messagebox.showinfo(
                "Filtros", "Carregue os dados da API antes de aplicar filtros.")
            return
        self._apply_filters_and_refresh()
        n = len(self._df) if self._df is not None else 0
        self._set_status(f"{n} registro(s) exibido(s) após filtros")

    def _apply_filters_and_refresh(self) -> None:
        df = self._df_raw
        if df is None or df.empty:
            self._df = df
        else:
            self._df = self._apply_filters(df)
        self._update_kpis()
        self._update_charts()
        self._update_table()

    def _apply_filters(self, df: pd.DataFrame) -> pd.DataFrame:
        filial = self._f_filial.get().strip()
        if filial:
            df = df[df["filial"] == filial]

        produto = self._f_produto.get().strip().lower()
        if produto:
            mask = (
                df["codigo"].fillna("").str.lower().str.contains(produto)
                | df["descricao"].fillna("").str.lower().str.contains(produto)
            )
            df = df[mask]

        return df

    # ── KPIs ──────────────────────────────────────────────────────────────────

    def _update_kpis(self) -> None:
        df = self._df
        if df is None or df.empty:
            for v in self._kpi_vars.values():
                v.set("—")
            return

        faturamento_total = float(df["faturamento"].sum())
        lucro_total = float(df["lucro_bruto"].sum())
        # Margem geral recalculada sobre os totais do recorte carregado (não
        # é a média da coluna "margem" — isso daria peso igual a grupos
        # pequenos e grandes; mesma lógica que a própria API usa por grupo).
        margem_geral = (lucro_total / faturamento_total * 100) if faturamento_total else 0.0

        self._kpi_vars["registros"].set(f'{len(df):,}'.replace(",", "."))
        self._kpi_vars["produtos"].set(f'{df["codigo"].nunique():,}'.replace(",", "."))
        self._kpi_vars["qtd_total"].set(_qtd(df["quantidade"].sum()))
        self._kpi_vars["faturamento"].set(_brl(faturamento_total))
        self._kpi_vars["lucro_bruto"].set(_brl(lucro_total))
        self._kpi_vars["margem_geral"].set(_pct(margem_geral))

    # ── gráficos ──────────────────────────────────────────────────────────────

    def _draw_empty_charts(self) -> None:
        for ax in self._axes.flat:
            ax.clear()
            ax.text(0.5, 0.5, "Sem dados — clique em \"Carregar do Protheus\"",
                    ha="center", va="center", color="#aab7b8", fontsize=10.5)
            ax.set_facecolor("#f8f9fa")
            ax.axis("off")
        self._canvas.draw()

    def _update_charts(self) -> None:
        for ax in self._axes.flat:
            ax.clear()

        df = self._df
        if df is None or df.empty:
            self._draw_empty_charts()
            return

        ax_top_fat, ax_filial = self._axes[0]
        ax_dia,     ax_top_lucro = self._axes[1]

        brl_fmt = mticker.FuncFormatter(
            lambda x, _: f"R${x/1000:.0f}k" if abs(x) >= 1000 else f"R${x:.0f}"
        )

        # ── Gráfico 1: barras horizontais — Top 10 Produtos por Faturamento ─
        rotulo = df["codigo"] + " — " + df["descricao"].fillna("").str.slice(0, 24)
        top_fat = (
            df.assign(rotulo=rotulo)
            .groupby("rotulo")["faturamento"]
            .sum()
            .nlargest(10)
            .sort_values(ascending=True)
        )
        bar_colors = CHART_COLORS[:len(top_fat)]
        hbars = ax_top_fat.barh(
            top_fat.index, top_fat.values,
            color=bar_colors[::-1], height=0.65, edgecolor="none",
        )
        ax_top_fat.set_title("Top 10 Produtos — Faturamento",
                             fontsize=13, fontweight="bold", pad=12)
        ax_top_fat.set_xlabel("Faturamento (R$)", fontsize=10.5)
        ax_top_fat.tick_params(axis="y", labelsize=9)
        ax_top_fat.tick_params(axis="x", labelsize=9)
        ax_top_fat.xaxis.set_major_formatter(brl_fmt)
        ax_top_fat.set_facecolor("#fafafa")
        ax_top_fat.margins(x=0.2)
        for bar in hbars:
            w = bar.get_width()
            ax_top_fat.text(
                w * 1.01, bar.get_y() + bar.get_height() / 2,
                _brl(w), va="center", fontsize=9, color="#555",
            )

        # ── Gráfico 2: pizza — Faturamento Total por Filial ────────────────
        by_filial = df.groupby("filial")["faturamento"].sum().sort_values(ascending=False)
        pie_colors = CHART_COLORS[:len(by_filial)]
        wedges, texts, autotexts = ax_filial.pie(
            by_filial,
            labels=by_filial.index,
            autopct="%1.1f%%",
            colors=pie_colors,
            startangle=90,
            textprops={"fontsize": 9.5},
            wedgeprops={"edgecolor": "white", "linewidth": 1.5},
        )
        for at in autotexts:
            at.set_fontsize(9)
        ax_filial.set_title("Faturamento Total por Filial",
                            fontsize=13, fontweight="bold", pad=12)

        # ── Gráfico 3: barras — Faturamento por Dia do Mês ─────────────────
        by_dia = df.groupby("dia")["faturamento"].sum().sort_index()
        ax_dia.bar(by_dia.index, by_dia.values,
                  color="#2980b9", width=0.7, edgecolor="none")
        ax_dia.set_title("Faturamento por Dia do Mês",
                         fontsize=13, fontweight="bold", pad=12)
        ax_dia.set_xlabel("Dia", fontsize=10.5)
        ax_dia.set_ylabel("Faturamento (R$)", fontsize=10.5)
        ax_dia.tick_params(axis="both", labelsize=9)
        ax_dia.yaxis.set_major_formatter(brl_fmt)
        ax_dia.set_facecolor("#fafafa")

        # ── Gráfico 4: barras horizontais — Top 10 Produtos por Lucro Bruto ─
        top_lucro = (
            df.assign(rotulo=rotulo)
            .groupby("rotulo")["lucro_bruto"]
            .sum()
            .nlargest(10)
            .sort_values(ascending=True)
        )
        bar_colors_lucro = CHART_COLORS[:len(top_lucro)]
        hbars_lucro = ax_top_lucro.barh(
            top_lucro.index, top_lucro.values,
            color=bar_colors_lucro[::-1], height=0.65, edgecolor="none",
        )
        ax_top_lucro.set_title("Top 10 Produtos — Lucro Bruto",
                               fontsize=13, fontweight="bold", pad=12)
        ax_top_lucro.set_xlabel("Lucro Bruto (R$)", fontsize=10.5)
        ax_top_lucro.tick_params(axis="y", labelsize=9)
        ax_top_lucro.tick_params(axis="x", labelsize=9)
        ax_top_lucro.xaxis.set_major_formatter(brl_fmt)
        ax_top_lucro.set_facecolor("#fafafa")
        ax_top_lucro.margins(x=0.2)
        for bar in hbars_lucro:
            w = bar.get_width()
            ax_top_lucro.text(
                w * 1.01, bar.get_y() + bar.get_height() / 2,
                _brl(w), va="center", fontsize=9, color="#555",
            )

        self._canvas.draw()

    # ── tabela ────────────────────────────────────────────────────────────────

    def _update_table(self) -> None:
        for row in self._tree.get_children():
            self._tree.delete(row)

        df = self._df
        if df is None or df.empty:
            return

        # itertuples() em vez de iterrows() (bem mais rápido em bases
        # grandes — mesmo cuidado já aplicado nos outros clients deste
        # projeto, ver app_cargas.py).
        for row in df.itertuples(index=False):
            self._tree.insert("", "end", values=(
                row.filial,
                int(row.dia),
                row.codigo,
                row.descricao,
                _qtd(row.quantidade),
                _brl(row.faturamento),
                _brl(row.preco_medio),
                _brl(row.lucro_bruto),
                _pct(row.margem),
            ))

    def _sort_tree(self, col: str) -> None:
        if self._df is None or self._df.empty:
            return
        asc = not self._sort_ascending.get(col, False)
        self._sort_ascending[col] = asc

        if col in ("dia", "quantidade", "faturamento", "preco_medio", "lucro_bruto", "margem"):
            self._df = self._df.sort_values(col, ascending=asc, na_position="last")
        else:
            self._df = self._df.sort_values(
                col, ascending=asc, na_position="last",
                key=lambda s: s.astype(str).str.lower(),
            )
        self._update_table()

    # ── exportar Excel ────────────────────────────────────────────────────────

    def exportar_excel(self) -> None:
        df = self._df
        if df is None or df.empty:
            messagebox.showwarning(
                "Exportar Excel",
                "Nenhum dado carregado. Faça uma consulta primeiro.",
            )
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("Todos os arquivos", "*.*")],
            title="Salvar planilha de Faturamento",
            initialfile=f"faturamento_{self._ano_var.get()}{self._mes_var.get()[:2]}.xlsx",
        )
        if not path:
            return

        try:
            self._set_status("Gerando planilha Excel…")
            self.root.update()
            _write_excel(df, path)
            self._set_status(f"Planilha salva em: {path}")
            messagebox.showinfo("Exportar Excel",
                                f"Arquivo salvo com sucesso:\n{path}")
        except Exception as exc:
            self._set_status("Erro ao exportar.")
            messagebox.showerror("Erro ao exportar", str(exc))

    # ── utilitários ───────────────────────────────────────────────────────────

    def limpar(self) -> None:
        for attr in ("_f_filial", "_f_produto"):
            getattr(self, attr).set("")
        if self._df_raw is not None:
            self._apply_filters_and_refresh()
        self._set_status("Filtros limpos.")

    def _clear_data(self) -> None:
        self._df_raw = None
        self._df = None
        for v in self._kpi_vars.values():
            v.set("—")
        for row in self._tree.get_children():
            self._tree.delete(row)
        self._draw_empty_charts()
        self._progress["value"] = 0
        self._pct_var.set("")

    def _set_status(self, msg: str, *, indeterminate: bool = False) -> None:
        self._status_var.set(msg)
        if indeterminate:
            self._progress.configure(mode="indeterminate")
            self._progress.start(12)
            self._pct_var.set("")
        else:
            self._progress.stop()
            self._progress.configure(mode="determinate")


# ── exportação Excel (função separada) ───────────────────────────────────────

def _write_excel(df: pd.DataFrame, path: str) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    HDR_FILL  = PatternFill("solid", fgColor="1A5276")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    ALT_FILL  = PatternFill("solid", fgColor="EBF5FB")
    THIN      = Side(style="thin", color="BDC3C7")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER    = Alignment(horizontal="center", vertical="center")
    RIGHT     = Alignment(horizontal="right",  vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")
    BRL       = '#,##0.00'
    QTD_FMT   = '#,##0.00'
    PCT       = '0.0%'
    BOLD10    = Font(bold=True, name="Calibri", size=10)

    def _hdr(ws, row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = BORDER
        return c

    # ── aba 1: Faturamento ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Faturamento"
    ws1.freeze_panes = "A2"

    cols_aba1 = [
        ("Filial",            "filial",       9),
        ("Dia",                "dia",          7),
        ("Código",             "codigo",      14),
        ("Descrição",         "descricao",    36),
        ("Quantidade",        "quantidade",   14),
        ("Faturamento (R$)",  "faturamento",  16),
        ("Preço Médio (R$)",  "preco_medio",  16),
        ("Lucro Bruto (R$)",  "lucro_bruto",  16),
        ("Margem (%)",        "margem",       12),
    ]
    for ci, (heading, _, w) in enumerate(cols_aba1, 1):
        _hdr(ws1, 1, ci, heading)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[1].height = 18
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(cols_aba1))}1"

    # itertuples() em vez de iterrows() — ver comentário equivalente em
    # _update_table().
    for ri, row in enumerate(df.itertuples(index=False), 2):
        row_fill = ALT_FILL if ri % 2 == 0 else None
        for ci, (_, field, _) in enumerate(cols_aba1, 1):
            val = getattr(row, field, None)
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if field in ("faturamento", "preco_medio", "lucro_bruto"):
                cell.number_format = BRL
                cell.alignment = RIGHT
            elif field == "quantidade":
                cell.number_format = QTD_FMT
                cell.alignment = RIGHT
            elif field == "margem":
                cell.number_format = '#,##0.00"%"'
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT
            if row_fill:
                cell.fill = row_fill

    # ── aba 2: Resumo ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 22

    ws2.cell(1, 1, "Resumo — Faturamento").font = Font(
        bold=True, color="1A5276", size=14, name="Calibri")
    ws2.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(
        italic=True, size=9, color="7F8C8D", name="Calibri")

    faturamento_total = float(df["faturamento"].sum())
    lucro_total = float(df["lucro_bruto"].sum())
    margem_geral = (lucro_total / faturamento_total) if faturamento_total else 0.0

    items_resumo = [
        ("Total de Registros",       len(df),                          None),
        ("Produtos Distintos",       int(df["codigo"].nunique()),      None),
        ("Filiais Distintas",        int(df["filial"].nunique()),      None),
        ("Quantidade Total",         float(df["quantidade"].sum()),    QTD_FMT),
        ("Faturamento Total (R$)",   faturamento_total,                BRL),
        ("Lucro Bruto Total (R$)",   lucro_total,                      BRL),
        ("Margem Geral (%)",         margem_geral,                     PCT),
    ]
    for i, (lbl, val, fmt) in enumerate(items_resumo, 4):
        c1 = ws2.cell(i, 1, lbl)
        c1.font = Font(bold=True, name="Calibri", size=10)
        c2 = ws2.cell(i, 2, val)
        c2.alignment = RIGHT
        if fmt:
            c2.number_format = fmt

    # ── aba 3: Por Filial ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Por Filial")
    by_filial = (
        df.groupby("filial", as_index=False)
        .agg(qtd_registros=("codigo", "count"),
             faturamento_total=("faturamento", "sum"),
             lucro_total=("lucro_bruto", "sum"))
        .sort_values("faturamento_total", ascending=False)
    )
    total_geral_filial = float(by_filial["faturamento_total"].sum())

    hdrs3 = ["Filial", "Qtd. Registros", "Faturamento (R$)", "Lucro Bruto (R$)", "% do Total"]
    widths3 = [12, 14, 18, 18, 12]
    for ci, (h, w) in enumerate(zip(hdrs3, widths3), 1):
        _hdr(ws3, 1, ci, h)
        ws3.column_dimensions[get_column_letter(ci)].width = w

    ri = 2
    for row in by_filial.itertuples(index=False):
        pct = (row.faturamento_total / total_geral_filial) if total_geral_filial else 0.0
        ws3.cell(ri, 1, row.filial)
        ws3.cell(ri, 2, int(row.qtd_registros))
        c3 = ws3.cell(ri, 3, float(row.faturamento_total)); c3.number_format = BRL
        c4 = ws3.cell(ri, 4, float(row.lucro_total)); c4.number_format = BRL
        c5 = ws3.cell(ri, 5, pct); c5.number_format = PCT
        if ri % 2 == 0:
            for ci in range(1, 6):
                ws3.cell(ri, ci).fill = ALT_FILL
        ri += 1

    ws3.cell(ri, 1, "TOTAL GERAL").font = BOLD10
    ws3.cell(ri, 2, int(by_filial["qtd_registros"].sum())).font = BOLD10
    c3 = ws3.cell(ri, 3, total_geral_filial); c3.number_format = BRL; c3.font = BOLD10
    c4 = ws3.cell(ri, 4, float(by_filial["lucro_total"].sum())); c4.number_format = BRL; c4.font = BOLD10
    c5 = ws3.cell(ri, 5, 1.0); c5.number_format = PCT; c5.font = BOLD10

    # ── aba 4: Top Produtos ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Top Produtos")
    by_produto = (
        df.groupby(["codigo", "descricao"], as_index=False, dropna=False)
        .agg(quantidade_total=("quantidade", "sum"),
             faturamento_total=("faturamento", "sum"),
             lucro_total=("lucro_bruto", "sum"))
        .sort_values("faturamento_total", ascending=False)
        .head(50)
    )
    hdrs4 = ["Código", "Descrição", "Quantidade Total", "Faturamento (R$)", "Lucro Bruto (R$)"]
    widths4 = [14, 36, 18, 18, 18]
    for ci, (h, w) in enumerate(zip(hdrs4, widths4), 1):
        _hdr(ws4, 1, ci, h)
        ws4.column_dimensions[get_column_letter(ci)].width = w
    for ri, row in enumerate(by_produto.itertuples(index=False), 2):
        ws4.cell(ri, 1, row.codigo)
        ws4.cell(ri, 2, row.descricao)
        c3 = ws4.cell(ri, 3, float(row.quantidade_total)); c3.number_format = QTD_FMT
        c4 = ws4.cell(ri, 4, float(row.faturamento_total)); c4.number_format = BRL
        c5 = ws4.cell(ri, 5, float(row.lucro_total)); c5.number_format = BRL
        if ri % 2 == 0:
            for ci in range(1, 6):
                ws4.cell(ri, ci).fill = ALT_FILL

    wb.save(path)


# ── entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    FaturamentoApp(root)
    root.mainloop()
