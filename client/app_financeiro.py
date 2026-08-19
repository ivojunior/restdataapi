"""
Cliente desktop para a RestDataAPI — Relatório Financeiro (endpoint /financeiro).

Exibe filtros, KPIs, gráficos analíticos, resumos por categoria/mês/filial
(mesma visão de CONTAS_A_PAGAR_ANALISE.xlsx) e exporta tudo para Excel.

O endpoint /financeiro exclui sempre os tipos de título PA/PR/NDF (regra
fixa). O período de vencimento (vencimento_de/vencimento_ate) e o status
("Em aberto"/"Vencido"/"Baixado") são enviados à API como query string — os
filtros "Vencimento De" e "Vencimento Até" já iniciam marcados com a data
atual do sistema (mesmo padrão que a API assume quando o parâmetro não é
enviado); o usuário pode alterar o período ou desmarcar antes de carregar.
Como o período padrão já começa hoje, para ver títulos "Vencido" é preciso
recuar "Vencimento De". Os demais filtros desta tela (filial, fornecedor,
tipo, tipo de operação e categoria) são aplicados no cliente, após o
carregamento completo dos dados.

A categoria de cada título não existe em nenhuma tabela do Protheus exposta
pela API — é resolvida no cliente a partir de regras editáveis mantidas em
categorias.xlsx (veja categorias.py).

Requisitos:  pip install requests matplotlib pandas openpyxl python-dotenv tkcalendar
"""

from __future__ import annotations

import threading
from datetime import date, datetime
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
from typing import Dict, List, Optional, Tuple

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from matplotlib.patches import Patch
import matplotlib.ticker as mticker
import pandas as pd
from tkcalendar import DateEntry

import categorias
from api_client import APIClient
from config import API_BASE_URL, API_KEY, API_KEY_NAME


# ── constantes ───────────────────────────────────────────────────────────────

TODAY = date.today().strftime("%Y%m%d")

STATUS_COLORS: Dict[str, str] = {
    "Em aberto": "#27ae60",
    "Vencido":   "#e74c3c",
    "Baixado":   "#95a5a6",
}

# Rótulo exibido -> valor do parâmetro "status" aceito pelo endpoint /financeiro.
STATUS_API_VALUES: Dict[str, str] = {
    "Em aberto": "em_aberto",
    "Vencido":   "vencido",
    "Baixado":   "baixado",
}
STATUS_FILTRO_PADRAO = "(Todos)"

CHART_COLORS = [
    "#2980b9", "#e74c3c", "#27ae60", "#f39c12", "#8e44ad",
    "#16a085", "#d35400", "#2c3e50", "#c0392b", "#1abc9c",
]

CATEGORIA_COLORS: Dict[str, str] = dict(
    zip(categorias.CATEGORIAS_CANONICAS, CHART_COLORS))

TREEVIEW_COLS: List[Tuple[str, str, int, str]] = [
    ("filial",             "Filial",         55,  "c"),
    ("numero",             "Número",         90,  "c"),
    ("parcela",            "Parcela",        60,  "c"),
    ("tipo",               "Tipo",           50,  "c"),
    ("codigo_operacao",    "Cód. Op.",       65,  "c"),
    ("descricao_operacao", "Tipo Operação", 150,  "w"),
    ("nome_fornecedor",    "Fornecedor",    180,  "w"),
    ("categoria",          "Categoria",     140,  "w"),
    ("emissao",            "Emissão",        85,  "c"),
    ("vencimento_real",    "Vencimento",     85,  "c"),
    ("valor",              "Valor (R$)",    115,  "e"),
    ("saldo",              "Saldo (R$)",    115,  "e"),
    ("historico",          "Histórico",     200,  "w"),
    ("recuperacao_judicial", "Rec. Judicial", 100, "c"),
    ("status",             "Status",         85,  "c"),
]


# ── helpers ──────────────────────────────────────────────────────────────────

def _brl(v) -> str:
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"


def _fmt_date(s) -> str:
    s = str(s or "").strip()
    if len(s) == 8 and s.isdigit():
        return f"{s[6:8]}/{s[4:6]}/{s[:4]}"
    return s or "—"


def _fmt_rec_judicial(v) -> str:
    """E2_YRJ: "1" = título de recuperação judicial; "2" ou vazio = não é."""
    return "Sim" if str(v or "").strip() == "1" else "Não"


def _status_from_row(row: Dict) -> str:
    baixa = (row.get("data_baixa") or "").strip()
    if baixa:
        return "Baixado"
    vcto = (row.get("vencimento_real") or "").strip()
    if vcto and vcto < TODAY:
        return "Vencido"
    return "Em aberto"


def _mes_ano(vencimento_real) -> str:
    """'AAAAMMDD' -> 'AAAAMM' (chave ordenável); vazio se a data for inválida."""
    s = str(vencimento_real or "").strip()
    return s[:6] if len(s) == 8 and s.isdigit() else ""


def _fmt_mes(mes_ano) -> str:
    """'AAAAMM' -> 'MM/AAAA', para exibição."""
    s = str(mes_ano or "")
    return f"{s[4:6]}/{s[:4]}" if len(s) == 6 else (s or "—")


# ── aplicação principal ───────────────────────────────────────────────────────

class FinanceiroApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Relatório Financeiro — Grupo LGH (sem o Broker)")
        self.root.geometry("1340x860")
        self.root.minsize(980, 640)
        self.root.configure(bg="#f0f2f5")

        self._df_raw: Optional[pd.DataFrame] = None
        self._df: Optional[pd.DataFrame] = None
        self._loading = False
        self._sort_ascending: Dict[str, bool] = {}

        self._setup_styles()
        self._build_ui()

    # ── estilos ──────────────────────────────────────────────────────────────

    def _setup_styles(self) -> None:
        s = ttk.Style(self.root)
        s.theme_use("clam")

        BG, WHITE, PRIMARY, ACCENT = "#f0f2f5", "#ffffff", "#1a5276", "#2980b9"

        s.configure("TFrame", background=BG)
        s.configure("White.TFrame", background=WHITE)
        s.configure("TLabel", background=BG, font=("Segoe UI", 10))
        s.configure("TLabelframe", background=BG, font=("Segoe UI", 10, "bold"))
        s.configure("TLabelframe.Label", background=BG, font=("Segoe UI", 10, "bold"),
                    foreground=PRIMARY)

        for name, bg, fg, hover in [
            ("Primary",   PRIMARY,   WHITE, ACCENT),
            ("Secondary", "#7f8c8d", WHITE, "#95a5a6"),
            ("Success",   "#27ae60", WHITE, "#2ecc71"),
        ]:
            s.configure(f"{name}.TButton",
                        background=bg, foreground=fg,
                        font=("Segoe UI", 10, "bold"), padding=(12, 5))
            s.map(f"{name}.TButton", background=[("active", hover), ("pressed", bg)])

        s.configure("TEntry", fieldbackground=WHITE, padding=4)
        s.configure("TNotebook", background=BG)
        s.configure("TNotebook.Tab", padding=(12, 6), font=("Segoe UI", 10))

        s.configure("Treeview", background=WHITE, fieldbackground=WHITE,
                    font=("Segoe UI", 10), rowheight=25)
        s.configure("Treeview.Heading",
                    font=("Segoe UI", 10, "bold"),
                    background=PRIMARY, foreground=WHITE)
        s.map("Treeview",
              background=[("selected", ACCENT)],
              foreground=[("selected", WHITE)])

    # ── construção da UI ─────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        self._build_conn_frame()
        self._build_periodo_frame()
        self._build_filter_frame()
        self._build_kpi_frame()
        self._build_toolbar()
        self._build_notebook()
        self._build_statusbar()

    # ── conexão ──────────────────────────────────────────────────────────────

    def _build_conn_frame(self) -> None:
        f = ttk.LabelFrame(self.root, text="Conexão", padding=(10, 6))
        f.pack(fill="x", padx=12, pady=(8, 3))

        ttk.Label(f, text="URL da API:").grid(row=0, column=0, sticky="w", padx=(0, 4))
        self._url_var = tk.StringVar(value=API_BASE_URL)
        ttk.Entry(f, textvariable=self._url_var, width=36).grid(
            row=0, column=1, sticky="ew", padx=(0, 18))

        ttk.Label(f, text="API Key:").grid(row=0, column=2, sticky="w", padx=(0, 4))
        self._key_var = tk.StringVar(value=API_KEY)
        ttk.Entry(f, textvariable=self._key_var, show="•", width=34).grid(
            row=0, column=3, sticky="ew", padx=(0, 12))

        ttk.Button(f, text="Testar conexão", style="Secondary.TButton",
                   command=self._test_connection).grid(row=0, column=4)

        f.columnconfigure(1, weight=1)
        f.columnconfigure(3, weight=1)

    # ── período (define vencimento_de/vencimento_ate enviados à API) ───────────

    def _build_periodo_frame(self) -> None:
        f = ttk.LabelFrame(
            self.root,
            text="Período e status (define vencimento_de/vencimento_ate/status enviados à API — recarrega ao consultar)",
            padding=(10, 6),
        )
        f.pack(fill="x", padx=12, pady=(8, 3))

        # caixa de seleção liga/desliga o envio do parâmetro; o campo de
        # calendário sempre contém uma data válida.
        self._f_vcto_de_on = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            f, text="Vencimento De:", variable=self._f_vcto_de_on,
            command=lambda: self._toggle_date_filter(
                self._de_vcto_de, self._f_vcto_de_on),
        ).grid(row=0, column=0, sticky="w", padx=(0, 3))
        self._f_vcto_de = tk.StringVar()
        self._de_vcto_de = DateEntry(
            f, textvariable=self._f_vcto_de, width=10,
            date_pattern="dd/mm/yyyy",
            background="#1a5276", foreground="white", borderwidth=1,
        )
        self._de_vcto_de.set_date(date.today())
        self._de_vcto_de.grid(row=0, column=1, sticky="w")

        self._f_vcto_ate_on = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            f, text="Vencimento Até:", variable=self._f_vcto_ate_on,
            command=lambda: self._toggle_date_filter(
                self._de_vcto_ate, self._f_vcto_ate_on),
        ).grid(row=0, column=2, sticky="w", padx=(14, 3))
        self._f_vcto_ate = tk.StringVar()
        self._de_vcto_ate = DateEntry(
            f, textvariable=self._f_vcto_ate, width=10,
            date_pattern="dd/mm/yyyy",
            background="#1a5276", foreground="white", borderwidth=1,
        )
        self._de_vcto_ate.set_date(date.today())
        self._de_vcto_ate.grid(row=0, column=3, sticky="w")

        ttk.Label(f, text="Status:").grid(row=0, column=4, sticky="w", padx=(14, 3))
        self._f_status = tk.StringVar(value=STATUS_FILTRO_PADRAO)
        ttk.Combobox(
            f, textvariable=self._f_status, width=14, state="readonly",
            values=[STATUS_FILTRO_PADRAO, *STATUS_API_VALUES.keys()],
        ).grid(row=0, column=5, sticky="w")

        ttk.Label(
            f,
            text="(período já inicia na data de hoje, igual ao padrão da API; desmarque para não limitar)",
            foreground="#7f8c8d",
        ).grid(row=0, column=6, sticky="w", padx=(10, 0))

    # ── filtros (aplicados no cliente, após o carregamento) ────────────────────

    def _build_filter_frame(self) -> None:
        f = ttk.LabelFrame(
            self.root,
            text="Filtros (aplicados localmente sobre os dados carregados)",
            padding=(10, 6),
        )
        f.pack(fill="x", padx=12, pady=3)

        text_filters = [
            ("Filial:",        "_f_filial",   6),
            ("Fornecedor:",    "_f_forn",    18),
            ("Tipo:",          "_f_tipo",     6),
            ("Tipo Operação:", "_f_tipo_op", 18),
        ]
        col = 0
        for label, attr, width in text_filters:
            ttk.Label(f, text=label).grid(
                row=0, column=col, sticky="w", padx=(0 if col == 0 else 14, 3))
            var = tk.StringVar()
            setattr(self, attr, var)
            ttk.Entry(f, textvariable=var, width=width).grid(
                row=0, column=col + 1, sticky="ew")
            col += 2

        ttk.Label(f, text="Categoria:").grid(
            row=0, column=col, sticky="w", padx=(14, 3))
        self._f_categoria = tk.StringVar()
        ttk.Combobox(
            f, textvariable=self._f_categoria, width=20, state="readonly",
            values=["", *categorias.CATEGORIAS_CANONICAS],
        ).grid(row=0, column=col + 1, sticky="ew")
        col += 2

        # botões
        btn = ttk.Frame(f)
        btn.grid(row=0, column=col, sticky="e", padx=(20, 0))
        ttk.Button(btn, text=" Carregar da API ", style="Primary.TButton",
                   command=self.consultar).pack(side="left", padx=(0, 6))
        ttk.Button(btn, text="Aplicar filtros", style="Secondary.TButton",
                   command=self._on_filtros_alterados).pack(side="left", padx=(0, 6))
        ttk.Button(btn, text="Limpar", style="Secondary.TButton",
                   command=self.limpar).pack(side="left")

        f.columnconfigure(col, weight=1)

    @staticmethod
    def _toggle_date_filter(entry: DateEntry, enabled: tk.BooleanVar) -> None:
        entry.configure(state="normal" if enabled.get() else "disabled")

    # ── KPI cards ─────────────────────────────────────────────────────────────

    def _build_kpi_frame(self) -> None:
        outer = tk.Frame(self.root, bg="#f0f2f5")
        outer.pack(fill="x", padx=12, pady=3)

        self._kpi_vars: Dict[str, tk.StringVar] = {}
        cards_cfg = [
            ("total_titulos", "Total de Títulos",  "#2c3e50"),
            ("valor_total",   "Valor Total",        "#1a5276"),
            ("saldo_total",   "Saldo a Pagar",      "#154360"),
            ("em_aberto",     "Em Aberto",          "#1e8449"),
            ("vencidos",      "Vencidos",           "#922b21"),
            ("baixados",      "Baixados",           "#626567"),
        ]

        for i, (key, label, fg) in enumerate(cards_cfg):
            card = tk.Frame(outer, bg="#ffffff",
                            highlightbackground="#d5d8dc", highlightthickness=1)
            card.grid(row=0, column=i, sticky="nsew", padx=4, pady=2, ipady=10)

            tk.Label(card, text=label, bg="#ffffff",
                     font=("Segoe UI", 9), fg="#7f8c8d").pack(pady=(4, 0))

            var = tk.StringVar(value="—")
            self._kpi_vars[key] = var
            tk.Label(card, textvariable=var, bg="#ffffff",
                     font=("Segoe UI", 14, "bold"), fg=fg).pack()

        for i in range(len(cards_cfg)):
            outer.columnconfigure(i, weight=1)

    # ── toolbar acima do notebook ─────────────────────────────────────────────

    def _build_toolbar(self) -> None:
        bar = ttk.Frame(self.root)
        bar.pack(fill="x", padx=12, pady=(4, 0))
        ttk.Button(bar, text="  Exportar Excel  ", style="Success.TButton",
                   command=self.exportar_excel).pack(side="right")

    # ── notebook ──────────────────────────────────────────────────────────────

    def _build_notebook(self) -> None:
        self._nb = ttk.Notebook(self.root)
        self._nb.pack(fill="both", expand=True, padx=12, pady=(2, 4))

        tab_graficos = ttk.Frame(self._nb)
        self._nb.add(tab_graficos, text="  Gráficos  ")
        self._build_charts_tab(tab_graficos)

        tab_dados = ttk.Frame(self._nb)
        self._nb.add(tab_dados, text="  Dados  ")
        self._build_data_tab(tab_dados)

        tab_categoria = ttk.Frame(self._nb)
        self._nb.add(tab_categoria, text="  Por Categoria  ")
        self._build_categoria_tab(tab_categoria)

        tab_evolucao = ttk.Frame(self._nb)
        self._nb.add(tab_evolucao, text="  Evolução Mensal  ")
        self._build_evolucao_tab(tab_evolucao)

        tab_filial = ttk.Frame(self._nb)
        self._nb.add(tab_filial, text="  Por Filial  ")
        self._build_filial_tab(tab_filial)

    def _build_charts_tab(self, parent: ttk.Frame) -> None:
        self._fig = Figure(figsize=(13, 6.4), dpi=100, facecolor="#f0f2f5")
        self._axes = self._fig.subplots(2, 2)
        self._fig.subplots_adjust(left=0.1, right=0.97, top=0.9,
                                  bottom=0.14, hspace=0.62, wspace=0.42)
        self._canvas = FigureCanvasTkAgg(self._fig, master=parent)
        self._canvas.get_tk_widget().pack(fill="both", expand=True, padx=4, pady=4)
        self._draw_empty_charts()

    def _build_data_tab(self, parent: ttk.Frame) -> None:
        cols = [c[0] for c in TREEVIEW_COLS]
        self._tree = ttk.Treeview(parent, columns=cols,
                                  show="headings", selectmode="browse")

        for col_id, heading, width, anchor in TREEVIEW_COLS:
            self._tree.heading(
                col_id, text=heading,
                command=lambda c=col_id: self._sort_tree(c),
            )
            self._tree.column(col_id, width=width, anchor=anchor,
                              stretch=(col_id in ("historico", "descricao_operacao")))

        self._tree.tag_configure("Em aberto", background="#eafaf1")
        self._tree.tag_configure("Vencido",   background="#fdecea")
        self._tree.tag_configure("Baixado",   background="#f2f3f4")

        vsb = ttk.Scrollbar(parent, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        parent.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)

    # ── abas de resumo (mesma visão de CONTAS_A_PAGAR_ANALISE.xlsx) ────────────

    @staticmethod
    def _build_summary_table(
        parent: tk.Widget,
        cols: List[Tuple[str, str, int, str]],
        height: int = 8,
    ) -> ttk.Treeview:
        frame = ttk.Frame(parent)
        frame.pack(fill="x", padx=4, pady=(4, 2))

        col_ids = [c[0] for c in cols]
        tree = ttk.Treeview(frame, columns=col_ids, show="headings",
                            selectmode="none", height=height)
        for col_id, heading, width, anchor in cols:
            tree.heading(col_id, text=heading)
            tree.column(col_id, width=width, anchor=anchor, stretch=True)
        tree.tag_configure("total", background="#eaf2f8", font=("Segoe UI", 9, "bold"))

        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="x", expand=True)
        vsb.pack(side="right", fill="y")
        return tree

    @staticmethod
    def _build_summary_chart(parent: tk.Widget, figsize=(13, 4.2)):
        frame = ttk.Frame(parent)
        frame.pack(fill="both", expand=True, padx=4, pady=(2, 4))
        fig = Figure(figsize=figsize, dpi=100, facecolor="#f0f2f5")
        ax = fig.add_subplot(111)
        fig.subplots_adjust(left=0.24, right=0.96, top=0.88, bottom=0.14)
        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.get_tk_widget().pack(fill="both", expand=True)
        return fig, ax, canvas

    @staticmethod
    def _draw_empty_ax(ax, canvas: FigureCanvasTkAgg) -> None:
        ax.clear()
        ax.text(0.5, 0.5, "Sem dados — clique em \"Carregar da API\"",
                ha="center", va="center", color="#aab7b8", fontsize=10.5)
        ax.set_facecolor("#f8f9fa")
        ax.axis("off")
        canvas.draw()

    def _build_categoria_tab(self, parent: ttk.Frame) -> None:
        cols = [
            ("categoria",   "Categoria",         180, "w"),
            ("qtd",         "Qtd. Títulos",       90, "c"),
            ("valor_total", "Valor Total (R$)",  140, "e"),
            ("pct_total",   "% do Total",         90, "c"),
            ("valor_medio", "Valor Médio (R$)",  140, "e"),
        ]
        self._tree_categoria = self._build_summary_table(parent, cols, height=10)
        self._fig_categoria, self._ax_categoria, self._canvas_categoria = \
            self._build_summary_chart(parent)
        self._draw_empty_ax(self._ax_categoria, self._canvas_categoria)

    def _build_filial_tab(self, parent: ttk.Frame) -> None:
        cols = [
            ("filial",      "Filial",            160, "w"),
            ("qtd",         "Qtd. Títulos",        90, "c"),
            ("valor_total", "Valor Total (R$)",   150, "e"),
            ("pct_total",   "% do Total",          90, "c"),
        ]
        self._tree_filial = self._build_summary_table(parent, cols, height=8)
        self._fig_filial, self._ax_filial, self._canvas_filial = \
            self._build_summary_chart(parent)
        self._draw_empty_ax(self._ax_filial, self._canvas_filial)

    def _build_evolucao_tab(self, parent: ttk.Frame) -> None:
        # a tabela tem colunas dinâmicas (um mês por coluna) — é reconstruída
        # a cada atualização de dados, então aqui só reservamos o container.
        self._evolucao_table_frame = ttk.Frame(parent)
        self._evolucao_table_frame.pack(fill="x")
        self._tree_evolucao: Optional[ttk.Treeview] = None

        chart_frame = ttk.Frame(parent)
        chart_frame.pack(fill="both", expand=True, padx=4, pady=(2, 4))
        self._fig_evolucao = Figure(figsize=(13, 4.1), dpi=100, facecolor="#f0f2f5")
        self._ax_evolucao = self._fig_evolucao.add_subplot(111)
        self._fig_evolucao.subplots_adjust(left=0.08, right=0.98, top=0.87, bottom=0.24)
        self._canvas_evolucao = FigureCanvasTkAgg(self._fig_evolucao, master=chart_frame)
        self._canvas_evolucao.get_tk_widget().pack(fill="both", expand=True)
        self._draw_empty_ax(self._ax_evolucao, self._canvas_evolucao)

    # ── barra de status ───────────────────────────────────────────────────────

    def _build_statusbar(self) -> None:
        bar = tk.Frame(self.root, bg="#dde1e7", height=24)
        bar.pack(fill="x", side="bottom")

        self._status_var = tk.StringVar(value="Pronto.")
        tk.Label(bar, textvariable=self._status_var,
                 bg="#dde1e7", font=("Segoe UI", 9), anchor="w").pack(
            side="left", padx=8, pady=2)

        self._pct_var = tk.StringVar()
        tk.Label(bar, textvariable=self._pct_var,
                 bg="#dde1e7", font=("Segoe UI", 9)).pack(side="right", padx=4)

        self._progress = ttk.Progressbar(bar, mode="determinate", length=200)
        self._progress.pack(side="right", padx=8, pady=3)

    # ── conexão ───────────────────────────────────────────────────────────────

    def _make_client(self) -> APIClient:
        return APIClient(
            base_url=self._url_var.get().strip(),
            api_key=self._key_var.get().strip(),
            api_key_name=API_KEY_NAME,
        )

    def _test_connection(self) -> None:
        try:
            self._make_client().health()
            messagebox.showinfo("Conexão OK",
                                "A API respondeu com sucesso!")
        except Exception as exc:
            messagebox.showerror("Falha na conexão", str(exc))

    # ── consulta (thread) ─────────────────────────────────────────────────────

    def consultar(self) -> None:
        if self._loading:
            return
        self._loading = True
        self._clear_data()

        # DateEntry sempre contém uma data válida; converte DD/MM/AAAA -> AAAAMMDD.
        vencimento_de = (
            self._de_vcto_de.get_date().strftime("%Y%m%d")
            if self._f_vcto_de_on.get() else None
        )
        vencimento_ate = (
            self._de_vcto_ate.get_date().strftime("%Y%m%d")
            if self._f_vcto_ate_on.get() else None
        )
        status_label = self._f_status.get()
        status = STATUS_API_VALUES.get(status_label)

        partes = []
        if vencimento_de:
            partes.append(f"de {self._f_vcto_de.get()}")
        if vencimento_ate:
            partes.append(f"até {self._f_vcto_ate.get()}")
        periodo = " ".join(partes) if partes else "sem limite de período"
        if status:
            periodo += f", status={status_label}"
        self._set_status(f"Consultando /financeiro na API… [{periodo}]", indeterminate=True)
        threading.Thread(
            target=self._fetch_thread, args=(vencimento_de, vencimento_ate, status), daemon=True
        ).start()

    def _fetch_thread(
        self, vencimento_de: Optional[str], vencimento_ate: Optional[str],
        status: Optional[str],
    ) -> None:
        try:
            client = self._make_client()
            items, total = client.get_all_financeiro(
                vencimento_de=vencimento_de, vencimento_ate=vencimento_ate, status=status,
                progress_callback=self._on_progress)
            self.root.after(0, self._on_data_ready, items, total)
        except Exception as exc:
            self.root.after(0, self._on_error, str(exc))

    def _on_progress(self, loaded: int, total: int) -> None:
        pct = int(100 * loaded / total) if total else 0
        self.root.after(0, self._update_progress_bar, loaded, total, pct)

    def _update_progress_bar(self, loaded: int, total: int, pct: int) -> None:
        self._progress.stop()
        self._progress.configure(mode="determinate", value=pct)
        self._pct_var.set(f"{pct}%  ({loaded}/{total})")
        self._status_var.set(f"Carregando… {loaded} de {total} registros")

    def _on_data_ready(self, items: List[Dict], total_api: int) -> None:
        self._loading = False

        for row in items:
            row["status"] = _status_from_row(row)

        df = pd.DataFrame(items) if items else pd.DataFrame()
        if not df.empty:
            df["valor"] = pd.to_numeric(df["valor"], errors="coerce").fillna(0.0)
            df["saldo"] = pd.to_numeric(df["saldo"], errors="coerce").fillna(0.0)
            df["mes_ano"] = df["vencimento_real"].apply(_mes_ano)

            regras = categorias.load_rules()
            df["categoria"] = df.apply(
                lambda r: categorias.classify(
                    r.get("nome_fornecedor", ""), r.get("historico", ""), regras),
                axis=1,
            )

        self._df_raw = df
        self._apply_filters_and_refresh()

        n = len(self._df) if self._df is not None else 0
        self._set_status(
            f"{n} título(s) exibido(s)"
            + (f"  [total carregado da API: {total_api}]" if n != total_api else "")
        )
        self._progress["value"] = 100
        self._pct_var.set("100%")

    def _on_error(self, msg: str) -> None:
        self._loading = False
        self._progress.stop()
        self._progress.configure(mode="determinate", value=0)
        self._pct_var.set("")
        self._set_status(f"Erro: {msg}")
        messagebox.showerror("Erro ao consultar a API", msg)

    # ── filtros (client-side) ────────────────────────────────────────────────

    def _on_filtros_alterados(self) -> None:
        if self._df_raw is None:
            messagebox.showinfo(
                "Filtros", "Carregue os dados da API antes de aplicar filtros.")
            return
        self._apply_filters_and_refresh()
        n = len(self._df) if self._df is not None else 0
        self._set_status(f"{n} título(s) exibido(s) após filtros")

    def _apply_filters_and_refresh(self) -> None:
        df = self._df_raw
        if df is None or df.empty:
            self._df = df
        else:
            self._df = self._apply_filters(df)
        self._update_kpis()
        self._update_charts()
        self._update_table()
        self._update_categoria_view()
        self._update_evolucao_view()
        self._update_filial_view()

    def _apply_filters(self, df: pd.DataFrame) -> pd.DataFrame:
        filial = self._f_filial.get().strip()
        if filial:
            df = df[df["filial"] == filial]

        fornecedor = self._f_forn.get().strip().lower()
        if fornecedor:
            df = df[df["nome_fornecedor"].fillna("").str.lower().str.contains(fornecedor)]

        tipo = self._f_tipo.get().strip().upper()
        if tipo:
            df = df[df["tipo"].fillna("").str.upper() == tipo]

        tipo_op = self._f_tipo_op.get().strip().lower()
        if tipo_op:
            df = df[df["descricao_operacao"].fillna("").str.lower().str.contains(tipo_op)]

        categoria = self._f_categoria.get().strip()
        if categoria:
            df = df[df["categoria"] == categoria]

        # período de vencimento já é filtrado no servidor (vencimento_de/vencimento_ate
        # enviados em consultar()) — não é reaplicado aqui.

        return df

    # ── KPIs ──────────────────────────────────────────────────────────────────

    def _update_kpis(self) -> None:
        df = self._df
        if df is None or df.empty:
            for v in self._kpi_vars.values():
                v.set("—")
            return

        ab = df[df["status"] == "Em aberto"]
        ve = df[df["status"] == "Vencido"]
        bx = df[df["status"] == "Baixado"]

        self._kpi_vars["total_titulos"].set(f'{len(df):,}'.replace(",", "."))
        self._kpi_vars["valor_total"].set(_brl(df["valor"].sum()))
        self._kpi_vars["saldo_total"].set(_brl(df["saldo"].sum()))
        self._kpi_vars["em_aberto"].set(
            f'{len(ab):,}'.replace(",", ".") + "\n" + _brl(ab["saldo"].sum()))
        self._kpi_vars["vencidos"].set(
            f'{len(ve):,}'.replace(",", ".") + "\n" + _brl(ve["saldo"].sum()))
        self._kpi_vars["baixados"].set(
            f'{len(bx):,}'.replace(",", ".") + "\n" + _brl(bx["valor"].sum()))

    # ── gráficos ──────────────────────────────────────────────────────────────

    def _draw_empty_charts(self) -> None:
        for ax in self._axes.flat:
            ax.clear()
            ax.text(0.5, 0.5, "Sem dados — clique em \"Carregar da API\"",
                    ha="center", va="center", color="#aab7b8", fontsize=10.5)
            ax.set_facecolor("#f8f9fa")
            ax.axis("off")
        self._canvas.draw()

    def _update_charts(self) -> None:
        for ax in self._axes.flat:
            ax.clear()

        df = self._df
        if df is None or df.empty:
            self._draw_empty_charts()
            return

        ax_status, ax_forn = self._axes[0]
        ax_mes,    ax_tipo = self._axes[1]

        brl_fmt = mticker.FuncFormatter(
            lambda x, _: f"R${x/1000:.0f}k" if abs(x) >= 1000 else f"R${x:.0f}"
        )

        # ── Gráfico 1: pizza — distribuição por status ────────────────────
        status_counts = df["status"].value_counts()
        pie_colors = [STATUS_COLORS.get(s, "#bdc3c7") for s in status_counts.index]
        wedges, texts, autotexts = ax_status.pie(
            status_counts,
            labels=status_counts.index,
            autopct="%1.1f%%",
            colors=pie_colors,
            startangle=90,
            textprops={"fontsize": 9.5},
            wedgeprops={"edgecolor": "white", "linewidth": 1.5},
        )
        for at in autotexts:
            at.set_fontsize(9)
        ax_status.set_title("Distribuição por Status",
                             fontsize=13, fontweight="bold", pad=12)

        # ── Gráfico 2: barras horizontais — Top 10 Fornecedores por Saldo ─
        top10 = (
            df.assign(nome_fornecedor=df["nome_fornecedor"].fillna("(sem nome)"))
            .groupby("nome_fornecedor")["saldo"]
            .sum()
            .nlargest(10)
            .sort_values(ascending=True)
        )
        bar_colors = CHART_COLORS[:len(top10)]
        hbars = ax_forn.barh(
            top10.index, top10.values,
            color=bar_colors[::-1], height=0.65,
            edgecolor="none",
        )
        ax_forn.set_title("Top 10 Fornecedores — Saldo",
                           fontsize=13, fontweight="bold", pad=12)
        ax_forn.set_xlabel("Saldo (R$)", fontsize=10.5)
        ax_forn.tick_params(axis="y", labelsize=9.5)
        ax_forn.tick_params(axis="x", labelsize=9)
        ax_forn.xaxis.set_major_formatter(brl_fmt)
        ax_forn.set_facecolor("#fafafa")
        ax_forn.margins(x=0.2)
        for bar in hbars:
            w = bar.get_width()
            ax_forn.text(
                w * 1.01, bar.get_y() + bar.get_height() / 2,
                _brl(w), va="center", fontsize=9, color="#555",
            )

        # ── Gráfico 3: barras — Saldo por Mês de Vencimento ───────────────
        df_open = df[df["status"] != "Baixado"].copy()
        df_open["mes"] = df_open["vencimento_real"].str[:6]
        by_month = (
            df_open.groupby("mes")["saldo"]
            .sum()
            .sort_index()
            .tail(14)
        )
        labels_mes = [f"{m[4:6]}/{m[:4]}" for m in by_month.index]
        bar_clrs_mes = [
            "#e74c3c" if m < TODAY[:6] else "#2980b9"
            for m in by_month.index
        ]
        ax_mes.bar(range(len(by_month)), by_month.values,
                   color=bar_clrs_mes, width=0.72, edgecolor="none")
        ax_mes.set_xticks(range(len(by_month)))
        ax_mes.set_xticklabels(labels_mes, rotation=45, ha="right", fontsize=9.5)
        ax_mes.set_title("Saldo por Mês de Vencimento (em aberto + vencidos)",
                         fontsize=13, fontweight="bold", pad=12)
        ax_mes.set_ylabel("Saldo (R$)", fontsize=10.5)
        ax_mes.tick_params(axis="y", labelsize=9.5)
        ax_mes.yaxis.set_major_formatter(brl_fmt)
        ax_mes.set_facecolor("#fafafa")
        ax_mes.legend(handles=[
            Patch(color="#e74c3c", label="Vencido"),
            Patch(color="#2980b9", label="A vencer"),
        ], fontsize=9.5, loc="upper left", framealpha=0.8)

        # ── Gráfico 4: barras horizontais — Saldo por Tipo de Operação ────
        by_tipo = (
            df.assign(descricao_operacao=df["descricao_operacao"].fillna("(não informado)"))
            .groupby("descricao_operacao")["saldo"]
            .sum()
            .nlargest(10)
            .sort_values(ascending=True)
        )
        bar_colors_tipo = CHART_COLORS[:len(by_tipo)]
        hbars_tipo = ax_tipo.barh(
            by_tipo.index, by_tipo.values,
            color=bar_colors_tipo[::-1], height=0.65,
            edgecolor="none",
        )
        ax_tipo.set_title("Saldo por Tipo de Operação",
                          fontsize=13, fontweight="bold", pad=12)
        ax_tipo.set_xlabel("Saldo (R$)", fontsize=10.5)
        ax_tipo.tick_params(axis="y", labelsize=9.5)
        ax_tipo.tick_params(axis="x", labelsize=9)
        ax_tipo.xaxis.set_major_formatter(brl_fmt)
        ax_tipo.set_facecolor("#fafafa")
        ax_tipo.margins(x=0.2)
        for bar in hbars_tipo:
            w = bar.get_width()
            ax_tipo.text(
                w * 1.01, bar.get_y() + bar.get_height() / 2,
                _brl(w), va="center", fontsize=9, color="#555",
            )

        self._canvas.draw()

    # ── resumos (mesma visão de CONTAS_A_PAGAR_ANALISE.xlsx) ────────────────────

    _brl_axis_fmt = mticker.FuncFormatter(
        lambda x, _: f"R${x/1000:.0f}k" if abs(x) >= 1000 else f"R${x:.0f}")

    def _update_categoria_view(self) -> None:
        tree, ax, canvas = self._tree_categoria, self._ax_categoria, self._canvas_categoria
        for row in tree.get_children():
            tree.delete(row)

        df = self._df
        if df is None or df.empty:
            self._draw_empty_ax(ax, canvas)
            return

        resumo = (
            df.groupby("categoria", as_index=False)
            .agg(qtd=("valor", "count"), valor_total=("valor", "sum"))
            .sort_values("valor_total", ascending=False)
        )
        total_geral = resumo["valor_total"].sum()

        for _, r in resumo.iterrows():
            pct = (r["valor_total"] / total_geral) if total_geral else 0
            valor_medio = (r["valor_total"] / r["qtd"]) if r["qtd"] else 0
            tree.insert("", "end", values=(
                r["categoria"], int(r["qtd"]), _brl(r["valor_total"]),
                f"{pct * 100:.1f}%", _brl(valor_medio),
            ))
        tree.insert("", "end", tags=("total",), values=(
            "TOTAL GERAL", int(resumo["qtd"].sum()), _brl(total_geral), "100,0%", "",
        ))

        ax.clear()
        plot_data = resumo.sort_values("valor_total", ascending=True)
        colors = [CATEGORIA_COLORS.get(c, "#7f8c8d") for c in plot_data["categoria"]]
        bars = ax.barh(plot_data["categoria"], plot_data["valor_total"],
                        color=colors, height=0.6, edgecolor="none")
        ax.set_title("Valor Total por Categoria", fontsize=13, fontweight="bold", pad=12)
        ax.set_xlabel("Valor (R$)", fontsize=10.5)
        ax.tick_params(axis="both", labelsize=9.5)
        ax.xaxis.set_major_formatter(self._brl_axis_fmt)
        ax.set_facecolor("#fafafa")
        ax.margins(x=0.2)
        for bar in bars:
            w = bar.get_width()
            ax.text(w * 1.01, bar.get_y() + bar.get_height() / 2,
                    _brl(w), va="center", fontsize=9, color="#555")
        canvas.draw()

    def _update_filial_view(self) -> None:
        tree, ax, canvas = self._tree_filial, self._ax_filial, self._canvas_filial
        for row in tree.get_children():
            tree.delete(row)

        df = self._df
        if df is None or df.empty:
            self._draw_empty_ax(ax, canvas)
            return

        resumo = (
            df.groupby("filial", as_index=False)
            .agg(qtd=("valor", "count"), valor_total=("valor", "sum"))
            .sort_values("valor_total", ascending=False)
        )
        total_geral = resumo["valor_total"].sum()

        for _, r in resumo.iterrows():
            pct = (r["valor_total"] / total_geral) if total_geral else 0
            tree.insert("", "end", values=(
                r["filial"], int(r["qtd"]), _brl(r["valor_total"]), f"{pct * 100:.1f}%",
            ))
        tree.insert("", "end", tags=("total",), values=(
            "TOTAL GERAL", int(resumo["qtd"].sum()), _brl(total_geral), "100,0%",
        ))

        ax.clear()
        pie_colors = CHART_COLORS[:len(resumo)]
        wedges, texts, autotexts = ax.pie(
            resumo["valor_total"], labels=resumo["filial"], autopct="%1.1f%%",
            colors=pie_colors, startangle=90, textprops={"fontsize": 9.5},
            wedgeprops={"edgecolor": "white", "linewidth": 1.5},
        )
        for at in autotexts:
            at.set_fontsize(9)
        ax.set_title("Valor Total por Filial", fontsize=13, fontweight="bold", pad=12)
        canvas.draw()

    def _update_evolucao_view(self) -> None:
        for child in self._evolucao_table_frame.winfo_children():
            child.destroy()
        self._tree_evolucao = None

        df = self._df
        if df is None or df.empty:
            self._draw_empty_ax(self._ax_evolucao, self._canvas_evolucao)
            return

        meses = sorted(m for m in df["mes_ano"].unique() if m)
        pivot = (
            df.groupby(["categoria", "mes_ano"])["valor"].sum()
            .unstack(fill_value=0.0)
            .reindex(columns=meses, fill_value=0.0)
        )
        pivot["Total"] = pivot.sum(axis=1)
        pivot = pivot.sort_values("Total", ascending=False)

        cols = [("categoria", "Categoria", 180, "w")]
        cols += [(m, _fmt_mes(m), 95, "e") for m in meses]
        cols.append(("total", "Total", 110, "e"))
        tree = self._build_summary_table(
            self._evolucao_table_frame, cols, height=min(11, len(pivot) + 2))
        self._tree_evolucao = tree

        for categoria, row in pivot.iterrows():
            values = [categoria] + [_brl(row[m]) for m in meses] + [_brl(row["Total"])]
            tree.insert("", "end", values=values)
        total_row = (
            ["TOTAL MENSAL"] + [_brl(pivot[m].sum()) for m in meses]
            + [_brl(pivot["Total"].sum())]
        )
        tree.insert("", "end", tags=("total",), values=total_row)

        ax = self._ax_evolucao
        ax.clear()
        x = range(len(meses))
        bottom = [0.0] * len(meses)
        for categoria in pivot.index:
            vals = [pivot.loc[categoria, m] for m in meses]
            ax.bar(x, vals, bottom=bottom, width=0.65, edgecolor="none",
                   color=CATEGORIA_COLORS.get(categoria, "#7f8c8d"), label=categoria)
            bottom = [b + v for b, v in zip(bottom, vals)]
        ax.set_xticks(list(x))
        ax.set_xticklabels([_fmt_mes(m) for m in meses], fontsize=9.5,
                           rotation=45, ha="right")
        ax.set_title("Evolução dos Custos por Categoria (mês de vencimento)",
                     fontsize=13, fontweight="bold", pad=12)
        ax.set_ylabel("Valor (R$)", fontsize=10.5)
        ax.tick_params(axis="y", labelsize=9.5)
        ax.yaxis.set_major_formatter(self._brl_axis_fmt)
        ax.set_facecolor("#fafafa")
        ax.legend(fontsize=9, loc="upper left", ncol=3, framealpha=0.85)
        self._canvas_evolucao.draw()

    # ── tabela ────────────────────────────────────────────────────────────────

    def _update_table(self) -> None:
        for row in self._tree.get_children():
            self._tree.delete(row)

        df = self._df
        if df is None or df.empty:
            return

        for _, row in df.iterrows():
            status = row.get("status", "")
            self._tree.insert("", "end", tags=(status,), values=(
                row.get("filial", ""),
                row.get("numero", ""),
                row.get("parcela", ""),
                row.get("tipo", ""),
                row.get("codigo_operacao", ""),
                row.get("descricao_operacao", ""),
                row.get("nome_fornecedor", ""),
                row.get("categoria", ""),
                _fmt_date(row.get("emissao")),
                _fmt_date(row.get("vencimento_real")),
                _brl(row.get("valor", 0)),
                _brl(row.get("saldo", 0)),
                row.get("historico", ""),
                _fmt_rec_judicial(row.get("recuperacao_judicial")),
                status,
            ))

    def _sort_tree(self, col: str) -> None:
        if self._df is None or self._df.empty:
            return
        asc = not self._sort_ascending.get(col, False)
        self._sort_ascending[col] = asc

        if col in ("valor", "saldo"):
            self._df = self._df.sort_values(col, ascending=asc, na_position="last")
        else:
            self._df = self._df.sort_values(
                col, ascending=asc, na_position="last",
                key=lambda s: s.astype(str).str.lower(),
            )
        self._update_table()

    # ── exportar Excel ────────────────────────────────────────────────────────

    def exportar_excel(self) -> None:
        df = self._df
        if df is None or df.empty:
            messagebox.showwarning(
                "Exportar Excel",
                "Nenhum dado carregado. Faça uma consulta primeiro.",
            )
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("Todos os arquivos", "*.*")],
            title="Salvar planilha do Relatório Financeiro",
            initialfile=f"financeiro_{date.today().strftime('%Y%m%d')}.xlsx",
        )
        if not path:
            return

        try:
            self._set_status("Gerando planilha Excel…")
            self.root.update()
            _write_excel(df, path)
            self._set_status(f"Planilha salva em: {path}")
            messagebox.showinfo("Exportar Excel",
                                f"Arquivo salvo com sucesso:\n{path}")
        except Exception as exc:
            self._set_status("Erro ao exportar.")
            messagebox.showerror("Erro ao exportar", str(exc))

    # ── utilitários ───────────────────────────────────────────────────────────

    def limpar(self) -> None:
        # O período e o status (Vencimento De/Até/Status) ficam fora deste
        # botão: são enviados à API e só são reaplicados ao clicar em
        # "Carregar da API" novamente.
        for attr in ("_f_filial", "_f_forn", "_f_tipo", "_f_tipo_op", "_f_categoria"):
            getattr(self, attr).set("")
        if self._df_raw is not None:
            self._apply_filters_and_refresh()
        self._set_status("Filtros limpos.")

    def _clear_data(self) -> None:
        self._df_raw = None
        self._df = None
        for v in self._kpi_vars.values():
            v.set("—")
        for row in self._tree.get_children():
            self._tree.delete(row)
        self._draw_empty_charts()
        self._update_categoria_view()
        self._update_evolucao_view()
        self._update_filial_view()
        self._progress["value"] = 0
        self._pct_var.set("")

    def _set_status(self, msg: str, *, indeterminate: bool = False) -> None:
        self._status_var.set(msg)
        if indeterminate:
            self._progress.configure(mode="indeterminate")
            self._progress.start(12)
            self._pct_var.set("")
        else:
            self._progress.stop()
            self._progress.configure(mode="determinate")


# ── exportação Excel (função separada) ───────────────────────────────────────

def _write_excel(df: pd.DataFrame, path: str) -> None:
    import openpyxl
    from openpyxl.styles import (Alignment, Border, Font,
                                 PatternFill, Side)
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── estilos compartilhados ────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1A5276")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    ALT_FILL  = PatternFill("solid", fgColor="EBF5FB")
    VENC_FILL = PatternFill("solid", fgColor="FDEDEC")
    BAIX_FILL = PatternFill("solid", fgColor="F2F3F4")
    THIN      = Side(style="thin", color="BDC3C7")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER    = Alignment(horizontal="center", vertical="center")
    RIGHT     = Alignment(horizontal="right",  vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")
    BRL       = '#,##0.00'
    PCT       = '0.0%'
    DATE_COLS = {"emissao", "vencimento_real"}
    CURR_COLS = {"valor", "saldo"}

    def _hdr(ws, row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = BORDER
        return c

    # ── aba 1: Financeiro ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Financeiro"
    ws1.freeze_panes = "A2"

    cols_aba1 = [
        ("Filial",         "filial",              7),
        ("Número",         "numero",             13),
        ("Parcela",        "parcela",             9),
        ("Tipo",           "tipo",                7),
        ("Cód. Operação",  "codigo_operacao",    12),
        ("Tipo Operação",  "descricao_operacao", 26),
        ("Fornecedor",     "nome_fornecedor",    30),
        ("Categoria",      "categoria",          22),
        ("Emissão",        "emissao",            13),
        ("Vencimento",     "vencimento_real",    13),
        ("Mês/Ano",        "mes_ano",            10),
        ("Valor (R$)",     "valor",              16),
        ("Saldo (R$)",     "saldo",              16),
        ("Histórico",      "historico",          42),
        ("Rec. Judicial",  "recuperacao_judicial", 14),
        ("Status",         "status",             12),
    ]

    for ci, (heading, _, w) in enumerate(cols_aba1, 1):
        _hdr(ws1, 1, ci, heading)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[1].height = 18
    ws1.auto_filter.ref = (
        f"A1:{get_column_letter(len(cols_aba1))}1"
    )

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        status = row.get("status", "")
        row_fill = (VENC_FILL if status == "Vencido"
                    else BAIX_FILL if status == "Baixado"
                    else (ALT_FILL if ri % 2 == 0 else None))
        for ci, (_, field, _) in enumerate(cols_aba1, 1):
            val = row.get(field)
            if field in DATE_COLS:
                val = _fmt_date(val)
            elif field == "mes_ano":
                val = _fmt_mes(val)
            elif field == "recuperacao_judicial":
                val = _fmt_rec_judicial(val)
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if field in CURR_COLS:
                cell.number_format = BRL
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT
            if row_fill:
                cell.fill = row_fill

    # ── aba 2: Resumo ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 22

    ws2.cell(1, 1, "Resumo — Relatório Financeiro").font = Font(
        bold=True, color="1A5276", size=14, name="Calibri")
    ws2.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(
        italic=True, size=9, color="7F8C8D", name="Calibri")

    aberto = df[df["status"] == "Em aberto"]
    vencido = df[df["status"] == "Vencido"]
    baixado = df[df["status"] == "Baixado"]

    items_resumo = [
        ("Total de Títulos",         len(df),                       None),
        ("Valor Total (R$)",         float(df["valor"].sum()),      BRL),
        ("Saldo Total (R$)",         float(df["saldo"].sum()),      BRL),
        ("",                          "",                            None),
        ("Em Aberto — Qtd",          len(aberto),                   None),
        ("Em Aberto — Saldo (R$)",   float(aberto["saldo"].sum()),  BRL),
        ("",                          "",                            None),
        ("Vencidos — Qtd",           len(vencido),                  None),
        ("Vencidos — Saldo (R$)",    float(vencido["saldo"].sum()), BRL),
        ("",                          "",                            None),
        ("Baixados — Qtd",           len(baixado),                  None),
        ("Baixados — Valor (R$)",    float(baixado["valor"].sum()), BRL),
    ]
    for i, (lbl, val, fmt) in enumerate(items_resumo, 4):
        c1 = ws2.cell(i, 1, lbl)
        c1.font = Font(bold=bool(lbl), name="Calibri", size=10)
        c2 = ws2.cell(i, 2, val if val != "" else None)
        c2.alignment = RIGHT
        if fmt:
            c2.number_format = fmt

    # ── aba 3: Por Fornecedor ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Por Fornecedor")
    by_forn = (
        df.assign(nome_fornecedor=df["nome_fornecedor"].fillna("(sem nome)"))
        .groupby("nome_fornecedor", as_index=False)
        .agg(qtd=("numero", "count"),
             valor_total=("valor", "sum"),
             saldo_total=("saldo", "sum"))
        .sort_values("saldo_total", ascending=False)
    )
    hdrs3 = ["Fornecedor", "Qtd Títulos", "Valor Total (R$)", "Saldo Total (R$)"]
    widths3 = [30, 14, 20, 20]
    for ci, (h, w) in enumerate(zip(hdrs3, widths3), 1):
        _hdr(ws3, 1, ci, h)
        ws3.column_dimensions[get_column_letter(ci)].width = w
    for ri, row in enumerate(by_forn.itertuples(index=False), 2):
        ws3.cell(ri, 1, row.nome_fornecedor)
        ws3.cell(ri, 2, int(row.qtd))
        c3 = ws3.cell(ri, 3, float(row.valor_total))
        c3.number_format = BRL
        c4 = ws3.cell(ri, 4, float(row.saldo_total))
        c4.number_format = BRL
        if ri % 2 == 0:
            for ci in range(1, 5):
                ws3.cell(ri, ci).fill = ALT_FILL

    # ── aba 4: Por Tipo de Operação ─────────────────────────────────────────
    ws4 = wb.create_sheet("Por Tipo de Operação")
    by_tipo = (
        df.assign(descricao_operacao=df["descricao_operacao"].fillna("(não informado)"))
        .groupby("descricao_operacao", as_index=False)
        .agg(qtd=("numero", "count"),
             valor_total=("valor", "sum"),
             saldo_total=("saldo", "sum"))
        .sort_values("saldo_total", ascending=False)
    )
    hdrs4 = ["Tipo Operação", "Qtd Títulos", "Valor Total (R$)", "Saldo Total (R$)"]
    widths4 = [26, 14, 20, 20]
    for ci, (h, w) in enumerate(zip(hdrs4, widths4), 1):
        _hdr(ws4, 1, ci, h)
        ws4.column_dimensions[get_column_letter(ci)].width = w
    for ri, row in enumerate(by_tipo.itertuples(index=False), 2):
        ws4.cell(ri, 1, row.descricao_operacao)
        ws4.cell(ri, 2, int(row.qtd))
        c3 = ws4.cell(ri, 3, float(row.valor_total))
        c3.number_format = BRL
        c4 = ws4.cell(ri, 4, float(row.saldo_total))
        c4.number_format = BRL
        if ri % 2 == 0:
            for ci in range(1, 5):
                ws4.cell(ri, ci).fill = ALT_FILL

    # ── aba 5: Resumo por Categoria ─────────────────────────────────────────
    ws5 = wb.create_sheet("Resumo por Categoria")
    ws5.cell(1, 1, "Análise de Contas a Pagar por Categoria").font = Font(
        bold=True, color="1A5276", size=13, name="Calibri")
    ws5.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(
        italic=True, size=9, color="7F8C8D", name="Calibri")

    by_cat = (
        df.groupby("categoria", as_index=False)
        .agg(qtd=("valor", "count"), valor_total=("valor", "sum"))
        .sort_values("valor_total", ascending=False)
    )
    total_cat = float(by_cat["valor_total"].sum())

    hdr_row5 = 4
    hdrs5 = ["Categoria", "Qtd. Títulos", "Valor Total (R$)", "% do Total", "Valor Médio (R$)"]
    widths5 = [26, 14, 20, 12, 20]
    for ci, (h, w) in enumerate(zip(hdrs5, widths5), 1):
        _hdr(ws5, hdr_row5, ci, h)
        ws5.column_dimensions[get_column_letter(ci)].width = w
    ws5.freeze_panes = f"A{hdr_row5 + 1}"

    ri = hdr_row5 + 1
    for _, r in by_cat.iterrows():
        pct = (r["valor_total"] / total_cat) if total_cat else 0.0
        valor_medio = (r["valor_total"] / r["qtd"]) if r["qtd"] else 0.0
        ws5.cell(ri, 1, r["categoria"])
        ws5.cell(ri, 2, int(r["qtd"]))
        c3 = ws5.cell(ri, 3, float(r["valor_total"])); c3.number_format = BRL
        c4 = ws5.cell(ri, 4, pct); c4.number_format = PCT
        c5 = ws5.cell(ri, 5, float(valor_medio)); c5.number_format = BRL
        if (ri - hdr_row5) % 2 == 0:
            for ci in range(1, 6):
                ws5.cell(ri, ci).fill = ALT_FILL
        ri += 1

    BOLD10 = Font(bold=True, name="Calibri", size=10)
    ws5.cell(ri, 1, "TOTAL GERAL").font = BOLD10
    ws5.cell(ri, 2, int(by_cat["qtd"].sum())).font = BOLD10
    c3 = ws5.cell(ri, 3, total_cat); c3.number_format = BRL; c3.font = BOLD10
    c4 = ws5.cell(ri, 4, 1.0); c4.number_format = PCT; c4.font = BOLD10

    # ── aba 6: Evolução Mensal ──────────────────────────────────────────────
    ws6 = wb.create_sheet("Evolução Mensal")
    ws6.cell(1, 1, "Evolução dos Custos ao Longo do Tempo").font = Font(
        bold=True, color="1A5276", size=13, name="Calibri")
    ws6.cell(2, 1, "Valores por vencimento, agrupados por mês e categoria").font = Font(
        italic=True, size=9, color="7F8C8D", name="Calibri")

    meses = sorted(m for m in df["mes_ano"].unique() if m)
    pivot = (
        df.groupby(["categoria", "mes_ano"])["valor"].sum()
        .unstack(fill_value=0.0)
        .reindex(columns=meses, fill_value=0.0)
    )
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False)

    hdr_row6 = 4
    headers6 = ["Categoria"] + [_fmt_mes(m) for m in meses] + ["Total"]
    for ci, h in enumerate(headers6, 1):
        _hdr(ws6, hdr_row6, ci, h)
    ws6.column_dimensions["A"].width = 26
    for ci in range(2, len(headers6) + 1):
        ws6.column_dimensions[get_column_letter(ci)].width = 14
    ws6.freeze_panes = f"B{hdr_row6 + 1}"

    ri = hdr_row6 + 1
    for categoria, row in pivot.iterrows():
        ws6.cell(ri, 1, categoria)
        for ci, m in enumerate(meses, 2):
            c = ws6.cell(ri, ci, float(row[m]))
            c.number_format = BRL
        ctotal = ws6.cell(ri, len(meses) + 2, float(row["Total"]))
        ctotal.number_format = BRL
        ctotal.font = BOLD10
        ri += 1

    ws6.cell(ri, 1, "TOTAL MENSAL").font = BOLD10
    for ci, m in enumerate(meses, 2):
        c = ws6.cell(ri, ci, float(pivot[m].sum()))
        c.number_format = BRL
        c.font = BOLD10
    ctotal_geral = ws6.cell(ri, len(meses) + 2, float(pivot["Total"].sum()))
    ctotal_geral.number_format = BRL
    ctotal_geral.font = BOLD10

    # ── aba 7: Resumo por Filial ────────────────────────────────────────────
    ws7 = wb.create_sheet("Resumo por Filial")
    ws7.cell(1, 1, "Contas a Pagar por Filial").font = Font(
        bold=True, color="1A5276", size=13, name="Calibri")

    by_filial = (
        df.groupby("filial", as_index=False)
        .agg(qtd=("valor", "count"), valor_total=("valor", "sum"))
        .sort_values("valor_total", ascending=False)
    )
    total_filial = float(by_filial["valor_total"].sum())

    hdr_row7 = 3
    hdrs7 = ["Filial", "Qtd. Títulos", "Valor Total (R$)", "% do Total"]
    widths7 = [18, 14, 20, 12]
    for ci, (h, w) in enumerate(zip(hdrs7, widths7), 1):
        _hdr(ws7, hdr_row7, ci, h)
        ws7.column_dimensions[get_column_letter(ci)].width = w

    ri = hdr_row7 + 1
    for _, r in by_filial.iterrows():
        pct = (r["valor_total"] / total_filial) if total_filial else 0.0
        ws7.cell(ri, 1, r["filial"])
        ws7.cell(ri, 2, int(r["qtd"]))
        c3 = ws7.cell(ri, 3, float(r["valor_total"])); c3.number_format = BRL
        c4 = ws7.cell(ri, 4, pct); c4.number_format = PCT
        if (ri - hdr_row7) % 2 == 0:
            for ci in range(1, 5):
                ws7.cell(ri, ci).fill = ALT_FILL
        ri += 1

    ws7.cell(ri, 1, "TOTAL GERAL").font = BOLD10
    ws7.cell(ri, 2, int(by_filial["qtd"].sum())).font = BOLD10
    c3 = ws7.cell(ri, 3, total_filial); c3.number_format = BRL; c3.font = BOLD10
    c4 = ws7.cell(ri, 4, 1.0); c4.number_format = PCT; c4.font = BOLD10

    wb.save(path)


# ── entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    FinanceiroApp(root)
    root.mainloop()
