"""Categorização de títulos a partir de uma planilha de apoio editável.

A API /financeiro não expõe categoria de título (não existe essa coluna em
nenhuma tabela do Protheus mapeada pelo projeto — SE2070, SA2070, PA6000).
Por isso a categoria é resolvida no cliente, a partir de regras mantidas pelo
próprio usuário em "categorias.xlsx" (aba "Categorias"), lida a cada consulta.

Formato da planilha:
    Fornecedor (contém) | Histórico (contém) | Categoria | Observação

As regras são avaliadas de cima para baixo; a primeira cujo texto (sem
acento, case-insensitive) esteja contido no fornecedor e/ou no histórico do
título vence. Deixe uma das duas colunas de filtro em branco para não
restringir por ela. Títulos sem regra correspondente recebem a categoria
"Não Classificado".
"""

from __future__ import annotations

import unicodedata
from pathlib import Path
from typing import List, NamedTuple

import pandas as pd

CATEGORIAS_PATH = Path(__file__).parent / "categorias.xlsx"
SHEET_NAME = "Categorias"
NAO_CLASSIFICADO = "Não Classificado"

CATEGORIAS_CANONICAS = [
    "Fornecedor", "Imposto", "Jurídico", "Manutenção Automotiva",
    "Prestador de Serviço", "T.I.", "Combustível", "Outros", NAO_CLASSIFICADO,
]

# Mapa fornecedor → categoria mais frequente, extraído do histórico de
# classificação manual em CONTAS_A_PAGAR_ANALISE.xlsx. Usado apenas para
# semear categorias.xlsx na primeira execução — o arquivo gerado é do
# usuário a partir daí e nunca mais é sobrescrito por este módulo.
# O terceiro elemento indica fornecedores com classificação ambígua na
# planilha de origem (mais de uma categoria para o mesmo fornecedor):
# vale revisar essas linhas.
SEED_FORNECEDOR_CATEGORIA = [
    ('58.640.232 KAWANNY E', 'Outros', False),
    ('60.132.893 MILENA SA', 'Manutenção Automotiva', False),
    ('64.073.032IVANILDOMA', 'Outros', False),
    ('ABC HIDRAULICA', 'Manutenção Automotiva', False),
    ('ACOMAIS LTDA', 'Outros', False),
    ('ACTIVE SEGURANCA ELE', 'Fornecedor', False),
    ('ADVOCACIA PIMENTEL', 'Jurídico', False),
    ('AGF CONSULTORIA', 'Prestador de Serviço', True),
    ('AGIBEN BENEFICIOS LT', 'Fornecedor', False),
    ('ALGAR TELECOM S/A', 'T.I.', True),
    ('ALINCA ESQUADRIA', 'Manutenção Automotiva', False),
    ('ALIRIO MORAES', 'Jurídico', False),
    ('ALMIR FRADIQUE DE AL', 'T.I.', False),
    ('APREDILETA CENTER', 'Outros', False),
    ('ARMAZEM CORAL', 'Outros', False),
    ('ARMAZEM SITIO NOVO', 'Fornecedor', False),
    ('ARQUIVEI', 'T.I.', True),
    ('ASSAI ATACADISTA', 'Outros', False),
    ('ASSOCIACAO PERNAMBUC', 'Fornecedor', True),
    ('AURE INFORMATICA', 'Outros', False),
    ('B.SHOP TECNOLOGIA', 'T.I.', False),
    ('BANKTEC COMERCIO LTD', 'Outros', False),
    ('BC COM.DE PECAS E SE', 'Manutenção Automotiva', False),
    ('BELEM AUTO PECAS E E', 'Manutenção Automotiva', True),
    ('BR DIGITAL TELECOM', 'T.I.', True),
    ('BSFRIO', 'Outros', True),
    ('CAIXA PEQUENO', 'Outros', True),
    ('CAR PLUS VEICULOS LT', 'Manutenção Automotiva', False),
    ('CAROATA HOME CENTER', 'Manutenção Automotiva', True),
    ('CASA DO BORRACHEIRO', 'Outros', False),
    ('CELPE', 'Fornecedor', False),
    ('CENTRAL DA CONST', 'Outros', False),
    ('CILO DETETIZACOES', 'Outros', False),
    ('CLARO', 'Fornecedor', False),
    ('CLIN - PLANO ODONT.', 'Outros', True),
    ('COMERCIAL JURIVAN', 'Manutenção Automotiva', True),
    ('COMPESA', 'Fornecedor', False),
    ('CONDOMINIO DO EDF. B', 'Outros', False),
    ('CONEXFER', 'Manutenção Automotiva', False),
    ('CONTROL INFORMATICA', 'T.I.', False),
    ('CREATIVE SISTEMAS', 'T.I.', False),
    ('DETRAN', 'Manutenção Automotiva', False),
    ('DISTRIBUIDORA DE AGU', 'Fornecedor', False),
    ('ECONET', 'T.I.', False),
    ('ECONET SOLUCOES E SE', 'T.I.', True),
    ('ESPERANCA NORDESTE', 'Outros', False),
    ('EXODO LIMPEZA', 'Fornecedor', False),
    ('EXTINNBAT', 'Manutenção Automotiva', False),
    ('FERREIRA COSTA CIA .', 'Manutenção Automotiva', False),
    ('FORTPEL', 'Fornecedor', False),
    ('GLOBAL MOTOS ATACADO', 'Manutenção Automotiva', False),
    ('GRUPO API', 'Outros', False),
    ('HABILITE SERVICOS', 'Outros', False),
    ('HAPVIDA', 'Outros', True),
    ('HC PNEUS', 'Manutenção Automotiva', True),
    ('HEBROM CONSTRUCOES L', 'Outros', False),
    ('INMETRO', 'Manutenção Automotiva', False),
    ('INTERSYS TECNOLOGIA', 'Prestador de Serviço', False),
    ('IPROSTORE', 'Outros', False),
    ('IRON MOUNTAIN BRASIL', 'Fornecedor', False),
    ('IVALDA ALMEIDA DE SO', 'Prestador de Serviço', True),
    ('J J MOLAS', 'Manutenção Automotiva', False),
    ('JATOBARRETTO CENTRO', 'Fornecedor', True),
    ('JC EXTINTORES', 'Outros', False),
    ('L2BR', 'Jurídico', False),
    ('LIG-CONTATO DIARIO F', 'Jurídico', False),
    ('LUIS GUSTAVO SABINO', 'Jurídico', False),
    ('LUIZ ANDREY OLIVEIR', 'Jurídico', False),
    ('MADEIREIRA SAO JOSE', 'Outros', False),
    ('MADIS', 'Outros', False),
    ('MARITEL', 'Outros', False),
    ('MATOS ADVOGADOS', 'Jurídico', False),
    ('MB PARABRISAS', 'Manutenção Automotiva', False),
    ('MEGA DEDETIZACAO', 'Outros', False),
    ('METLIFE PLANOS', 'Fornecedor', False),
    ('MOLAS NORDESTE', 'Manutenção Automotiva', False),
    ('NADJA ISABELLA C. DA', 'Manutenção Automotiva', False),
    ('OI', 'Fornecedor', False),
    ('ONLINE DATA CLOUND', 'T.I.', False),
    ('PARAIBA VIAGENS', 'Outros', False),
    ('PAULO DE PAIVA AZEVE', 'Outros', False),
    ('PLUXEE BENEFICIOS BR', 'Fornecedor', False),
    ('PNEU 1000', 'Manutenção Automotiva', False),
    ('POPULINE TELECOM SER', 'T.I.', False),
    ('POSTO IPIRANGAO', 'Outros', True),
    ('POWER CHIPS', 'Manutenção Automotiva', False),
    ('PPK ASSESSORIA E GES', 'Jurídico', False),
    ('PRESERVE SEG E TRASN', 'Fornecedor', True),
    ('PRODUTIVA', 'Outros', False),
    ('PROTECT', 'Manutenção Automotiva', False),
    ('QUEIROZ & SERENO CON', 'Jurídico', False),
    ('R M CARGO', 'Manutenção Automotiva', False),
    ('RAJ TECNOLOGIA LTDA', 'T.I.', True),
    ('RECUPERA SOLUCOES EM', 'Jurídico', False),
    ('REDIESEL', 'Manutenção Automotiva', True),
    ('RENATO CANUTO NETO S', 'Jurídico', False),
    ('RENOVE', 'Manutenção Automotiva', False),
    ('RM LOC SERVICE', 'Manutenção Automotiva', True),
    ('ROBERTO ANTONIO ALV', 'Fornecedor', False),
    ('ROBSON FERNANDES DA', 'Prestador de Serviço', False),
    ('RODOTEC LTDA', 'Manutenção Automotiva', False),
    ('SAFETEC', 'T.I.', True),
    ('SEC DA FAZENDA PE', 'Imposto', True),
    ('SEGUROS SURA S.A', 'Fornecedor', False),
    ('SIGA MAIS', 'Manutenção Automotiva', False),
    ('SKY.ONE TECNOLOGIA E', 'T.I.', False),
    ('SOCIEDADE AGUIAR LEI', 'Outros', True),
    ('SOLIVETTI', 'Manutenção Automotiva', True),
    ('TAGUS TEC', 'Outros', False),
    ('TARGET ENGENHARIA LT', 'Outros', True),
    ('TERRA LIMPA', 'Outros', True),
    ('THIAGO LUIZ DA SILVA', 'Outros', False),
    ('TICKET LOG', 'Combustível', False),
    ('TOTVS', 'T.I.', False),
    ('TRUCK DIESEL', 'Manutenção Automotiva', False),
    ('TRUCK-CAR', 'Manutenção Automotiva', False),
    ('UNAUDO UNA', 'Manutenção Automotiva', False),
    ('UNIAO', 'Imposto', True),
    ('VALE TRANSPORTE', 'Fornecedor', True),
    ('VALORE CONSULTORES A', 'Prestador de Serviço', False),
    ('VANGAS', 'Manutenção Automotiva', True),
    ('VIPER TECNOLOGIA LTD', 'T.I.', False),
    ('ZAQUEU GAS E AGUA', 'Manutenção Automotiva', True),
]


class Regra(NamedTuple):
    fornecedor: str
    historico: str
    categoria: str


def _normalize(s) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.strip().lower()


def load_rules(path: Path = CATEGORIAS_PATH) -> List[Regra]:
    """Lê categorias.xlsx e devolve as regras na ordem da planilha.

    Se o arquivo não existir, cria-o com a categorização inicial extraída
    de CONTAS_A_PAGAR_ANALISE.xlsx.
    """
    ensure_default_file(path)
    df = pd.read_excel(path, sheet_name=SHEET_NAME, header=3, dtype=str).fillna("")

    regras: List[Regra] = []
    for _, row in df.iterrows():
        categoria = str(row.get("Categoria", "")).strip()
        fornecedor = _normalize(row.get("Fornecedor (contém)", ""))
        historico = _normalize(row.get("Histórico (contém)", ""))
        if not categoria or (not fornecedor and not historico):
            continue
        regras.append(Regra(fornecedor=fornecedor, historico=historico, categoria=categoria))
    return regras


def classify(nome_fornecedor: str, historico: str, regras: List[Regra]) -> str:
    forn_norm = _normalize(nome_fornecedor)
    hist_norm = _normalize(historico)
    for regra in regras:
        if regra.fornecedor and regra.fornecedor not in forn_norm:
            continue
        if regra.historico and regra.historico not in hist_norm:
            continue
        return regra.categoria
    return NAO_CLASSIFICADO


def ensure_default_file(path: Path = CATEGORIAS_PATH) -> None:
    """Cria categorias.xlsx com a categorização inicial, se ainda não existir.

    Nunca sobrescreve um arquivo já existente — a partir da primeira criação
    o arquivo passa a ser mantido pelo usuário.
    """
    if path.exists():
        return

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.cell(1, 1, "Categorização de títulos (planilha de apoio)").font = Font(
        bold=True, color="1A5276", size=13, name="Calibri")
    ws.cell(2, 1, (
        "Cada linha é uma regra: se o fornecedor do título contiver o texto da "
        "coluna B (e/ou o histórico contiver o texto da coluna C, quando preenchida), "
        "o título recebe a categoria da coluna D. Regras são avaliadas de cima para "
        "baixo — a primeira que casar vence. Deixe B ou C em branco para não filtrar "
        "por ela. Títulos sem regra correspondente ficam como \"Não Classificado\". "
        "Edite, adicione ou remova linhas livremente; este arquivo não é sobrescrito "
        "automaticamente."
    )).font = Font(italic=True, size=9, color="7F8C8D", name="Calibri")
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 45
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")

    headers = ["Fornecedor (contém)", "Histórico (contém)", "Categoria", "Observação"]
    hdr_row = 4
    hdr_fill = PatternFill("solid", fgColor="1A5276")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = hdr_font
        c.fill = hdr_fill
    widths = [26, 26, 22, 46]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{hdr_row + 1}"
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(headers))}{hdr_row}"

    for ri, (fornecedor, categoria, ambiguo) in enumerate(
            SEED_FORNECEDOR_CATEGORIA, hdr_row + 1):
        ws.cell(ri, 1, fornecedor)
        ws.cell(ri, 3, categoria)
        if ambiguo:
            ws.cell(ri, 4, (
                "Fornecedor teve mais de uma categoria na planilha de origem — "
                "revise; use a coluna B (Histórico) para diferenciar por lançamento."
            ))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
